#!/usr/bin/env python3
"""Create Appendix Table B9 for the service-destination estimand comparison.

The table reports the validated Heavy-scenario paired comparison between the
implemented any-same-class reachability estimand and the restrictive
fixed-baseline-destination continuity comparator. No simulation is rerun here;
the table consumes the byte-stable audit summary created by
``revision_service_destination_rerouting_audit.py``.
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[2]
SOURCE = (
    ROOT
    / "data/exp/revision/reviewer-2-comment-8/service_destination_estimand_summary.csv"
)
OUT = ROOT / "data/results/tables/Table_service_destination_estimand_comparison.xlsx"
SHEET_NAME = "Estimand Comparison"
TABLE_TITLE = "Service-Destination Estimand and Rerouting Comparison"
SERVICE_ORDER = ["Shelter", "Fire service", "Municipal facility", "Emergency water"]

COLUMNS = [
    "Service Class",
    "Resolved / Total Source Facilities",
    "Road-Attached Facilities",
    "Baseline-Eligible Communities",
    "Baseline-Eligible Population",
    "Any-Same-Class Loss Population Mean",
    "Any-Same-Class Five-Seed Range",
    "Fixed-Destination Loss Population Mean",
    "Fixed-Destination Five-Seed Range",
    "Rerouting Benefit Population Mean",
    "Rerouting Benefit Share of Fixed Loss",
    "Communities with Positive Rerouting Benefit",
]


def _range_text(low: float, high: float) -> str:
    return f"{low:,.1f}-{high:,.1f}"


def build_table() -> pd.DataFrame:
    """Load the validated audit summary and return four reader-facing rows."""
    source = pd.read_csv(SOURCE)
    if set(source["Service Class"]) != set(SERVICE_ORDER) or len(source) != 4:
        raise RuntimeError("The validated service-estimand summary must contain four classes.")
    source = source.set_index("Service Class").loc[SERVICE_ORDER].reset_index()
    table = pd.DataFrame(
        {
            "Service Class": source["Service Class"].replace(
                {"Emergency water": "Emergency water (conditional)"}
            ),
            "Resolved / Total Source Facilities": (
                source["Resolved Source Facilities"].astype(int).astype(str)
                + " / "
                + source["Source Facilities Total"].astype(int).astype(str)
            ),
            "Road-Attached Facilities": source["Road-Attached Facilities"].astype(int),
            "Baseline-Eligible Communities": source[
                "Baseline-Eligible Communities"
            ].astype(int),
            "Baseline-Eligible Population": source[
                "Baseline-Eligible Population"
            ].astype(float),
            "Any-Same-Class Loss Population Mean": source[
                "Any-Same-Class Loss Population Mean"
            ].astype(float),
            "Any-Same-Class Five-Seed Range": [
                _range_text(low, high)
                for low, high in zip(
                    source["Any-Same-Class Loss Population Min"],
                    source["Any-Same-Class Loss Population Max"],
                )
            ],
            "Fixed-Destination Loss Population Mean": source[
                "Fixed-Destination Loss Population Mean"
            ].astype(float),
            "Fixed-Destination Five-Seed Range": [
                _range_text(low, high)
                for low, high in zip(
                    source["Fixed-Destination Loss Population Min"],
                    source["Fixed-Destination Loss Population Max"],
                )
            ],
            "Rerouting Benefit Population Mean": source[
                "Rerouting Benefit Population Mean"
            ].astype(float),
            "Rerouting Benefit Share of Fixed Loss": source[
                "Rerouting Benefit Share of Fixed Loss"
            ].astype(float),
            "Communities with Positive Rerouting Benefit": source[
                "Communities with Positive Rerouting Benefit"
            ].astype(int),
        }
    )
    return table.loc[:, COLUMNS]


def style_workbook(path: Path) -> None:
    """Apply the established project table style and one-page print settings."""
    workbook = load_workbook(path)
    worksheet = workbook[SHEET_NAME]
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A3"
    worksheet.sheet_view.zoomScale = 80
    worksheet.print_area = "A1:L8"
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A3
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 1
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_margins = PageMargins(
        left=0.25, right=0.25, top=0.35, bottom=0.35, header=0.10, footer=0.10
    )

    worksheet.merge_cells("A1:L1")
    title = worksheet["A1"]
    title.value = TABLE_TITLE
    title.fill = PatternFill("solid", fgColor="D9EAF7")
    title.font = Font(name="Aptos Display", size=15, bold=True, color="17365D")
    title.alignment = Alignment(horizontal="left", vertical="center")
    worksheet.row_dimensions[1].height = 28

    header_fill = PatternFill("solid", fgColor="17365D")
    header_font = Font(name="Aptos", size=9, bold=True, color="FFFFFF")
    body_font = Font(name="Aptos", size=9, color="172033")
    border = Border(bottom=Side(style="thin", color="D0D5DD"))
    for cell in worksheet[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[2].height = 64

    for row_number in range(3, 7):
        for column in range(1, 13):
            cell = worksheet.cell(row_number, column)
            cell.font = body_font
            cell.border = border
            cell.alignment = Alignment(
                horizontal="left" if column == 1 else "right",
                vertical="center",
                wrap_text=True,
            )
        if row_number % 2 == 0:
            for cell in worksheet[row_number]:
                cell.fill = PatternFill("solid", fgColor="F5F8FC")
        worksheet.row_dimensions[row_number].height = 30
    for row_number in range(3, 7):
        for column in (3, 4, 12):
            worksheet.cell(row_number, column).number_format = "#,##0"
        for column in (5, 6, 8, 10):
            worksheet.cell(row_number, column).number_format = "#,##0.0"
        worksheet.cell(row_number, 11).number_format = "0.0%"

    worksheet.merge_cells("A8:L8")
    note = worksheet["A8"]
    note.value = (
        "Note: Heavy-scenario means and ranges use five predeclared seeds with 1,000 paired "
        "closure draws per seed. Emergency-water estimates are conditional on the 10 of 36 "
        "announced destinations with resolved coordinates."
    )
    note.font = Font(name="Aptos", size=8.5, italic=True, color="475467")
    note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    worksheet.row_dimensions[8].height = 34

    widths = {
        "A": 24,
        "B": 18,
        "C": 16,
        "D": 18,
        "E": 17,
        "F": 21,
        "G": 18,
        "H": 21,
        "I": 18,
        "J": 20,
        "K": 18,
        "L": 19,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    excel_table = Table(displayName="ServiceDestinationEstimands", ref="A2:L6")
    excel_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(excel_table)
    workbook.save(path)


def verify_workbook(path: Path, expected: pd.DataFrame) -> None:
    """Verify workbook structure, typed values, and audit-summary reconciliation."""
    workbook = load_workbook(path, data_only=False)
    if workbook.sheetnames != [SHEET_NAME]:
        raise RuntimeError(f"Unexpected workbook sheets: {workbook.sheetnames}")
    worksheet = workbook[SHEET_NAME]
    if worksheet.max_column != 12 or worksheet.max_row != 8:
        raise RuntimeError(
            f"Expected 8 rows and 12 columns; found {worksheet.max_row} x "
            f"{worksheet.max_column}."
        )
    if worksheet["A1"].value != TABLE_TITLE:
        raise RuntimeError("Workbook title row is missing or incorrect.")
    headers = [worksheet.cell(2, column).value for column in range(1, 13)]
    if headers != COLUMNS:
        raise RuntimeError("Workbook column headers do not match the declared schema.")
    error_tokens = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NUM!"}
    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value in error_tokens:
                raise RuntimeError(f"Spreadsheet error token in {cell.coordinate}: {cell.value}")
    for row_offset, record in enumerate(expected.to_dict("records"), start=3):
        for column_offset, header in enumerate(COLUMNS, start=1):
            observed = worksheet.cell(row_offset, column_offset).value
            target = record[header]
            if isinstance(target, float):
                if abs(float(observed) - target) > 1e-8:
                    raise RuntimeError(f"Numeric mismatch at {worksheet.cell(row_offset, column_offset).coordinate}.")
            elif observed != target:
                raise RuntimeError(f"Value mismatch at {worksheet.cell(row_offset, column_offset).coordinate}.")


def main() -> None:
    table = build_table()
    OUT.parent.mkdir(parents=True, exist_ok=True)
    table.to_excel(
        OUT,
        index=False,
        sheet_name=SHEET_NAME,
        engine="openpyxl",
        startrow=1,
    )
    style_workbook(OUT)
    verify_workbook(OUT, table)
    print(f"Saved: {OUT.relative_to(ROOT)}")
    print(f"Rows: {len(table):,}; columns: {len(table.columns):,}")
    print("Workbook verification: passed")


if __name__ == "__main__":
    main()
