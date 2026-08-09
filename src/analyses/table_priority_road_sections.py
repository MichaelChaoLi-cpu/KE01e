#!/usr/bin/env python3
"""Priority Road Sections.

Plan: List the 30 road sections with the largest consequence-aware intervention
relevance under the accepted Heavy-rainfall screening scenario.
Framework: AnaSOP Sections 5-7 use relative road disruption scores, the central
score-to-closure mapping, single-section consequence checks, and robust median
benefit-per-planning-cost ranking. Results are screening priorities rather than
engineering recommendations or calibrated closure probabilities.
"""
from __future__ import annotations

import os
from pathlib import Path

os.environ.pop("PROJ_LIB", None)
os.environ.pop("PROJ_DATA", None)

import numpy as np
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd
from PIL import Image, ImageDraw, ImageFont
from rasterio.transform import from_bounds
from scipy.sparse import coo_matrix
from scipy.sparse.csgraph import connected_components
import shapely

import figure_basic_service_reachability_loss as service_loss
import figure_community_isolation_frequency_and_exposed_population as isolation
import figure_intervention_priorities_and_budgeted_benefits as intervention
import figure_road_disruption_exposure_and_observed_restriction_evidence as road_exposure


ROOT = Path(__file__).resolve().parents[2]
PROCESSED = ROOT / "data/processed"
ADMIN_PATH = PROCESSED / "administrative_areas_preprocessed.parquet"
ROAD_PATH = PROCESSED / "road_sections_preprocessed.parquet"
EDGE_PATH = PROCESSED / "road_edges_preprocessed.parquet"
NODE_PATH = PROCESSED / "road_nodes_preprocessed.parquet"
OUT = ROOT / "data/results/tables/Table_priority_road_sections.xlsx"
PREVIEW_OUT = ROOT / "data/exp/table_previews/Table_priority_road_sections.png"
SHEET_NAME = "Priority Roads"
TABLE_TITLE = "Priority Road Sections"
TOP_ROADS = 30
MUNICIPALITY_ENGLISH_BY_CODE = {
    "43101": "Kumamoto City, Chuo Ward",
    "43102": "Kumamoto City, Higashi Ward",
    "43103": "Kumamoto City, Nishi Ward",
    "43104": "Kumamoto City, Minami Ward",
    "43105": "Kumamoto City, Kita Ward",
    "43202": "Yatsushiro City",
    "43203": "Hitoyoshi City",
    "43204": "Arao City",
    "43205": "Minamata City",
    "43206": "Tamana City",
    "43208": "Yamaga City",
    "43210": "Kikuchi City",
    "43211": "Uto City",
    "43212": "Kamiamakusa City",
    "43213": "Uki City",
    "43214": "Aso City",
    "43215": "Amakusa City",
    "43216": "Koushi City",
    "43348": "Misato Town",
    "43364": "Gyokuto Town",
    "43367": "Nankan Town",
    "43368": "Nagasu Town",
    "43369": "Nagomi Town",
    "43403": "Ozu Town",
    "43404": "Kikuyo Town",
    "43423": "Minamioguni Town",
    "43424": "Oguni Town",
    "43425": "Ubuyama Village",
    "43428": "Takamori Town",
    "43432": "Nishihara Village",
    "43433": "Minamiaso Village",
    "43441": "Mifune Town",
    "43442": "Kashima Town",
    "43443": "Mashiki Town",
    "43444": "Kosa Town",
    "43447": "Yamato Town",
    "43468": "Hikawa Town",
    "43482": "Ashikita Town",
    "43484": "Tsunagi Town",
    "43501": "Nishiki Town",
    "43505": "Taragi Town",
    "43506": "Yunomae Town",
    "43507": "Mizukami Village",
    "43510": "Sagara Village",
    "43511": "Itsuki Village",
    "43512": "Yamae Village",
    "43513": "Kuma Village",
    "43514": "Asagiri Town",
    "43531": "Reihoku Town",
}


def _preview_font(size: int, bold: bool = False) -> ImageFont.FreeTypeFont:
    """Return a readable system font for the long PNG review copy."""
    candidates = (
        "/System/Library/Fonts/Hiragino Sans GB.ttc",
        "/System/Library/Fonts/Supplemental/Arial Bold.ttf"
        if bold
        else "/System/Library/Fonts/Supplemental/Arial.ttf",
        "/System/Library/Fonts/Supplemental/Arial.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"
        if bold
        else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    )
    for candidate in candidates:
        if Path(candidate).exists():
            return ImageFont.truetype(candidate, size=size)
    return ImageFont.load_default()


def _wrap_preview_text(
    draw: ImageDraw.ImageDraw,
    value: object,
    font: ImageFont.ImageFont,
    max_width: int,
) -> str:
    """Wrap a cell value to its pixel width without truncating words."""
    words = str(value).replace("; ", ";\n").split()
    lines: list[str] = []
    current = ""
    for word in words:
        candidate = word if not current else f"{current} {word}"
        if draw.textbbox((0, 0), candidate, font=font)[2] <= max_width:
            current = candidate
        else:
            if current:
                lines.append(current)
            current = word
    if current:
        lines.append(current)
    return "\n".join(lines)


def render_preview(path: Path, output: Path = PREVIEW_OUT) -> None:
    """Render the complete 30-row workbook as one continuous review PNG."""
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook[SHEET_NAME]
    values = list(worksheet.values)
    title = str(values[0][0])
    headers = [str(value) for value in values[1]]
    rows = [list(row) for row in values[2:]]

    widths = [100, 210, 240, 150, 180, 175, 170, 225, 360, 260, 160, 175]
    margin = 24
    title_height = 88
    header_height = 104
    row_height = 66
    table_width = sum(widths)
    image = Image.new(
        "RGB",
        (table_width + 2 * margin, title_height + header_height + row_height * len(rows) + margin),
        "white",
    )
    draw = ImageDraw.Draw(image)
    title_font = _preview_font(28, bold=True)
    note_font = _preview_font(16)
    header_font = _preview_font(17, bold=True)
    body_font = _preview_font(16)
    body_bold = _preview_font(16, bold=True)
    draw.text((margin, 18), title, font=title_font, fill="#17365D")
    draw.text(
        (margin, 54),
        "Top 30 consequence-aware screening priorities under the Heavy rainfall scenario",
        font=note_font,
        fill="#52606D",
    )

    x_positions = [margin]
    for width in widths:
        x_positions.append(x_positions[-1] + width)
    header_top = title_height
    for column, header in enumerate(headers):
        x0, x1 = x_positions[column], x_positions[column + 1]
        draw.rectangle((x0, header_top, x1, header_top + header_height), fill="#17365D")
        wrapped = _wrap_preview_text(draw, header, header_font, widths[column] - 16)
        bbox = draw.multiline_textbbox((0, 0), wrapped, font=header_font, spacing=3, align="center")
        text_height = bbox[3] - bbox[1]
        draw.multiline_text(
            ((x0 + x1) / 2, header_top + (header_height - text_height) / 2),
            wrapped,
            font=header_font,
            fill="white",
            spacing=3,
            anchor="ma",
            align="center",
        )

    numeric_right = {0, 3, 4, 5, 6, 10, 11}
    formats = {
        3: lambda value: f"{float(value):.3f}",
        4: lambda value: f"{float(value):.1%}",
        5: lambda value: f"{float(value):.3f}",
        6: lambda value: f"{int(round(float(value))):,}",
        10: lambda value: f"{float(value):,.2f}",
        11: lambda value: f"{float(value):,.2f}",
    }
    conditional_columns = {4, 5, 11}
    column_ranges = {
        column: (
            min(float(row[column]) for row in rows),
            max(float(row[column]) for row in rows),
        )
        for column in conditional_columns
    }

    def scale_colour(value: float, low: float, high: float) -> str:
        ratio = 0.5 if high == low else (value - low) / (high - low)
        if ratio <= 0.5:
            local = ratio / 0.5
            start, end = (255, 255, 255), (255, 235, 132)
        else:
            local = (ratio - 0.5) / 0.5
            start, end = (255, 235, 132), (248, 105, 107)
        rgb = tuple(round(a + (b - a) * local) for a, b in zip(start, end))
        return "#" + "".join(f"{channel:02X}" for channel in rgb)

    for row_number, row in enumerate(rows):
        y0 = header_top + header_height + row_number * row_height
        y1 = y0 + row_height
        base_fill = "#FFFFFF" if row_number % 2 == 0 else "#F8FAFC"
        rank = int(row[0])
        rank_fill = "#F4CCCC" if rank <= 10 else "#FCE5CD" if rank <= 30 else "#FFF2CC"
        for column, value in enumerate(row):
            x0, x1 = x_positions[column], x_positions[column + 1]
            fill = rank_fill if column == 0 else base_fill
            if column in conditional_columns:
                low, high = column_ranges[column]
                fill = scale_colour(float(value), low, high)
            draw.rectangle((x0, y0, x1, y1), fill=fill, outline="#D0D5DD", width=1)
            display_value = "" if value is None else formats.get(column, str)(value)
            wrapped = _wrap_preview_text(draw, display_value, body_font, widths[column] - 14)
            font = body_bold if column in {0, 11} else body_font
            if column in numeric_right:
                draw.multiline_text(
                    (x1 - 7, y0 + 8),
                    wrapped,
                    font=font,
                    fill="#172033",
                    spacing=2,
                    anchor="ra",
                    align="right",
                )
            else:
                draw.multiline_text(
                    (x0 + 7, y0 + 8),
                    wrapped,
                    font=font,
                    fill="#172033",
                    spacing=2,
                )

    output.parent.mkdir(parents=True, exist_ok=True)
    image.save(output, optimize=True)


def component_community_access(
    labels: np.ndarray,
    component_roots: np.ndarray,
    attachment_community: np.ndarray,
    attachment_root: np.ndarray,
    community_count: int,
) -> np.ndarray:
    """Return community access to any root in the declared root set."""
    component_has_target = np.zeros(int(labels.max()) + 1, dtype=bool)
    component_has_target[labels[component_roots]] = True
    root_accessible = component_has_target[labels]
    community_accessible = np.zeros(community_count, dtype="uint8")
    np.maximum.at(
        community_accessible,
        attachment_community,
        root_accessible[attachment_root].astype("uint8"),
    )
    return community_accessible.astype(bool)


def graph_labels(
    candidate_u: np.ndarray,
    candidate_v: np.ndarray,
    candidate_edge_section: np.ndarray,
    root_count: int,
    closed_section: int | None = None,
) -> np.ndarray:
    """Return connected-component labels with all or one candidate section open."""
    if closed_section is None:
        edge_open = np.ones(len(candidate_edge_section), dtype=bool)
    else:
        edge_open = candidate_edge_section != closed_section
    u = candidate_u[edge_open]
    v = candidate_v[edge_open]
    graph = coo_matrix(
        (
            np.ones(len(u) * 2, dtype="uint8"),
            (np.concatenate([u, v]), np.concatenate([v, u])),
        ),
        shape=(root_count, root_count),
    ).tocsr()
    _, labels = connected_components(graph, directed=False, return_labels=True)
    return labels.astype("int32")


def build_table() -> tuple[pd.DataFrame, dict[str, object]]:
    """Reproduce the accepted intervention ranking and detail its top 30 roads."""
    admin = pd.read_parquet(
        ADMIN_PATH,
        columns=["Municipality Code", "Municipality Name", "Geometry"],
    )
    admin_geometry = road_exposure.decode_geometry(admin.pop("Geometry"))
    admin_union = shapely.union_all(admin_geometry)
    min_x, min_y, max_x, max_y = shapely.bounds(admin_union)
    pad_x = (max_x - min_x) * 0.025
    pad_y = (max_y - min_y) * 0.025
    extent = (min_x - pad_x, max_x + pad_x, min_y - pad_y, max_y + pad_y)
    west, east, south, north = extent
    display_height = max(
        650,
        round(intervention.DISPLAY_WIDTH * (north - south) / (east - west)),
    )
    display_shape = (display_height, intervention.DISPLAY_WIDTH)
    display_transform = from_bounds(
        west,
        south,
        east,
        north,
        intervention.DISPLAY_WIDTH,
        display_height,
    )

    terrain_scores, _, model_mode, elevation_grid = road_exposure.build_landslide_scores(
        admin,
        admin_geometry,
        admin_union,
        extent,
        display_shape,
        display_transform,
    )
    roads = pd.read_parquet(
        ROAD_PATH,
        columns=[
            "Road Section ID",
            "Road Section Length (m)",
            "Road Category",
            "Emergency Route Membership",
            "Network Analysis Eligible",
            "Geometry",
        ],
    )
    roads = roads.loc[roads["Network Analysis Eligible"]].reset_index(drop=True)
    road_geometry = road_exposure.decode_geometry(roads.pop("Geometry"))
    road_scores = road_exposure.road_scores(road_geometry, terrain_scores, extent, elevation_grid)
    heavy_lower = isolation.positive_score_quantile(
        road_scores["Heavy"], isolation.CANDIDATE_QUANTILE
    )
    heavy_upper = isolation.positive_score_quantile(
        road_scores["Heavy"], isolation.UPPER_MAPPING_QUANTILE
    )
    candidate = np.isfinite(road_scores["Heavy"]) & (road_scores["Heavy"] >= heavy_lower)
    candidate_ids = roads.loc[candidate, "Road Section ID"].reset_index(drop=True)
    candidate_position = pd.Series(
        np.arange(len(candidate_ids), dtype="int32"),
        index=candidate_ids,
    )
    candidate_road_index = roads.index[candidate].to_numpy(dtype="int32")
    candidate_score = road_scores["Heavy"][candidate]
    section_propensity = isolation.closure_propensity(
        candidate_score,
        heavy_lower,
        heavy_upper,
    )

    nodes = pd.read_parquet(
        NODE_PATH,
        columns=["Network Node ID", "Network Component ID", "Geometry"],
    )
    node_geometry = road_exposure.decode_geometry(nodes.pop("Geometry"))
    node_index = pd.Index(nodes["Network Node ID"])
    edges = pd.read_parquet(
        EDGE_PATH,
        columns=[
            "Road Section ID",
            "From Node ID",
            "To Node ID",
            "Network Component ID",
            "Emergency Route Membership",
            "Network Analysis Eligible",
        ],
    )
    edges = edges.loc[edges["Network Analysis Eligible"]].reset_index(drop=True)
    edge_u = node_index.get_indexer(edges["From Node ID"])
    edge_v = node_index.get_indexer(edges["To Node ID"])
    if np.any(edge_u < 0) or np.any(edge_v < 0):
        raise RuntimeError("Road edges reference missing network nodes.")
    edge_candidate = edges["Road Section ID"].isin(candidate_ids).to_numpy()

    stable_u = edge_u[~edge_candidate]
    stable_v = edge_v[~edge_candidate]
    stable_graph = coo_matrix(
        (
            np.ones(len(stable_u) * 2, dtype="uint8"),
            (
                np.concatenate([stable_u, stable_v]),
                np.concatenate([stable_v, stable_u]),
            ),
        ),
        shape=(len(nodes), len(nodes)),
    ).tocsr()
    root_count, stable_labels = connected_components(
        stable_graph,
        directed=False,
        return_labels=True,
    )
    stable_labels = stable_labels.astype("int32")

    candidate_u = stable_labels[edge_u[edge_candidate]]
    candidate_v = stable_labels[edge_v[edge_candidate]]
    candidate_edge_section = (
        edges.loc[edge_candidate, "Road Section ID"]
        .map(candidate_position)
        .to_numpy(dtype="int32")
    )
    between_root = candidate_u != candidate_v
    candidate_u = candidate_u[between_root]
    candidate_v = candidate_v[between_root]
    candidate_edge_section = candidate_edge_section[between_root]

    target_definitions, target_network_components = isolation.external_target_definitions(
        nodes,
        node_geometry,
        stable_labels,
        edges,
        edge_u,
        edge_v,
        admin_union,
    )
    target_roots = target_definitions["Primary boundary gateways"]
    (
        community,
        attachment_community,
        attachment_root,
        community_diagnostics,
        _,
        _,
    ) = isolation.build_baseline_communities(
        nodes,
        node_geometry,
        stable_labels,
        target_network_components,
    )
    community_population = community["Total_Population"].to_numpy(dtype=float)
    community_older = community["Population_Age_65"].to_numpy(dtype=float)
    print("Simulating Heavy baseline for priority-road screening")
    baseline_frequency = isolation.simulate_isolation(
        candidate_u,
        candidate_v,
        candidate_edge_section,
        section_propensity,
        root_count,
        target_roots,
        attachment_community,
        attachment_root,
        len(community),
        isolation.RANDOM_SEED + 10_000,
    )

    attachment_count = np.bincount(
        attachment_community,
        minlength=len(community),
    ).astype(float)
    attachment_share_burden = (
        community_population * baseline_frequency / np.maximum(attachment_count, 1.0)
    )
    root_burden = np.zeros(root_count, dtype="float64")
    np.add.at(
        root_burden,
        attachment_root,
        attachment_share_burden[attachment_community],
    )
    root_degree = np.bincount(
        np.concatenate([candidate_u, candidate_v]),
        minlength=root_count,
    ).astype(float)
    section_burden = np.zeros(len(candidate_ids), dtype="float64")
    section_scarcity = np.zeros(len(candidate_ids), dtype="float64")
    edge_burden = root_burden[candidate_u] + root_burden[candidate_v]
    edge_scarcity = 1.0 / np.sqrt(
        np.maximum(np.minimum(root_degree[candidate_u], root_degree[candidate_v]), 1.0)
    )
    np.maximum.at(section_burden, candidate_edge_section, edge_burden)
    np.maximum.at(section_scarcity, candidate_edge_section, edge_scarcity)
    emergency_candidate = (
        roads.loc[candidate_road_index, "Emergency Route Membership"]
        .astype("string")
        .ne("None")
        .to_numpy()
    )
    preliminary_score = (
        candidate_score
        * np.log1p(section_burden)
        * (1.0 + section_scarcity)
        * np.where(emergency_candidate, 1.20, 1.0)
    )

    screen_positions = np.argsort(preliminary_score)[-intervention.SINGLE_CLOSE_SCREEN_COUNT :]
    single_close_population = np.zeros(len(candidate_ids), dtype="float64")
    for count, position in enumerate(screen_positions, start=1):
        single_close_population[position] = intervention.single_section_closed_population(
            int(position),
            candidate_u,
            candidate_v,
            candidate_edge_section,
            root_count,
            target_roots,
            attachment_community,
            attachment_root,
            community_population,
        )
        if count % 250 == 0:
            print(
                f"  completed {count:,}/{len(screen_positions):,} "
                "single-road consequence checks"
            )

    consequence_proxy = single_close_population + 0.15 * section_burden
    candidate_length_km = (
        roads.loc[candidate_road_index, "Road Section Length (m)"].to_numpy(dtype=float)
        / 1000.0
    )
    actions = intervention.action_assignment(
        roads.loc[candidate_road_index, "Emergency Route Membership"],
        candidate_score,
        heavy_upper,
        section_scarcity,
    )
    base_cost = np.select(
        [
            actions == "Temporary reinforcement",
            actions == "Alternative-route protection",
        ],
        [3.0 + 2.0 * candidate_length_km, 2.5 + 1.2 * candidate_length_km],
        default=1.5 + 0.5 * candidate_length_km,
    ).astype("float64")

    sensitivity_scores: list[np.ndarray] = []
    for setting in ("Conservative", "Central", "Optimistic"):
        effect = np.array(
            [intervention.ACTION_EFFECT[action][setting] for action in actions]
        )
        cost = base_cost * intervention.COST_MULTIPLIER[setting]
        sensitivity_scores.append(
            consequence_proxy * effect / np.maximum(cost, 1e-6)
        )
    priority_score = np.median(np.vstack(sensitivity_scores), axis=0)
    priority_order = np.argsort(priority_score)[::-1]
    top_positions = priority_order[:TOP_ROADS]

    service_geometry, _ = service_loss.service_geometries()
    service_roots, _, _ = service_loss.attach_services_to_roots(
        service_geometry,
        node_geometry,
        stable_labels,
    )
    baseline_labels = graph_labels(
        candidate_u,
        candidate_v,
        candidate_edge_section,
        root_count,
    )
    baseline_service_access = {
        name: component_community_access(
            baseline_labels,
            service_roots[name],
            attachment_community,
            attachment_root,
            len(community),
        )
        for name in service_loss.SERVICE_CLASSES
    }

    rows: list[dict[str, object]] = []
    for rank, position in enumerate(top_positions, start=1):
        labels = graph_labels(
            candidate_u,
            candidate_v,
            candidate_edge_section,
            root_count,
            closed_section=int(position),
        )
        external_access = component_community_access(
            labels,
            target_roots,
            attachment_community,
            attachment_root,
            len(community),
        )
        isolated_mask = ~external_access
        dependent_services: list[str] = []
        for name in service_loss.SERVICE_CLASSES:
            access = component_community_access(
                labels,
                service_roots[name],
                attachment_community,
                attachment_root,
                len(community),
            )
            if np.any(baseline_service_access[name] & ~access):
                dependent_services.append(name)
        road_index = int(candidate_road_index[position])
        road_geometry_value = road_geometry[road_index]
        municipality_matches = np.flatnonzero(
            shapely.intersects(admin_geometry, road_geometry_value)
        )
        if municipality_matches.size:
            overlap_lengths = shapely.length(
                shapely.intersection(
                    admin_geometry[municipality_matches],
                    road_geometry_value,
                )
            )
            municipality_index = int(
                municipality_matches[int(np.argmax(overlap_lengths))]
            )
            municipality_code = str(
                admin.iloc[municipality_index]["Municipality Code"]
            )
        else:
            road_midpoint = shapely.line_interpolate_point(
                road_geometry_value,
                0.5,
                normalized=True,
            )
            municipality_index = int(
                np.argmin(shapely.distance(admin_geometry, road_midpoint))
            )
            municipality_code = str(
                admin.iloc[municipality_index]["Municipality Code"]
            )
        municipality_name = MUNICIPALITY_ENGLISH_BY_CODE.get(municipality_code)
        if municipality_name is None:
            raise RuntimeError(
                f"Missing English municipality name for code {municipality_code}."
            )
        action = str(actions[position])
        dependent_total = float(community_population[isolated_mask].sum())
        dependent_older = float(community_older[isolated_mask].sum())
        rows.append(
            {
                "Priority Rank": rank,
                "Municipality / Ward": municipality_name,
                "Road Category": str(roads.at[road_index, "Road Category"]),
                "Road Section Length (km)": float(candidate_length_km[position]),
                "Heavy Road Disruption Score": float(candidate_score[position]),
                "Alternative Scarcity Score": float(section_scarcity[position]),
                "Affected Community Count": int(isolated_mask.sum()),
                "Dependent Population (Total / Age 65+)": (
                    f"{dependent_total:,.0f} / {dependent_older:,.0f}"
                ),
                "Dependent Service Classes": (
                    "; ".join(dependent_services) if dependent_services else "None detected"
                ),
                "Assigned Intervention Type": action,
                "Central Planning Cost": float(base_cost[position]),
                "Robust Priority Score": float(priority_score[position]),
            }
        )

    table = pd.DataFrame(rows)
    if table.shape != (TOP_ROADS, 12):
        raise RuntimeError(f"Expected a {TOP_ROADS} × 12 table, found {table.shape}.")
    selected_road_ids = candidate_ids.iloc[top_positions].astype(str)
    if selected_road_ids.duplicated().any():
        raise RuntimeError("Priority road-section identifiers must be unique.")
    if not table["Priority Rank"].equals(
        pd.Series(range(1, TOP_ROADS + 1), name="Priority Rank")
    ):
        raise RuntimeError(f"Priority ranks are not sequential from 1 to {TOP_ROADS}.")
    diagnostics = {
        "Eligible Population": community_diagnostics["Eligible Population"],
        "Candidate Roads": len(candidate_ids),
        "Model Mode": model_mode,
        "Heavy Lower": heavy_lower,
        "Heavy Upper": heavy_upper,
        "Priority Order": priority_order,
        "Base Cost": base_cost,
        "Actions": actions,
        "Section Propensity": section_propensity,
        "Candidate U": candidate_u,
        "Candidate V": candidate_v,
        "Candidate Edge Section": candidate_edge_section,
        "Root Count": root_count,
        "Target Roots": target_roots,
        "Attachment Community": attachment_community,
        "Attachment Root": attachment_root,
        "Community Population": community_population,
        "Community Population Age 65+": community_older,
        "Baseline Frequency": baseline_frequency,
        "Candidate Score": candidate_score,
        "Emergency Candidate": emergency_candidate,
        "Candidate Road Category": roads.loc[
            candidate_road_index, "Road Category"
        ].reset_index(drop=True),
        "Consequence Proxy": consequence_proxy,
    }
    return table, diagnostics


def style_workbook(path: Path) -> None:
    """Apply compact grouped formatting to the top-30 screening table."""
    workbook = load_workbook(path)
    worksheet = workbook[SHEET_NAME]
    worksheet.insert_rows(1)
    worksheet.merge_cells("A1:L1")
    worksheet["A1"] = TABLE_TITLE
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "C3"
    worksheet.auto_filter.ref = f"A2:L{worksheet.max_row}"
    worksheet.sheet_view.zoomScale = 90
    worksheet.print_area = f"A1:L{worksheet.max_row}"
    worksheet.print_title_rows = "1:2"
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A3
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_margins = PageMargins(
        left=0.20, right=0.20, top=0.30, bottom=0.30, header=0.10, footer=0.10
    )

    header_fill = PatternFill("solid", fgColor="17365D")
    header_font = Font(name="Aptos", size=9, bold=True, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor="D9EAF7")
    title_font = Font(name="Aptos Display", size=14, bold=True, color="17365D")
    body_font = Font(name="Aptos", size=8.8, color="172033")
    subtle_border = Border(bottom=Side(style="thin", color="D0D5DD"))
    rank_fills = {
        "top": PatternFill("solid", fgColor="F4CCCC"),
        "middle": PatternFill("solid", fgColor="FCE5CD"),
        "lower": PatternFill("solid", fgColor="FFF2CC"),
    }
    worksheet["A1"].fill = title_fill
    worksheet["A1"].font = title_font
    worksheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    worksheet.row_dimensions[1].height = 30
    for cell in worksheet[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[2].height = 48

    for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = subtle_border
        rank = int(row[0].value)
        row[0].fill = (
            rank_fills["top"]
            if rank <= 10
            else rank_fills["middle"]
            if rank <= 30
            else rank_fills["lower"]
        )
        row[3].number_format = "0.000"
        row[4].number_format = "0.0%"
        row[5].number_format = "0.000"
        row[6].number_format = "#,##0"
        for column in (11, 12):
            row[column - 1].number_format = "#,##0.00"
        for cell in (*row[:1], *row[3:7], *row[10:]):
            cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
        worksheet.row_dimensions[row[0].row].height = 28

    widths = {
        "A": 12,
        "B": 21,
        "C": 18,
        "D": 19,
        "E": 21,
        "F": 20,
        "G": 18,
        "H": 28,
        "I": 30,
        "J": 26,
        "K": 18,
        "L": 20,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width
    for column in ("E", "F", "L"):
        worksheet.conditional_formatting.add(
            f"{column}3:{column}{worksheet.max_row}",
            ColorScaleRule(
                start_type="min",
                start_color="FFFFFF",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="F8696B",
            ),
        )

    excel_table = Table(displayName="PriorityRoadSections", ref=f"A2:L{worksheet.max_row}")
    excel_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(excel_table)
    workbook.save(path)


def verify_workbook(path: Path) -> None:
    """Verify dimensions, ranks, ranges, and absence of spreadsheet errors."""
    workbook = load_workbook(path, data_only=False)
    if workbook.sheetnames != [SHEET_NAME]:
        raise RuntimeError(f"Unexpected workbook sheets: {workbook.sheetnames}")
    worksheet = workbook[SHEET_NAME]
    if worksheet.max_row != TOP_ROADS + 2 or worksheet.max_column != 12:
        raise RuntimeError(
            f"Expected {TOP_ROADS + 2} rows including title and header and 12 columns; found "
            f"{worksheet.max_row} × {worksheet.max_column}."
        )
    if worksheet["A1"].value != TABLE_TITLE:
        raise RuntimeError("Workbook title row is missing or incorrect.")
    ranks = [worksheet.cell(row, 1).value for row in range(3, TOP_ROADS + 3)]
    if ranks != list(range(1, TOP_ROADS + 1)):
        raise RuntimeError(f"Workbook priority ranks are not 1 through {TOP_ROADS}.")
    error_tokens = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"}
    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value in error_tokens:
                raise RuntimeError(f"Spreadsheet error token in {cell.coordinate}: {cell.value}")
    for row in range(3, TOP_ROADS + 3):
        for column in (5, 6):
            value = float(worksheet.cell(row, column).value)
            if not 0 <= value <= 1:
                raise RuntimeError(f"Score outside [0, 1] at row {row}, column {column}.")


def main() -> None:
    table, diagnostics = build_table()
    OUT.parent.mkdir(parents=True, exist_ok=True)
    table.to_excel(OUT, index=False, sheet_name=SHEET_NAME, engine="openpyxl")
    style_workbook(OUT)
    verify_workbook(OUT)
    render_preview(OUT)
    print(f"Saved: {OUT.relative_to(ROOT)}")
    print(f"Preview: {PREVIEW_OUT.relative_to(ROOT)}")
    print(f"Rows: {len(table):,}; columns: {len(table.columns):,}")
    print(f"Candidate roads: {diagnostics['Candidate Roads']:,}")
    print(
        f"Top-road location: {table.iloc[0]['Municipality / Ward']}; "
        f"priority score={table.iloc[0]['Robust Priority Score']:.3f}"
    )
    print(f"Terrain-score construction: {diagnostics['Model Mode']}")
    print("Workbook verification: passed")


if __name__ == "__main__":
    main()
