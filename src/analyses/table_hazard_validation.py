#!/usr/bin/env python3
"""Hazard Validation.

Plan: Compare five pre-specified terrain-score specifications using spatially blocked
presence-pseudo-background validation.
Framework: AnaSOP Sections 5-7 require mean and fold-specific AUC, held-out
top-quartile inventory capture, and rank correspondence. The selected transparent score
is retained for stability and auditability, not because every metric is superior.
"""
from __future__ import annotations

from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd
from scipy.stats import spearmanr
from sklearn.metrics import roc_auc_score
from sklearn.model_selection import GroupKFold

import _hazard_validation_shared as shared


ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "data/results/tables/Table_hazard_validation.xlsx"
SHEET_NAME = "Hazard Validation"
TABLE_TITLE = "Hazard Validation"


def model_for(
    model_kind: str,
    feature_indices: list[int],
    train_matrix: np.ndarray,
    train_outcome: np.ndarray,
) -> object | None:
    """Fit one declared validation specification on a training block."""
    if model_kind == "raw":
        return None
    if model_kind == "fixed":
        weights = {
            shared.terrain_score.FEATURE_NAMES[index]: shared.terrain_score.FALLBACK_WEIGHTS[
                shared.terrain_score.FEATURE_NAMES[index]
            ]
            for index in feature_indices
        }
        return shared.terrain_score.TransparentStandardizedScore(weights).fit(train_matrix)
    model = shared.logistic_model()
    model.fit(train_matrix, train_outcome)
    return model


def score_with(model_kind: str, model: object | None, matrix: np.ndarray) -> np.ndarray:
    """Return ranking scores without converting them to probabilities."""
    if model_kind == "raw":
        return matrix[:, 0].astype(float)
    if model is None:
        raise RuntimeError("A fitted model is required for this specification.")
    return np.asarray(model.decision_function(matrix), dtype=float)


def validate_specification(
    matrix: np.ndarray,
    outcome: np.ndarray,
    groups: np.ndarray,
    feature_indices: list[int],
    model_kind: str,
) -> tuple[dict[str, object], np.ndarray]:
    """Compute fold-specific AUC and capture plus a full-sample ranking score."""
    selected = matrix[:, feature_indices]
    auc_values: list[float] = []
    capture_values: list[float] = []
    splitter = GroupKFold(n_splits=5)
    for train, test in splitter.split(selected, outcome, groups):
        if len(np.unique(outcome[test])) < 2:
            continue
        fitted = model_for(model_kind, feature_indices, selected[train], outcome[train])
        test_score = score_with(model_kind, fitted, selected[test])
        auc_values.append(float(roc_auc_score(outcome[test], test_score)))
        presence = outcome[test] == 1
        threshold = float(np.quantile(test_score, 0.75))
        if model_kind != "raw":
            capture_values.append(float(np.mean(test_score[presence] >= threshold)))
    if len(auc_values) < 4:
        raise RuntimeError("Spatial validation produced fewer than four evaluable folds.")

    fitted_full = model_for(model_kind, feature_indices, selected, outcome)
    full_score = score_with(model_kind, fitted_full, selected)
    metrics = {
        "Spatial Folds": len(auc_values),
        "Mean Spatial AUC": float(np.mean(auc_values)),
        "Spatial AUC Range": f"{np.min(auc_values):.3f}–{np.max(auc_values):.3f}",
        "Held-Out Top-Quartile Capture": (
            float(np.mean(capture_values)) if capture_values else np.nan
        ),
    }
    return metrics, full_score


def validate_frozen_score(
    model: object,
    matrix: np.ndarray,
    outcome: np.ndarray,
    groups: np.ndarray,
) -> tuple[dict[str, object], np.ndarray]:
    """Evaluate the already propagated fixed score without refitting it."""
    full_score = np.asarray(model.decision_function(matrix), dtype=float)
    auc_values: list[float] = []
    capture_values: list[float] = []
    splitter = GroupKFold(n_splits=5)
    for _, test in splitter.split(matrix, outcome, groups):
        if len(np.unique(outcome[test])) < 2:
            continue
        test_score = full_score[test]
        auc_values.append(float(roc_auc_score(outcome[test], test_score)))
        threshold = float(np.quantile(test_score, 0.75))
        capture_values.append(
            float(np.mean(test_score[outcome[test] == 1] >= threshold))
        )
    if len(auc_values) < 4:
        raise RuntimeError("Frozen-score validation produced fewer than four folds.")
    return (
        {
            "Spatial Folds": len(auc_values),
            "Mean Spatial AUC": float(np.mean(auc_values)),
            "Spatial AUC Range": f"{np.min(auc_values):.3f}–{np.max(auc_values):.3f}",
            "Held-Out Top-Quartile Capture": float(np.mean(capture_values)),
        },
        full_score,
    )


def build_table() -> pd.DataFrame:
    """Build the planned five-row hazard validation table."""
    context, frozen_model = shared.prepare_historical_alignment_context()
    matrix = np.asarray(context["matrix"], dtype=float)
    outcome = np.asarray(context["outcome"], dtype=int)
    groups = np.asarray(context["groups"])
    validation_warning_counts = context["validation_warning_counts"]
    sample_label = (
        f"{context['presence_count']:,} / {context['background_count']:,}; "
        f"{validation_warning_counts['selected']:,} pre-sequence zones; "
        f"{context['interpretation_footprint_pct']:.1f}% GSI footprint"
    )
    specifications = [
        (
            "Full terrain + warning-zone logistic",
            [0, 1, 2, 3],
            "logistic",
            "Fitted comparator using warning zones designated before the 2016 earthquake sequence; diagnostic only.",
        ),
        (
            "Terrain-only logistic",
            [0, 1, 2],
            "logistic",
            "Diagnostic fitted comparator only.",
        ),
        (
            "Elevation + warning-zone logistic",
            [0, 3],
            "logistic",
            "Reduced comparator using warning zones designated before the 2016 earthquake sequence.",
        ),
        (
            "Warning-zone-only indicator",
            [3],
            "raw",
            "Pre-event binary indicator; AUC is diagnostic and top-quartile capture is not rank-resolving because of ties.",
        ),
        (
            "Fixed standardized terrain score",
            [0, 1, 2, 3],
            "fixed",
            "Frozen propagated terrain-context ranking evaluated without reviewer-driven refitting; not an occurrence probability.",
        ),
    ]

    rows: list[dict[str, object]] = []
    full_reference: np.ndarray | None = None
    for specification, indices, model_kind, interpretation in specifications:
        if model_kind == "fixed":
            metrics, full_score = validate_frozen_score(
                frozen_model, matrix, outcome, groups
            )
        else:
            metrics, full_score = validate_specification(
                matrix,
                outcome,
                groups,
                indices,
                model_kind,
            )
        if full_reference is None:
            full_reference = full_score
        correlation = float(
            spearmanr(full_score, full_reference, nan_policy="omit").statistic
        )
        rows.append(
            {
                "Specification": specification,
                "Validation Sample (Presence / Pseudo-Background)": sample_label,
                **metrics,
                "Rank Correlation vs Full": correlation,
                "Permitted Interpretation": interpretation,
            }
        )
    table = pd.DataFrame(rows)
    if table.shape != (5, 8):
        raise RuntimeError(f"Expected a 5 × 8 table, found {table.shape}.")
    return table


def style_workbook(path: Path) -> None:
    """Apply the accepted compact, Word-safe scientific-table style."""
    workbook = load_workbook(path)
    worksheet = workbook[SHEET_NAME]
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A3"
    worksheet.auto_filter.ref = f"A2:H{worksheet.max_row}"
    worksheet.sheet_view.zoomScale = 95
    worksheet.print_area = f"A1:H{worksheet.max_row}"
    worksheet.print_title_rows = "1:2"
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A3
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 1
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_margins = PageMargins(
        left=0.25,
        right=0.25,
        top=0.35,
        bottom=0.35,
        header=0.10,
        footer=0.10,
    )

    header_fill = PatternFill("solid", fgColor="17365D")
    header_font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
    body_font = Font(name="Aptos", size=9.5, color="172033")
    subtle_border = Border(bottom=Side(style="thin", color="D0D5DD"))

    worksheet.merge_cells("A1:H1")
    title_cell = worksheet["A1"]
    title_cell.value = TABLE_TITLE
    title_cell.fill = PatternFill("solid", fgColor="D9EAF7")
    title_cell.font = Font(name="Aptos Display", size=15, bold=True, color="17365D")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    worksheet.row_dimensions[1].height = 28

    for cell in worksheet[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[2].height = 46

    for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = subtle_border
        row[0].fill = PatternFill("solid", fgColor="E4EEF7")
        if row[0].value == "Fixed standardized terrain score":
            row[0].fill = PatternFill("solid", fgColor="D9EDE9")
        elif row[0].value == "Warning-zone-only indicator":
            row[0].fill = PatternFill("solid", fgColor="F4D6D3")
        row[3].number_format = "0.0%"
        row[5].number_format = "0.0%"
        row[6].number_format = "0.000"
        for cell in row[1:7]:
            cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
        worksheet.row_dimensions[row[0].row].height = 44

    widths = {
        "A": 37,
        "B": 27,
        "C": 13,
        "D": 18,
        "E": 19,
        "F": 25,
        "G": 21,
        "H": 49,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    for column in ("D", "G"):
        worksheet.conditional_formatting.add(
            f"{column}3:{column}{worksheet.max_row}",
            ColorScaleRule(
                start_type="min",
                start_color="F8696B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="63BE7B",
            ),
        )
    excel_table = Table(displayName="HazardValidation", ref=f"A2:H{worksheet.max_row}")
    excel_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(excel_table)
    workbook.save(path)


def verify_workbook(path: Path) -> None:
    """Verify dimensions, numeric fields, and absence of spreadsheet errors."""
    workbook = load_workbook(path, data_only=False)
    if workbook.sheetnames != [SHEET_NAME]:
        raise RuntimeError(f"Unexpected workbook sheets: {workbook.sheetnames}")
    worksheet = workbook[SHEET_NAME]
    if worksheet.max_row != 7 or worksheet.max_column != 8:
        raise RuntimeError(
            f"Expected 7 rows including title and header and 8 columns; found "
            f"{worksheet.max_row} × {worksheet.max_column}."
        )
    if worksheet["A1"].value != TABLE_TITLE:
        raise RuntimeError("Workbook title row is missing or incorrect.")
    error_tokens = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"}
    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value in error_tokens:
                raise RuntimeError(f"Spreadsheet error token in {cell.coordinate}: {cell.value}")
    for row in range(3, 8):
        for column in (3, 4, 7):
            if not isinstance(worksheet.cell(row, column).value, (int, float)):
                raise RuntimeError(f"Expected numeric value at row {row}, column {column}.")
        capture = worksheet.cell(row, 6).value
        if row != 6 and not isinstance(capture, (int, float)):
            raise RuntimeError(f"Expected numeric capture value at row {row}.")


def main() -> None:
    table = build_table()
    OUT.parent.mkdir(parents=True, exist_ok=True)
    table.to_excel(
        OUT,
        index=False,
        sheet_name=SHEET_NAME,
        engine="openpyxl",
        startrow=1,
    )
    style_workbook(OUT)
    verify_workbook(OUT)
    print(f"Saved: {OUT.relative_to(ROOT)}")
    print(f"Rows: {len(table):,}; columns: {len(table.columns):,}")
    for row in table.itertuples(index=False):
        print(
            f"{row[0]}: AUC={row[3]:.3f}; range={row[4]}; "
            f"capture={row[5]:.3f}; rho={row[6]:.3f}"
        )
    print("Workbook verification: passed")


if __name__ == "__main__":
    main()
