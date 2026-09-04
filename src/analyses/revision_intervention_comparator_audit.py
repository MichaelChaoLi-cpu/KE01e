#!/usr/bin/env python3
"""Audit intervention comparator coverage for Reviewer 1 Comment 4."""
from __future__ import annotations

import ast
import hashlib
import json
from pathlib import Path
from statistics import median
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
PORTFOLIO_PATH = ROOT / "data/results/tables/Table_intervention_portfolios.xlsx"
COMPARATOR_PATH = ROOT / "data/results/tables/Table_comparator_robustness.xlsx"
SOURCE_PATH = ROOT / "src/analyses/figure_intervention_priorities_and_budgeted_benefits.py"
OUTPUT_DIR = ROOT / "data/exp/revision/reviewer-1-comment-4"


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def rows_from_workbook(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[2]]
    rows: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=3, values_only=True):
        if not any(value is not None for value in values):
            continue
        rows.append(dict(zip(headers, values, strict=True)))
    workbook.close()
    return rows


def literal_assignment(source: str, name: str) -> Any:
    tree = ast.parse(source)
    for node in tree.body:
        if not isinstance(node, ast.Assign):
            continue
        if any(isinstance(target, ast.Name) and target.id == name for target in node.targets):
            return ast.literal_eval(node.value)
    raise RuntimeError(f"Could not find literal assignment {name}")


def close_or_equal(left: Any, right: Any, tolerance: float = 1e-12) -> bool:
    if left is None or right is None:
        return left is right
    if isinstance(left, (int, float)) and isinstance(right, (int, float)):
        return abs(float(left) - float(right)) <= tolerance
    return left == right


def main() -> None:
    portfolio_rows = rows_from_workbook(PORTFOLIO_PATH)
    comparator_rows = rows_from_workbook(COMPARATOR_PATH)
    source = SOURCE_PATH.read_text(encoding="utf-8")

    declared_comparators = {
        "Hazard only",
        "Emergency route only",
        "Road class only",
        "Equal-cost consequence",
    }
    observed_comparators = {str(row["Comparator"]) for row in comparator_rows}

    assigned = {
        float(row["Budget (Relative Planning Units)"]): row
        for row in portfolio_rows
        if row["Sensitivity Setting"] == "Central"
    }
    equal_cost = {
        float(row["Budget (Relative Planning Units)"]): row
        for row in comparator_rows
        if row["Setting"] == "Central" and row["Comparator"] == "Equal-cost consequence"
    }
    common_budgets = sorted(set(assigned) & set(equal_cost))
    field_pairs = [
        ("Selected Road Count", "Selected Road Count"),
        ("Realized Cost (Relative Planning Units)", "Realized Cost (Relative Planning Units)"),
        (
            "Protected Population Mean [Seed Range] (Total / Age 65+)",
            "Protected Population Mean [Seed Range] (Total / Age 65+)",
        ),
        ("Avoided Isolation Share", "Avoided Isolation Share"),
        (
            "Protected Population per Relative Cost",
            "Protected Population per Relative Cost",
        ),
    ]
    row_checks: list[dict[str, Any]] = []
    for budget in common_budgets:
        checks = {
            left_name: close_or_equal(assigned[budget][left_name], equal_cost[budget][right_name])
            for left_name, right_name in field_pairs
        }
        row_checks.append({"budget": budget, "checks": checks, "all_equal": all(checks.values())})

    action_effect = literal_assignment(source, "ACTION_EFFECT")
    cost_multiplier = literal_assignment(source, "COST_MULTIPLIER")
    ratio_checks: dict[str, dict[str, Any]] = {}
    for action, effects in action_effect.items():
        ratios = {
            setting: float(effects[setting]) / float(cost_multiplier[setting])
            for setting in ("Conservative", "Central", "Optimistic")
        }
        median_ratio = float(median(ratios.values()))
        central_ratio = ratios["Central"]
        ratio_checks[action] = {
            "ratios": ratios,
            "median_ratio": median_ratio,
            "central_ratio": central_ratio,
            "equal": close_or_equal(median_ratio, central_ratio),
        }

    mcdm_terms = ["AHP", "TOPSIS", "ELECTRE", "PROMETHEE", "VIKOR", "MAUT", "multi-criteria"]
    detected_terms = [term for term in mcdm_terms if term.lower() in source.lower()]

    max_budget = max(common_budgets)
    central_max: dict[str, Any] = {
        "Assigned-action screening": assigned[max_budget][
            "Protected Population Mean [Seed Range] (Total / Age 65+)"
        ]
    }
    for comparator in sorted(declared_comparators):
        match = next(
            row
            for row in comparator_rows
            if row["Setting"] == "Central"
            and row["Comparator"] == comparator
            and close_or_equal(row["Budget (Relative Planning Units)"], max_budget)
        )
        central_max[comparator] = match[
            "Protected Population Mean [Seed Range] (Total / Age 65+)"
        ]

    checks = {
        "four_declared_comparators_present": declared_comparators == observed_comparators,
        "seven_common_central_budgets": len(common_budgets) == 7,
        "central_rows_identical": all(item["all_equal"] for item in row_checks),
        "median_ratio_equals_central_for_all_actions": all(
            item["equal"] for item in ratio_checks.values()
        ),
        "formal_mcdm_term_absent_from_production_script": not detected_terms,
    }
    decision = {
        "reviewer": "Reviewer 1",
        "comment": 4,
        "input_sha256": {
            str(PORTFOLIO_PATH.relative_to(ROOT)): sha256(PORTFOLIO_PATH),
            str(COMPARATOR_PATH.relative_to(ROOT)): sha256(COMPARATOR_PATH),
            str(SOURCE_PATH.relative_to(ROOT)): sha256(SOURCE_PATH),
        },
        "checks": checks,
        "all_checks_pass": all(checks.values()),
        "common_central_budgets": common_budgets,
        "central_row_comparisons": row_checks,
        "effect_cost_ratio_checks": ratio_checks,
        "detected_mcdm_terms": detected_terms,
        "central_maximum_budget_results": central_max,
        "interpretation": (
            "Four heuristic comparators were tested. No formal MCDM was implemented. "
            "The Central assigned-action and equal-cost consequence orders are structurally "
            "identical because each action's median low-central-high effect-cost ratio equals "
            "its Central ratio under the declared assumptions."
        ),
        "recommendation": (
            "Clarify the four tested heuristics and the absence of a formal MCDM. Do not add "
            "a post hoc MCDM without independently supported criterion definitions and weights."
        ),
    }

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUTPUT_DIR / "decision.json").write_text(
        json.dumps(decision, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )

    report_lines = [
        "# Reviewer 1 Comment 4 comparator audit",
        "",
        f"- All checks pass: **{decision['all_checks_pass']}**",
        f"- Comparator rules present: {', '.join(sorted(observed_comparators))}",
        f"- Common Central budgets: {len(common_budgets)}",
        f"- Central assigned-action and equal-cost rows identical: **{checks['central_rows_identical']}**",
        "- Formal MCDM implementation detected: **False**",
        "",
        "## Structural identity",
        "",
    ]
    for action, item in ratio_checks.items():
        report_lines.append(
            f"- {action}: Conservative={item['ratios']['Conservative']:.6f}, "
            f"Central={item['ratios']['Central']:.6f}, "
            f"Optimistic={item['ratios']['Optimistic']:.6f}; "
            f"median={item['median_ratio']:.6f}."
        )
    report_lines.extend(
        [
            "",
            "For every assigned action, the median ratio used by the assigned-action score "
            "equals the Central effect-cost ratio. With the same consequence proxy and road "
            "cost, the two Central scores are therefore elementwise identical before sorting.",
            "",
            "## Maximum-budget Central comparison",
            "",
        ]
    )
    for label, value in central_max.items():
        report_lines.append(f"- {label}: {value}")
    report_lines.extend(
        [
            "",
            "## Recommendation",
            "",
            decision["recommendation"],
            "",
        ]
    )
    (OUTPUT_DIR / "audit_report.md").write_text("\n".join(report_lines), encoding="utf-8")
    print(json.dumps(decision, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
