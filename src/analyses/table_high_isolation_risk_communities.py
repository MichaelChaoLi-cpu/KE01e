#!/usr/bin/env python3
"""High-Isolation-Risk Communities.

Plan: List the 30 communities with the largest Heavy-scenario population and
older-population isolation burden in a compact operational summary.
Framework: AnaSOP Sections 5-7 use the accepted baseline community definition,
1,000-draw scenario-conditional isolation frequencies, Heavy-scenario service
loss simulation, and candidate gateway-road dependence. Frequencies are model-
conditional screening results, not calibrated real-world probabilities.
"""
from __future__ import annotations

import os
from pathlib import Path

os.environ.pop("PROJ_LIB", None)
os.environ.pop("PROJ_DATA", None)

import numpy as np
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd
from PIL import Image, ImageDraw, ImageFont
import shapely

import figure_basic_service_reachability_loss as service_loss
import table_municipality_isolation_and_service_loss_summary as municipality_summary
import table_priority_road_sections as priority_roads


ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "data/results/tables/Table_high_isolation_risk_communities.xlsx"
PREVIEW_OUT = ROOT / "data/exp/table_previews/Table_high_isolation_risk_communities.png"
SHEET_NAME = "High-Risk Communities"
TABLE_TITLE = "High-Isolation-Risk Communities"
TOP_COMMUNITIES = 30


def community_admin_names(context: dict[str, object]) -> np.ndarray:
    """Assign each population-weighted community centroid to an English admin name."""
    community = context["community"]
    admin = context["admin"].reset_index(drop=True)
    admin_geometry = context["admin_geometry"]
    points = shapely.points(
        community["Longitude"].to_numpy(dtype=float),
        community["Latitude"].to_numpy(dtype=float),
    )
    names: list[str] = []
    for point in points:
        matches = np.flatnonzero(shapely.covers(admin_geometry, point))
        if matches.size:
            admin_position = int(matches[0])
        else:
            admin_position = int(np.argmin(shapely.distance(admin_geometry, point)))
        code = str(admin.iloc[admin_position]["Municipality Code"])
        name = priority_roads.MUNICIPALITY_ENGLISH_BY_CODE.get(code)
        if name is None:
            raise RuntimeError(f"Missing English municipality name for code {code}.")
        names.append(name)
    return np.asarray(names, dtype=object)


def gateway_section_counts(context: dict[str, object]) -> np.ndarray:
    """Count Heavy-active candidate sections crossing each community's attached roots."""
    community_count = len(context["community"])
    attachment_community = context["attachment_community"]
    attachment_root = context["attachment_root"]
    candidate_u = context["candidate_u"]
    candidate_v = context["candidate_v"]
    candidate_edge_section = context["candidate_edge_section"]
    heavy_propensity = context["heavy_propensity"]
    heavy_active_edge = heavy_propensity[candidate_edge_section] > 0
    counts = np.zeros(community_count, dtype="int32")
    for position in range(community_count):
        roots = np.unique(attachment_root[attachment_community == position])
        touches_u = np.isin(candidate_u, roots)
        touches_v = np.isin(candidate_v, roots)
        crossing = heavy_active_edge & np.logical_xor(touches_u, touches_v)
        counts[position] = np.unique(candidate_edge_section[crossing]).size
    return counts


def build_table() -> tuple[pd.DataFrame, dict[str, object]]:
    """Build the top-30 community ranking from accepted simulation components."""
    context = municipality_summary.prepare_network_and_outcomes()
    community = context["community"].copy()
    frequencies = context["frequencies"]
    total = community["Total_Population"].to_numpy(dtype=float)
    older = community["Population_Age_65"].to_numpy(dtype=float)
    heavy = frequencies["Heavy"].astype(float)
    burden = heavy * (total + 0.5 * older)
    order = np.lexsort((-frequencies["Extreme"], -burden))[:TOP_COMMUNITIES]
    municipality_names = community_admin_names(context)
    gateway_counts = gateway_section_counts(context)

    rows: list[dict[str, object]] = []
    for rank, position in enumerate(order, start=1):
        service_values = {
            name: float(context["loss_frequency"][name][position])
            for name in service_loss.SERVICE_CLASSES
            if np.isfinite(context["loss_frequency"][name][position])
        }
        if service_values:
            principal_service = max(service_values, key=service_values.get)
            principal_value = service_values[principal_service]
            principal_loss = (
                f"{principal_service} ({principal_value:.1%})"
                if principal_value >= 0.001
                else "None detected"
            )
        else:
            principal_loss = "Baseline unreachable"

        rows.append(
            {
                "Priority Rank": rank,
                "Municipality / Ward": municipality_names[position],
                "Community Centroid (Lon, Lat)": (
                    f"{community.iloc[position]['Longitude']:.4f}, "
                    f"{community.iloc[position]['Latitude']:.4f}"
                ),
                "Mesh Count": int(community.iloc[position]["Mesh_Count"]),
                "Population (Total / Age 65+)": f"{total[position]:,.0f} / {older[position]:,.0f}",
                "Candidate Gateway Section Count": int(gateway_counts[position]),
                "Moderate Isolation Frequency": float(frequencies["Moderate"][position]),
                "Heavy Isolation Frequency": float(heavy[position]),
                "Extreme Isolation Frequency": float(frequencies["Extreme"][position]),
                "Heavy Expected Isolated Population (Total / Age 65+)": (
                    f"{total[position] * heavy[position]:,.1f} / "
                    f"{older[position] * heavy[position]:,.1f}"
                ),
                "Principal Heavy Service Loss": principal_loss,
            }
        )

    table = pd.DataFrame(rows)
    if table.shape != (TOP_COMMUNITIES, 11):
        raise RuntimeError(
            f"Expected a {TOP_COMMUNITIES} × 11 table, found {table.shape}."
        )
    if not table["Priority Rank"].equals(
        pd.Series(range(1, TOP_COMMUNITIES + 1), name="Priority Rank")
    ):
        raise RuntimeError("Community ranks are not sequential.")
    diagnostics = {
        "Baseline Communities": len(community),
        "Top Heavy Frequency": float(table.iloc[0]["Heavy Isolation Frequency"]),
        "Top Location": str(table.iloc[0]["Municipality / Ward"]),
        "Model Mode": context["model_mode"],
    }
    return table, diagnostics


def style_workbook(path: Path) -> None:
    """Style the title-first compact top-community workbook."""
    workbook = load_workbook(path)
    worksheet = workbook[SHEET_NAME]
    worksheet.insert_rows(1)
    worksheet.merge_cells("A1:K1")
    worksheet["A1"] = TABLE_TITLE
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "C3"
    worksheet.auto_filter.ref = f"A2:K{worksheet.max_row}"
    worksheet.sheet_view.zoomScale = 90
    worksheet.print_area = f"A1:K{worksheet.max_row}"
    worksheet.print_title_rows = "1:2"
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A3
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_margins = PageMargins(
        left=0.20, right=0.20, top=0.30, bottom=0.30, header=0.10, footer=0.10
    )

    title_fill = PatternFill("solid", fgColor="D9EAF7")
    title_font = Font(name="Aptos Display", size=14, bold=True, color="17365D")
    header_fill = PatternFill("solid", fgColor="17365D")
    header_font = Font(name="Aptos", size=9, bold=True, color="FFFFFF")
    body_font = Font(name="Aptos", size=8.8, color="172033")
    subtle_border = Border(bottom=Side(style="thin", color="D0D5DD"))
    rank_fills = {
        "top": PatternFill("solid", fgColor="F4CCCC"),
        "middle": PatternFill("solid", fgColor="FCE5CD"),
        "lower": PatternFill("solid", fgColor="FFF2CC"),
    }
    worksheet["A1"].fill = title_fill
    worksheet["A1"].font = title_font
    worksheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    worksheet.row_dimensions[1].height = 30
    for cell in worksheet[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[2].height = 50

    for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = subtle_border
        rank = int(row[0].value)
        row[0].fill = (
            rank_fills["top"]
            if rank <= 10
            else rank_fills["middle"]
            if rank <= 20
            else rank_fills["lower"]
        )
        for column in (1, 4, 6):
            row[column - 1].number_format = "#,##0"
        for column in (7, 8, 9):
            row[column - 1].number_format = "0.0%"
        for cell in (*row[:1], *row[3:4], *row[5:9]):
            cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
        worksheet.row_dimensions[row[0].row].height = 31

    widths = {
        "A": 12,
        "B": 24,
        "C": 25,
        "D": 13,
        "E": 27,
        "F": 23,
        "G": 20,
        "H": 19,
        "I": 20,
        "J": 34,
        "K": 31,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width
    for column in ("G", "H", "I"):
        worksheet.conditional_formatting.add(
            f"{column}3:{column}{worksheet.max_row}",
            ColorScaleRule(
                start_type="min",
                start_color="FFFFFF",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="F8696B",
            ),
        )

    excel_table = Table(
        displayName="HighIsolationRiskCommunities",
        ref=f"A2:K{worksheet.max_row}",
    )
    excel_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(excel_table)
    workbook.save(path)


def render_preview(path: Path, output: Path = PREVIEW_OUT) -> None:
    """Render the complete workbook as one title-first review PNG."""
    workbook = load_workbook(path, data_only=True)
    values = list(workbook[SHEET_NAME].values)
    title = str(values[0][0])
    headers = [str(value) for value in values[1]]
    rows = [list(row) for row in values[2:]]
    widths = [95, 210, 230, 115, 225, 190, 175, 170, 175, 280, 300]
    margin = 22
    title_height = 84
    header_height = 105
    row_height = 60
    table_width = sum(widths)
    image = Image.new(
        "RGB",
        (table_width + 2 * margin, title_height + header_height + row_height * len(rows) + margin),
        "white",
    )
    draw = ImageDraw.Draw(image)
    title_font = priority_roads._preview_font(28, bold=True)
    note_font = priority_roads._preview_font(16)
    header_font = priority_roads._preview_font(16, bold=True)
    body_font = priority_roads._preview_font(16)
    body_bold = priority_roads._preview_font(16, bold=True)
    draw.text((margin, 16), title, font=title_font, fill="#17365D")
    draw.text(
        (margin, 52),
        "Top 30 communities ranked by Heavy-scenario population and age-65+ isolation burden",
        font=note_font,
        fill="#52606D",
    )
    x_positions = [margin]
    for width in widths:
        x_positions.append(x_positions[-1] + width)
    for column, header in enumerate(headers):
        x0, x1 = x_positions[column], x_positions[column + 1]
        draw.rectangle((x0, title_height, x1, title_height + header_height), fill="#17365D")
        wrapped = priority_roads._wrap_preview_text(
            draw, header, header_font, widths[column] - 14
        )
        bbox = draw.multiline_textbbox((0, 0), wrapped, font=header_font, spacing=3)
        text_height = bbox[3] - bbox[1]
        draw.multiline_text(
            ((x0 + x1) / 2, title_height + (header_height - text_height) / 2),
            wrapped,
            font=header_font,
            fill="white",
            spacing=3,
            anchor="ma",
            align="center",
        )

    frequency_columns = (6, 7, 8)
    ranges = {
        column: (
            min(float(row[column]) for row in rows),
            max(float(row[column]) for row in rows),
        )
        for column in frequency_columns
    }

    def scale_colour(value: float, low: float, high: float) -> str:
        ratio = 0.5 if high == low else (value - low) / (high - low)
        if ratio <= 0.5:
            local, start, end = ratio / 0.5, (255, 255, 255), (255, 235, 132)
        else:
            local, start, end = (ratio - 0.5) / 0.5, (255, 235, 132), (248, 105, 107)
        rgb = tuple(round(a + (b - a) * local) for a, b in zip(start, end))
        return "#" + "".join(f"{channel:02X}" for channel in rgb)

    numeric_right = {0, 3, 5, 6, 7, 8}
    for row_number, row in enumerate(rows):
        y0 = title_height + header_height + row_number * row_height
        y1 = y0 + row_height
        base_fill = "#FFFFFF" if row_number % 2 == 0 else "#F8FAFC"
        rank = int(row[0])
        rank_fill = "#F4CCCC" if rank <= 10 else "#FCE5CD" if rank <= 20 else "#FFF2CC"
        for column, value in enumerate(row):
            x0, x1 = x_positions[column], x_positions[column + 1]
            fill = rank_fill if column == 0 else base_fill
            if column in frequency_columns:
                low, high = ranges[column]
                fill = scale_colour(float(value), low, high)
            draw.rectangle((x0, y0, x1, y1), fill=fill, outline="#D0D5DD", width=1)
            if column in frequency_columns:
                display_value = f"{float(value):.1%}"
            elif column in {0, 3, 5}:
                display_value = f"{int(value):,}"
            else:
                display_value = "" if value is None else str(value)
            wrapped = priority_roads._wrap_preview_text(
                draw, display_value, body_font, widths[column] - 14
            )
            font = body_bold if column == 0 else body_font
            if column in numeric_right:
                draw.multiline_text(
                    (x1 - 7, y0 + 8),
                    wrapped,
                    font=font,
                    fill="#172033",
                    spacing=2,
                    anchor="ra",
                    align="right",
                )
            else:
                draw.multiline_text(
                    (x0 + 7, y0 + 8),
                    wrapped,
                    font=font,
                    fill="#172033",
                    spacing=2,
                )
    output.parent.mkdir(parents=True, exist_ok=True)
    image.save(output, optimize=True)


def verify_workbook(path: Path) -> None:
    """Verify title/header placement, dimensions, ranks, and numeric ranges."""
    workbook = load_workbook(path, data_only=False)
    if workbook.sheetnames != [SHEET_NAME]:
        raise RuntimeError(f"Unexpected workbook sheets: {workbook.sheetnames}")
    worksheet = workbook[SHEET_NAME]
    if worksheet.max_row != TOP_COMMUNITIES + 2 or worksheet.max_column != 11:
        raise RuntimeError(
            f"Expected {TOP_COMMUNITIES + 2} rows and 11 columns; found "
            f"{worksheet.max_row} × {worksheet.max_column}."
        )
    if worksheet["A1"].value != TABLE_TITLE:
        raise RuntimeError("Workbook title row is missing or incorrect.")
    ranks = [worksheet.cell(row, 1).value for row in range(3, TOP_COMMUNITIES + 3)]
    if ranks != list(range(1, TOP_COMMUNITIES + 1)):
        raise RuntimeError("Workbook ranks are not sequential.")
    error_tokens = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"}
    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value in error_tokens:
                raise RuntimeError(f"Spreadsheet error token in {cell.coordinate}: {cell.value}")
    for row in range(3, TOP_COMMUNITIES + 3):
        for column in (7, 8, 9):
            value = float(worksheet.cell(row, column).value)
            if not 0 <= value <= 1:
                raise RuntimeError(f"Isolation frequency outside [0, 1] at row {row}.")


def main() -> None:
    table, diagnostics = build_table()
    OUT.parent.mkdir(parents=True, exist_ok=True)
    table.to_excel(OUT, index=False, sheet_name=SHEET_NAME, engine="openpyxl")
    style_workbook(OUT)
    verify_workbook(OUT)
    render_preview(OUT)
    print(f"Saved: {OUT.relative_to(ROOT)}")
    print(f"Preview: {PREVIEW_OUT.relative_to(ROOT)}")
    print(f"Rows: {len(table):,}; columns: {len(table.columns):,}")
    print(f"Baseline communities: {diagnostics['Baseline Communities']:,}")
    print(
        f"Top location: {diagnostics['Top Location']}; "
        f"Heavy isolation frequency={diagnostics['Top Heavy Frequency']:.3f}"
    )
    print(f"Terrain-score construction: {diagnostics['Model Mode']}")
    print("Workbook verification: passed")


if __name__ == "__main__":
    main()
