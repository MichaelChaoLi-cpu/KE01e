#!/usr/bin/env python3
"""Municipality Isolation and Service-Loss Summary.

Plan: Summarize simulation-conditional isolation, exposed total and older
population, service reachability loss, and excess travel time for all 49
municipalities or wards.
Framework: AnaSOP Sections 5-7 use the accepted 1,000-draw section-closure
simulation. Population consequences are allocated through eligible 125 m meshes;
service loss is evaluated under the Heavy rainfall scenario used in the accepted
service-reachability figure.
"""
from __future__ import annotations

import os
from pathlib import Path

os.environ.pop("PROJ_LIB", None)
os.environ.pop("PROJ_DATA", None)

import numpy as np
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd
from rasterio.transform import from_bounds
from scipy.sparse import coo_matrix
from scipy.sparse.csgraph import connected_components
import shapely

import figure_basic_service_reachability_loss as service_loss
import figure_community_isolation_frequency_and_exposed_population as isolation
import figure_road_disruption_exposure_and_observed_restriction_evidence as road_exposure
import table_priority_road_sections as priority_roads


ROOT = Path(__file__).resolve().parents[2]
PROCESSED = ROOT / "data/processed"
ADMIN_PATH = PROCESSED / "administrative_areas_preprocessed.parquet"
ROAD_PATH = PROCESSED / "road_sections_preprocessed.parquet"
EDGE_PATH = PROCESSED / "road_edges_preprocessed.parquet"
NODE_PATH = PROCESSED / "road_nodes_preprocessed.parquet"
OUT = ROOT / "data/results/tables/Table_municipality_isolation_and_service_loss_summary.xlsx"
SHEET_NAME = "Municipality Summary"
TABLE_TITLE = "Municipality Isolation and Service-Loss Summary"


def prepare_network_and_outcomes() -> dict[str, object]:
    """Reproduce the accepted road, community, isolation, and service simulations."""
    admin = pd.read_parquet(
        ADMIN_PATH,
        columns=["Municipality Code", "Municipality Label", "Municipality Name", "Geometry"],
    )
    admin_geometry = road_exposure.decode_geometry(admin.pop("Geometry"))
    admin_union = shapely.union_all(admin_geometry)
    min_x, min_y, max_x, max_y = shapely.bounds(admin_union)
    pad_x = (max_x - min_x) * 0.025
    pad_y = (max_y - min_y) * 0.025
    extent = (min_x - pad_x, max_x + pad_x, min_y - pad_y, max_y + pad_y)
    west, east, south, north = extent
    display_height = max(
        650,
        round(road_exposure.DISPLAY_WIDTH * (north - south) / (east - west)),
    )
    display_shape = (display_height, road_exposure.DISPLAY_WIDTH)
    display_transform = from_bounds(
        west,
        south,
        east,
        north,
        road_exposure.DISPLAY_WIDTH,
        display_height,
    )

    terrain_scores, _, model_mode, elevation_grid = road_exposure.build_landslide_scores(
        admin,
        admin_geometry,
        admin_union,
        extent,
        display_shape,
        display_transform,
    )
    roads = pd.read_parquet(
        ROAD_PATH,
        columns=[
            "Road Section ID",
            "Emergency Route Membership",
            "Network Analysis Eligible",
            "Geometry",
        ],
    )
    roads = roads.loc[roads["Network Analysis Eligible"]].reset_index(drop=True)
    road_geometry = road_exposure.decode_geometry(roads.pop("Geometry"))
    road_scores = road_exposure.road_scores(road_geometry, terrain_scores, extent, elevation_grid)

    heavy_lower = isolation.positive_score_quantile(
        road_scores["Heavy"], isolation.CANDIDATE_QUANTILE
    )
    heavy_upper = isolation.positive_score_quantile(
        road_scores["Heavy"], isolation.UPPER_MAPPING_QUANTILE
    )
    candidate = np.isfinite(road_scores["Heavy"]) & (road_scores["Heavy"] >= heavy_lower)
    candidate_ids = roads.loc[candidate, "Road Section ID"].reset_index(drop=True)
    candidate_position = pd.Series(
        np.arange(len(candidate_ids), dtype="int32"),
        index=candidate_ids,
    )

    nodes = pd.read_parquet(
        NODE_PATH,
        columns=["Network Node ID", "Network Component ID", "Geometry"],
    )
    node_geometry = road_exposure.decode_geometry(nodes.pop("Geometry"))
    node_index = pd.Index(nodes["Network Node ID"])
    edges = pd.read_parquet(
        EDGE_PATH,
        columns=[
            "Road Section ID",
            "From Node ID",
            "To Node ID",
            "Network Component ID",
            "Emergency Route Membership",
            "Baseline Edge Travel Time (min)",
            "Network Analysis Eligible",
        ],
    )
    edges = edges.loc[edges["Network Analysis Eligible"]].reset_index(drop=True)
    edge_u = node_index.get_indexer(edges["From Node ID"])
    edge_v = node_index.get_indexer(edges["To Node ID"])
    if np.any(edge_u < 0) or np.any(edge_v < 0):
        raise RuntimeError("Road edges reference missing network nodes.")
    edge_candidate = edges["Road Section ID"].isin(candidate_ids).to_numpy()
    edge_candidate_position_full = (
        edges["Road Section ID"].map(candidate_position).fillna(-1).to_numpy(dtype="int32")
    )

    stable_u = edge_u[~edge_candidate]
    stable_v = edge_v[~edge_candidate]
    stable_graph = coo_matrix(
        (
            np.ones(len(stable_u) * 2, dtype="uint8"),
            (
                np.concatenate([stable_u, stable_v]),
                np.concatenate([stable_v, stable_u]),
            ),
        ),
        shape=(len(nodes), len(nodes)),
    ).tocsr()
    root_count, stable_labels = connected_components(
        stable_graph,
        directed=False,
        return_labels=True,
    )
    stable_labels = stable_labels.astype("int32")

    candidate_u = stable_labels[edge_u[edge_candidate]]
    candidate_v = stable_labels[edge_v[edge_candidate]]
    candidate_edge_section = (
        edges.loc[edge_candidate, "Road Section ID"]
        .map(candidate_position)
        .to_numpy(dtype="int32")
    )
    candidate_edge_time = edges.loc[
        edge_candidate, "Baseline Edge Travel Time (min)"
    ].to_numpy(dtype="float64")
    between_root = candidate_u != candidate_v
    candidate_u = candidate_u[between_root]
    candidate_v = candidate_v[between_root]
    candidate_edge_section = candidate_edge_section[between_root]
    candidate_edge_time = candidate_edge_time[between_root]

    target_definitions, target_network_components = isolation.external_target_definitions(
        nodes,
        node_geometry,
        stable_labels,
        edges,
        edge_u,
        edge_v,
        admin_union,
    )
    target_roots = target_definitions["Primary boundary gateways"]

    (
        community,
        attachment_community,
        attachment_root,
        community_diagnostics,
        selected_mesh,
        selected_mesh_geometry,
    ) = isolation.build_baseline_communities(
        nodes,
        node_geometry,
        stable_labels,
        target_network_components,
    )

    frequencies: dict[str, np.ndarray] = {}
    candidate_scores: dict[str, np.ndarray] = {}
    for scenario_index, scenario in enumerate(("Moderate", "Heavy", "Extreme")):
        candidate_scores[scenario] = (
            pd.Series(road_scores[scenario], index=roads["Road Section ID"])
            .reindex(candidate_ids)
            .to_numpy(dtype="float32")
        )
        propensity = isolation.closure_propensity(
            candidate_scores[scenario], heavy_lower, heavy_upper
        )
        print(
            f"Simulating municipality isolation, {scenario}: "
            f"{np.count_nonzero(propensity):,} non-zero candidate sections"
        )
        frequencies[scenario] = isolation.simulate_isolation(
            candidate_u,
            candidate_v,
            candidate_edge_section,
            propensity,
            root_count,
            target_roots,
            attachment_community,
            attachment_root,
            len(community),
            isolation.RANDOM_SEED,
        )

    heavy_propensity = isolation.closure_propensity(
        candidate_scores["Heavy"], heavy_lower, heavy_upper
    )
    pair_reduction = service_loss.prepare_pair_reduction(
        candidate_u,
        candidate_v,
        candidate_edge_section,
        candidate_edge_time,
        root_count,
    )
    service_geometry, source_counts = service_loss.service_geometries()
    service_roots, service_nodes, attached_counts = service_loss.attach_services_to_roots(
        service_geometry,
        node_geometry,
        stable_labels,
    )
    for service in service_loss.SERVICE_CLASSES:
        if len(service_roots[service]) == 0:
            raise RuntimeError(f"No {service} features attach to the road network.")
    loss_frequency, _, _ = service_loss.simulate_service_loss(
        heavy_propensity,
        pair_reduction,
        root_count,
        service_roots,
        attachment_community,
        attachment_root,
        len(community),
    )
    mean_excess, baseline_distance = service_loss.full_graph_service_excess_time(
        edge_u,
        edge_v,
        edges["Baseline Edge Travel Time (min)"].to_numpy(dtype="float64"),
        edge_candidate_position_full,
        heavy_propensity,
        len(nodes),
        service_nodes,
        selected_mesh["Community Position"].to_numpy(dtype="int32"),
        selected_mesh["Attached Node Position"].to_numpy(dtype="int32"),
        len(community),
    )

    return {
        "admin": admin,
        "admin_geometry": admin_geometry,
        "community": community,
        "attachment_community": attachment_community,
        "attachment_root": attachment_root,
        "selected_mesh": selected_mesh,
        "selected_mesh_geometry": selected_mesh_geometry,
        "frequencies": frequencies,
        "candidate_u": candidate_u,
        "candidate_v": candidate_v,
        "candidate_edge_section": candidate_edge_section,
        "heavy_propensity": heavy_propensity,
        "loss_frequency": loss_frequency,
        "mean_excess": mean_excess,
        "baseline_distance": baseline_distance,
        "community_diagnostics": community_diagnostics,
        "source_counts": source_counts,
        "attached_counts": attached_counts,
        "model_mode": model_mode,
    }


def administrative_positions(
    mesh_geometry: np.ndarray,
    admin_geometry: np.ndarray,
) -> np.ndarray:
    """Assign eligible meshes by centroid, then maximum overlap at boundaries."""
    centroids = shapely.centroid(mesh_geometry)
    positions = np.full(len(centroids), -1, dtype="int32")
    for position, polygon in enumerate(admin_geometry):
        unassigned = positions < 0
        if not unassigned.any():
            break
        positions[unassigned] = np.where(
            shapely.covers(polygon, centroids[unassigned]),
            position,
            -1,
        )
    for mesh_index in np.flatnonzero(positions < 0):
        overlaps = np.asarray(
            [
                float(shapely.area(shapely.intersection(mesh_geometry[mesh_index], polygon)))
                for polygon in admin_geometry
            ]
        )
        if overlaps.max() > 0:
            positions[mesh_index] = int(np.argmax(overlaps))
            continue
        distances = np.asarray(shapely.distance(admin_geometry, centroids[mesh_index]), dtype=float)
        nearest = int(np.argmin(distances))
        if distances[nearest] <= 0.003:
            positions[mesh_index] = nearest
    if np.any(positions < 0):
        raise RuntimeError(
            f"{np.count_nonzero(positions < 0):,} eligible meshes neither intersect nor lie "
            "within 0.003 degrees of an administrative polygon."
        )
    return positions


def build_table() -> tuple[pd.DataFrame, dict[str, object]]:
    """Aggregate mesh-allocated simulation consequences to 49 administrative units."""
    context = prepare_network_and_outcomes()
    admin = context["admin"]
    selected_mesh = context["selected_mesh"].copy()
    admin_position = administrative_positions(
        context["selected_mesh_geometry"], context["admin_geometry"]
    )
    community_position = selected_mesh["Community Position"].to_numpy(dtype="int32")
    population = selected_mesh["Total Population"].to_numpy(dtype=float)
    older_population = selected_mesh["Population Age 65+"].to_numpy(dtype=float)
    frequencies = context["frequencies"]
    loss_frequency = context["loss_frequency"]
    mean_excess = context["mean_excess"]

    service_excess = np.column_stack(
        [mean_excess[name][community_position] for name in service_loss.SERVICE_CLASSES]
    )
    finite_excess = np.isfinite(service_excess)
    service_excess_sum = np.nansum(service_excess, axis=1)
    service_excess_count = finite_excess.sum(axis=1)
    mesh_mean_excess = np.divide(
        service_excess_sum,
        service_excess_count,
        out=np.full(len(service_excess), np.nan, dtype=float),
        where=service_excess_count > 0,
    )

    rows: list[dict[str, object]] = []
    for position, admin_row in admin.reset_index(drop=True).iterrows():
        mask = admin_position == position
        eligible_population = float(population[mask].sum())
        eligible_older = float(older_population[mask].sum())
        row: dict[str, object] = {
            "Admin Area Code": str(admin_row["Municipality Code"]),
            "Municipality / Ward": priority_roads.MUNICIPALITY_ENGLISH_BY_CODE[
                str(admin_row["Municipality Code"])
            ],
            "Eligible Population": eligible_population,
            "Eligible Population Age 65+": eligible_older,
        }
        for scenario in ("Moderate", "Heavy", "Extreme"):
            mesh_frequency = frequencies[scenario][community_position[mask]]
            isolated_population = float(np.sum(population[mask] * mesh_frequency))
            row[f"{scenario} Isolation Frequency"] = (
                isolated_population / eligible_population
                if eligible_population > 0
                else np.nan
            )
            row[f"{scenario} Expected Isolated Population"] = isolated_population
        heavy_frequency = frequencies["Heavy"][community_position[mask]]
        row["Heavy Expected Isolated Population Age 65+"] = float(
            np.sum(older_population[mask] * heavy_frequency)
        )
        for service in service_loss.SERVICE_CLASSES:
            service_values = loss_frequency[service][community_position[mask]]
            evaluable = np.isfinite(service_values)
            expected_loss = float(
                np.sum(population[mask][evaluable] * service_values[evaluable])
            )
            row[f"Heavy {service} Loss Population (Baseline-Reachable)"] = expected_loss
        valid_excess = mask & np.isfinite(mesh_mean_excess)
        row["Heavy Population-Weighted Mean Excess Time (min)"] = (
            float(np.average(mesh_mean_excess[valid_excess], weights=population[valid_excess]))
            if valid_excess.any() and population[valid_excess].sum() > 0
            else np.nan
        )
        rows.append(row)

    table = pd.DataFrame(rows)
    columns = [
        "Admin Area Code",
        "Municipality / Ward",
        "Eligible Population",
        "Eligible Population Age 65+",
        "Moderate Isolation Frequency",
        "Moderate Expected Isolated Population",
        "Heavy Isolation Frequency",
        "Heavy Expected Isolated Population",
        "Extreme Isolation Frequency",
        "Extreme Expected Isolated Population",
        "Heavy Expected Isolated Population Age 65+",
        "Heavy Shelter Loss Population (Baseline-Reachable)",
        "Heavy Emergency water Loss Population (Baseline-Reachable)",
        "Heavy Fire service Loss Population (Baseline-Reachable)",
        "Heavy Municipal facility Loss Population (Baseline-Reachable)",
        "Heavy Population-Weighted Mean Excess Time (min)",
    ]
    table = table.loc[:, columns]
    if table.shape != (49, 16):
        raise RuntimeError(f"Expected a 49 × 16 table, found {table.shape}.")
    if table["Admin Area Code"].duplicated().any():
        raise RuntimeError("Administrative area codes must be unique.")
    if not np.isclose(
        table["Eligible Population"].sum(),
        context["community_diagnostics"]["Eligible Population"],
        rtol=0,
        atol=0.1,
    ):
        raise RuntimeError("Municipality population totals do not reconcile.")
    return table, context


def style_workbook(path: Path) -> None:
    """Apply grouped scientific-table formatting and review pagination."""
    workbook = load_workbook(path)
    worksheet = workbook[SHEET_NAME]
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "C2"
    worksheet.auto_filter.ref = f"A1:P{worksheet.max_row}"
    worksheet.sheet_view.zoomScale = 80
    worksheet.print_area = f"A1:P{worksheet.max_row}"
    worksheet.print_title_rows = "1:1"
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A3
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_margins = PageMargins(
        left=0.20, right=0.20, top=0.30, bottom=0.30, header=0.10, footer=0.10
    )

    header_fill = PatternFill("solid", fgColor="17365D")
    header_font = Font(name="Aptos", size=9, bold=True, color="FFFFFF")
    body_font = Font(name="Aptos", size=8.8, color="172033")
    subtle_border = Border(bottom=Side(style="thin", color="D0D5DD"))
    scenario_fills = {
        5: "E2F0EC",
        6: "E2F0EC",
        7: "FCE8D7",
        8: "FCE8D7",
        9: "F5DADB",
        10: "F5DADB",
    }
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[1].height = 48

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = subtle_border
        for column, color in scenario_fills.items():
            row[column - 1].fill = PatternFill("solid", fgColor=color)
        for column in (5, 7, 9):
            row[column - 1].number_format = "0.0%"
        for column in range(3, 16):
            if column not in (5, 7, 9):
                row[column - 1].number_format = "#,##0.0"
        row[15].number_format = "0.00"
        for cell in row[2:]:
            cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
        worksheet.row_dimensions[row[0].row].height = 27

    widths = {
        "A": 13,
        "B": 18,
        "C": 17,
        "D": 19,
        "E": 18,
        "F": 21,
        "G": 18,
        "H": 21,
        "I": 18,
        "J": 21,
        "K": 24,
        "L": 20,
        "M": 20,
        "N": 20,
        "O": 21,
        "P": 25,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width
    for column in ("E", "G", "I"):
        worksheet.conditional_formatting.add(
            f"{column}2:{column}{worksheet.max_row}",
            ColorScaleRule(
                start_type="min",
                start_color="FFFFFF",
                mid_type="percentile",
                mid_value=75,
                mid_color="FFEB84",
                end_type="max",
                end_color="F8696B",
            ),
        )

    excel_table = Table(displayName="MunicipalityServiceLoss", ref=f"A1:P{worksheet.max_row}")
    excel_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(excel_table)
    workbook.save(path)


def verify_workbook(path: Path) -> None:
    """Verify dimensions, numeric ranges, and absence of spreadsheet errors."""
    workbook = load_workbook(path, data_only=False)
    if workbook.sheetnames != [SHEET_NAME]:
        raise RuntimeError(f"Unexpected workbook sheets: {workbook.sheetnames}")
    worksheet = workbook[SHEET_NAME]
    if worksheet.max_row != 50 or worksheet.max_column != 16:
        raise RuntimeError(
            f"Expected 50 rows including the header and 16 columns; found "
            f"{worksheet.max_row} × {worksheet.max_column}."
        )
    error_tokens = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"}
    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value in error_tokens:
                raise RuntimeError(f"Spreadsheet error token in {cell.coordinate}: {cell.value}")
    for row in range(2, worksheet.max_row + 1):
        for column in (5, 7, 9):
            value = worksheet.cell(row, column).value
            if value is not None and not 0 <= float(value) <= 1:
                raise RuntimeError(f"Isolation frequency outside [0, 1] at row {row}.")


def main() -> None:
    table, context = build_table()
    OUT.parent.mkdir(parents=True, exist_ok=True)
    table.to_excel(OUT, index=False, sheet_name=SHEET_NAME, engine="openpyxl")
    style_workbook(OUT)
    verify_workbook(OUT)
    highest = table.sort_values("Heavy Isolation Frequency", ascending=False).iloc[0]
    print(f"Saved: {OUT.relative_to(ROOT)}")
    print(f"Rows: {len(table):,}; columns: {len(table.columns):,}")
    print(f"Eligible population: {table['Eligible Population'].sum():,.1f}")
    print(
        f"Highest Heavy isolation frequency: {highest['Municipality / Ward']} "
        f"({highest['Heavy Isolation Frequency']:.3f})"
    )
    print(f"Terrain-score construction: {context['model_mode']}")
    print("Workbook verification: passed")


if __name__ == "__main__":
    main()
