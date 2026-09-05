"""Apply only the five approved R2C10 synchronization parts; never write markup."""
from pathlib import Path
from zipfile import ZipFile
from copy import deepcopy
import argparse
import datetime as dt
import hashlib
import json
import os
import shutil
import difflib
from lxml import etree as E
from openpyxl import load_workbook
import table_threshold_baseline_and_rainfall_parameter_sensitivity as b5

ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT/'data/exp/revision/reviewer-2-comment-10'
BACKUPS = ROOT/'Rev/revision/.kila-backups'
NS = {'w':'http://schemas.openxmlformats.org/wordprocessingml/2006/main',
      'm':'http://schemas.openxmlformats.org/officeDocument/2006/math'}
BEFORE = ['Rainfall parameters','Matched road-evidence concordance','0.646','0.632–0.652',
          'Range across 15 combinations','93 matched road sections','Validation signal is stable']
AFTER = ['Rainfall parameters','Episode-weighted road-restriction correspondence','0.723','0.711–0.726',
         'Range across 15 combinations','10 physical episodes',
         'Supplementary dry-event correspondence; not rainfall-trigger validation']
HIGHLIGHTS = [
 ('Heavy rainfall isolated 12,000 residents, including 3,000 elderly.',
  'Heavy-scenario mean disconnection is 1,064 residents, including 577 aged 65+.'),
 ('Road disruption scores showed strong concordance with observed data.',
  'Ten dry-window restriction episodes provide supplementary ranking evidence.')]
ANCHOR = 'The geographic summary is descriptive and does not replace the community-level network unit.'
NOTE = ' Prefecture-wide totals are calculated from unrounded values; sums of the displayed municipality-level values can differ slightly because of rounding.'

def sha(p): return hashlib.sha256(p.read_bytes()).hexdigest()
def visible(p): return ''.join(p.xpath('.//w:t/text()',namespaces=NS))
def unique(seq):
    seq=list(seq)
    assert len(seq)==1,len(seq)
    return seq[0]

def patch_text(node,before,after):
    """Character edits preserve existing runs, run styles and all non-text XML."""
    nodes=node.xpath('.//w:t',namespaces=NS)
    full=''.join(n.text or '' for n in nodes)
    assert full.count(before)==1,(before,full)
    desired=full.replace(before,after,1)
    for tag,a,b,c,d in reversed(difflib.SequenceMatcher(None,full,desired,autojunk=False).get_opcodes()):
        if tag=='equal':continue
        spans=[]; pos=0
        for n in nodes:
            length=len(n.text or '');spans.append((n,pos,pos+length));pos+=length
        targets=[(n,s,e) for n,s,e in spans if e>a and s<b]
        if a==b:
            n,s,e=next((x for x in spans if x[1]<=a<x[2]),spans[-1])
            offset=a-s;n.text=(n.text or '')[:offset]+desired[c:d]+(n.text or '')[offset:]
        else:
            assert targets
            for i,(n,s,e) in enumerate(targets):
                val=n.text or '';lo=max(a-s,0);hi=min(b-s,len(val))
                n.text=val[:lo]+(desired[c:d] if i==0 else '')+val[hi:]
        for n in nodes:
            if (n.text or '').startswith(' ') or (n.text or '').endswith(' '):
                n.set('{http://www.w3.org/XML/1998/namespace}space','preserve')
    assert visible(node)==desired

def log(part,record):
    record.update(part=part,decision='KILA-D-20260905-020')
    path=OUT/f'part-{part:02d}.json';assert not path.exists()
    path.write_text(json.dumps(record,ensure_ascii=False,indent=2)+'\n')
    entry=f'\n## reviewer-2/comment-10#part-{part:02d}\n\n'
    entry+='- Approval: KILA-D-20260905-020; exact five-part proposal.\n'
    for k,v in record.items():
        if k in ('before','after'):
            entry+=f'\n### {k.title()}\n\n'+(v if isinstance(v,str) else json.dumps(v,ensure_ascii=False))+'\n'
        else:entry+=f'- {k}: {v}\n'
    with (ROOT/'Rev/docs/revisionchanges.md').open('a') as f:f.write(entry)
    print(json.dumps(record,ensure_ascii=False))

def doc_part(part):
    dest=ROOT/('Rev/revision/Highlights.docx' if part<3 else 'Rev/revision/Appendix.docx')
    src=ROOT/'article/deliverables/Highlights.docx' if part==1 else dest
    if part==1:assert not dest.exists()
    oldhash=sha(src)
    with ZipFile(src) as z:
        infos=z.infolist(); original={i.filename:z.read(i.filename) for i in infos}
    root=E.fromstring(original['word/document.xml']); baseline=deepcopy(root)
    if part in (1,2):
        before,after=HIGHLIGHTS[part-1]
        p=unique(p for p in root.xpath('.//w:body/w:p',namespaces=NS) if visible(p)==before)
        patch_text(p,before,after)
    elif part==4:
        t=unique(t for t in root.xpath('.//w:tbl',namespaces=NS) if visible(t).startswith('Table B5. Baseline-Threshold'))
        row=unique(r for r in t.findall('w:tr',NS) if len(r.findall('w:tc',NS))==7 and visible(r.findall('w:tc',NS)[1])==BEFORE[1])
        cells=row.findall('w:tc',NS);assert [visible(c) for c in cells]==BEFORE
        for cell,b,a in zip(cells,BEFORE,AFTER):
            if b!=a:patch_text(cell,b,a)
        before,after=BEFORE,AFTER
    elif part==5:
        before,after=ANCHOR,ANCHOR+NOTE
        p=unique(p for p in root.xpath('.//w:body/w:p',namespaces=NS) if ANCHOR in visible(p))
        patch_text(p,before,after)
    else:raise ValueError(part)
    # Outside approved text nodes, styles, formulas, fields and structure cannot change.
    for tree in (root,baseline):
        for n in tree.xpath('.//w:t',namespaces=NS):
            n.text='';n.attrib.pop('{http://www.w3.org/XML/1998/namespace}space',None)
    assert E.tostring(root)==E.tostring(baseline),'Non-text XML changed'
    # Repeat deterministic patch on a fresh tree after the structural comparison.
    root2=E.fromstring(original['word/document.xml'])
    if part in (1,2):patch_text(unique(p for p in root2.xpath('.//w:body/w:p',namespaces=NS) if visible(p)==before),before,after)
    elif part==5:patch_text(unique(p for p in root2.xpath('.//w:body/w:p',namespaces=NS) if ANCHOR in visible(p)),before,after)
    else:
        t=unique(t for t in root2.xpath('.//w:tbl',namespaces=NS) if visible(t).startswith('Table B5. Baseline-Threshold'))
        row=unique(r for r in t.findall('w:tr',NS) if len(r.findall('w:tc',NS))==7 and visible(r.findall('w:tc',NS)[1])==BEFORE[1])
        for cell,b,a in zip(row.findall('w:tc',NS),BEFORE,AFTER):
            if b!=a:patch_text(cell,b,a)
    updated=E.tostring(root2,xml_declaration=True,encoding='UTF-8',standalone=True)
    staged=OUT/f'part-{part:02d}.docx'
    with ZipFile(staged,'w') as z:
        for i in infos:z.writestr(i,updated if i.filename=='word/document.xml' else original[i.filename])
    with ZipFile(staged) as z:
        assert z.testzip() is None
        assert all(z.read(k)==v for k,v in original.items() if k!='word/document.xml')
    backup=BACKUPS/f'{src.stem}.before-r2c10-part{part:02d}.{dt.datetime.now().strftime("%Y%m%dT%H%M%S%f")}.docx'
    shutil.copy2(src,backup);assert sha(backup)==oldhash
    assert sha(src)==oldhash
    os.replace(staged,dest)
    log(part,dict(target=str(dest.relative_to(ROOT)),before=before,after=after,before_sha256=oldhash,after_sha256=sha(dest),
                  backup=str(backup.relative_to(ROOT)),verification='Only approved w:t text changed; other XML structure/styles and all other package members byte-preserved; direct edit, no markup touched'))

def workbook_part():
    dest=b5.OUT;oldhash=sha(dest)
    staged=OUT/'B5.generated.xlsx';b5.OUT=staged;b5.build_workbook();b5.verify_workbook()
    old=load_workbook(dest);new=load_workbook(staged)
    allowed={(b5.SUMMARY_SHEET,c) for c in ('B12','F12','G12')}
    allowed|={(b5.EVIDENCE_SHEET,f'{c}{r}') for r in (21,22,23) for c in ('B','C','D','F','G','H')}
    for s in old:
        for row in s:
            for cell in row:
                if (s.title,cell.coordinate) not in allowed:
                    assert cell.value==new[s.title][cell.coordinate].value,(s.title,cell.coordinate,cell.value,new[s.title][cell.coordinate].value)
    backup=BACKUPS/f'B5.before-r2c10.{dt.datetime.now().strftime("%Y%m%dT%H%M%S%f")}.xlsx'
    shutil.copy2(dest,backup)
    changes=[]
    for sheet,coordinate in sorted(allowed):
        c=old[sheet][coordinate];val=new[sheet][coordinate].value
        if c.value!=val:changes.append([sheet,coordinate,c.value,val]);c.value=val
    old[b5.SUMMARY_SHEET]['C12'].comment=deepcopy(new[b5.SUMMARY_SHEET]['C12'].comment)
    candidate=OUT/'B5.corrected.xlsx';old.save(candidate)
    check=load_workbook(candidate)
    for s in old:
        for row in s:
            for c in row:
                assert check[s.title][c.coordinate].value==c.value
    # Preserve every cell's style and every unrelated value/formula from source.
    original=load_workbook(backup)
    for s in original:
        for row in s:
            for c in row:
                n=check[s.title][c.coordinate]
                assert c._style==n._style
                if (s.title,c.coordinate) not in allowed:assert c.value==n.value
    assert sha(dest)==oldhash;os.replace(candidate,dest)
    b5.OUT=dest;b5.verify_workbook()
    log(3,dict(target=str(dest.relative_to(ROOT)),before='Legacy section-weighted B5 correspondence row and provenance',
               after=AFTER,changed_cells=changes,before_sha256=oldhash,after_sha256=sha(dest),backup=str(backup.relative_to(ROOT)),
               verification='Every unrelated value/formula and every cell style preserved; current producer agrees with current workbook outside approved cells; preview pending'))

if __name__=='__main__':
    p=argparse.ArgumentParser();p.add_argument('part',type=int,choices=range(1,6));a=p.parse_args()
    OUT.mkdir(parents=True,exist_ok=True);BACKUPS.mkdir(parents=True,exist_ok=True)
    assert not (OUT/f'part-{a.part:02d}.json').exists(),'Part already applied'
    if a.part==3:workbook_part()
    else:doc_part(a.part)
