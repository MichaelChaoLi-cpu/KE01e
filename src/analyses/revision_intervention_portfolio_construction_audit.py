#!/usr/bin/env python3
"""Audit the implemented relationship among manuscript Equations 15–17."""

from __future__ import annotations

import ast
import hashlib
import json
from pathlib import Path
from typing import Any

import numpy as np
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
PRODUCTION_SOURCE = (
    ROOT / "src/analyses/figure_intervention_priorities_and_budgeted_benefits.py"
)
TABLE_SOURCE = ROOT / "src/analyses/table_intervention_portfolios_and_robustness.py"
PORTFOLIO_WORKBOOK = ROOT / "data/results/tables/Table_intervention_portfolios.xlsx"
OUTPUT_DIR = ROOT / "data/exp/revision/reviewer-2-comment-9"


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def literal_assignment(tree: ast.Module, name: str) -> Any:
    for node in tree.body:
        if isinstance(node, ast.Assign):
            if any(isinstance(target, ast.Name) and target.id == name for target in node.targets):
                return ast.literal_eval(node.value)
    raise RuntimeError(f"Assignment {name} was not found")


def function_node(tree: ast.Module, name: str) -> ast.FunctionDef:
    for node in tree.body:
        if isinstance(node, ast.FunctionDef) and node.name == name:
            return node
    raise RuntimeError(f"Function {name} was not found")


def assignment_texts(function: ast.FunctionDef) -> dict[str, str]:
    found: dict[str, str] = {}
    for node in ast.walk(function):
        if not isinstance(node, ast.Assign) or len(node.targets) != 1:
            continue
        target = node.targets[0]
        if isinstance(target, ast.Name):
            found[target.id] = ast.unparse(node.value)
    return found


def call_names(function: ast.FunctionDef) -> list[str]:
    names: list[str] = []
    for node in ast.walk(function):
        if not isinstance(node, ast.Call):
            continue
        if isinstance(node.func, ast.Name):
            names.append(node.func.id)
        elif isinstance(node.func, ast.Attribute):
            names.append(ast.unparse(node.func))
    return sorted(set(names))


def run_select_under_budget(function: ast.FunctionDef) -> tuple[list[int], float]:
    module = ast.Module(body=[function], type_ignores=[])
    ast.fix_missing_locations(module)
    namespace: dict[str, Any] = {"np": np}
    exec(compile(module, str(PRODUCTION_SOURCE), "exec"), namespace)
    order = np.asarray([2, 0, 3, 1], dtype="int32")
    costs = np.asarray([5.0, 4.0, 8.0, 2.0])
    selected, spent = namespace["select_under_budget"](order, costs, 7.0)
    return selected.astype(int).tolist(), float(spent)


def workbook_audit() -> dict[str, Any]:
    workbook = load_workbook(PORTFOLIO_WORKBOOK, data_only=True, read_only=True)
    sheet = workbook["Portfolio Robustness"]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
    rows = [dict(zip(headers, values, strict=True)) for values in sheet.iter_rows(min_row=3, values_only=True)]
    settings = sorted({str(row["Sensitivity Setting"]) for row in rows})
    budgets = sorted({float(row["Budget (Relative Planning Units)"]) for row in rows})
    central_max = next(
        row
        for row in rows
        if row["Sensitivity Setting"] == "Central"
        and float(row["Budget (Relative Planning Units)"]) == max(budgets)
    )
    return {
        "data_rows": len(rows),
        "settings": settings,
        "budget_count": len(budgets),
        "maximum_budget": max(budgets),
        "central_maximum_selected_roads": int(central_max["Selected Road Count"]),
        "central_maximum_realized_cost": float(
            central_max["Realized Cost (Relative Planning Units)"]
        ),
    }


def main() -> None:
    source_text = PRODUCTION_SOURCE.read_text(encoding="utf-8")
    tree = ast.parse(source_text)
    main_node = function_node(tree, "main")
    selector_node = function_node(tree, "select_under_budget")
    assignments = assignment_texts(main_node)
    calls = call_names(main_node)

    candidate_count = literal_assignment(tree, "PORTFOLIO_CANDIDATE_COUNT")
    budget_count = literal_assignment(tree, "PORTFOLIO_BUDGET_COUNT")
    selected, spent = run_select_under_budget(selector_node)
    workbook = workbook_audit()

    checks = {
        "candidate_count_is_150": candidate_count == 150,
        "budget_count_is_7": budget_count == 7,
        "candidate_slice_uses_priority_order": assignments.get("portfolio_positions")
        == "priority_order[:PORTFOLIO_CANDIDATE_COUNT]",
        "maximum_budget_uses_first_100_central_costs": assignments.get("max_budget")
        == "float(base_cost[portfolio_positions[:100]].sum())",
        "budgets_are_evenly_spaced": assignments.get("budgets")
        == "np.linspace(0.0, max_budget, PORTFOLIO_BUDGET_COUNT)",
        "synthetic_ordered_skip_selection": selected == [0, 3] and abs(spent - 7.0) < 1e-12,
        "selector_called_by_production_main": "select_under_budget" in calls,
        "network_evaluation_occurs_after_selection": "cached_intervention_frequency" in calls,
        "no_named_combinatorial_solver_call": not any(
            token in name.lower()
            for name in calls
            for token in ("milp", "linprog", "optimize", "knapsack", "pulp", "ortools")
        ),
        "formal_table_has_21_rows": workbook["data_rows"] == 21,
        "formal_table_has_three_settings": workbook["settings"]
        == ["Central", "Conservative", "Optimistic"],
        "formal_table_has_seven_budgets": workbook["budget_count"] == 7,
        "central_maximum_has_100_roads": workbook["central_maximum_selected_roads"] == 100,
    }
    if not all(checks.values()):
        failures = [name for name, passed in checks.items() if not passed]
        raise RuntimeError(f"Portfolio-construction audit failed: {failures}")

    decision = {
        "reviewer_comment": "reviewer-2/comment-9",
        "implemented_procedure": "greedy rank-and-pack screening",
        "equation_15_role": "post-selection evaluation estimand and budget-feasibility condition",
        "equations_16_17_role": "construct the candidate screening order",
        "explicitly_optimizes_equation_15": False,
        "candidate_count": candidate_count,
        "budget_count": budget_count,
        "synthetic_selected_positions": selected,
        "synthetic_realized_cost": spent,
        "workbook": workbook,
        "checks": checks,
        "input_sha256": {
            str(PRODUCTION_SOURCE.relative_to(ROOT)): sha256(PRODUCTION_SOURCE),
            str(TABLE_SOURCE.relative_to(ROOT)): sha256(TABLE_SOURCE),
            str(PORTFOLIO_WORKBOOK.relative_to(ROOT)): sha256(PORTFOLIO_WORKBOOK),
        },
        "requires_production_recalculation": False,
    }

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUTPUT_DIR / "decision.json").write_text(
        json.dumps(decision, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )
    report = f"""# Reviewer 2 Comment 9 — Portfolio-Construction Audit

- Implemented procedure: greedy rank-and-pack screening.
- Equation 15 role: post-selection evaluation estimand and budget-feasibility condition.
- Equations 16–17 role: construct the candidate screening order.
- Explicit optimization of Equation 15: no.
- Planning candidate set: first {candidate_count} roads in descending priority-score order.
- Budgets: {budget_count} evenly spaced values from zero to the Central cost of the first 100 ranked candidates.
- Maximum budget in the formal table: {workbook['maximum_budget']:.3f} relative planning units.
- Central maximum-budget portfolio: {workbook['central_maximum_selected_roads']} roads at realized cost {workbook['central_maximum_realized_cost']:.3f}.
- Synthetic selector test: order `[2, 0, 3, 1]`, costs `[5, 4, 8, 2]`, budget `7` selects `{selected}` and spends `{spent:.1f}`, confirming skip-if-unaffordable ordered scanning.
- Production recalculation required: no.

The selected portfolio is evaluated after construction by applying the assumed
action effects to closure propensities and rerunning the network simulation.
The implementation does not search all feasible combinations and therefore
must not be described as maximizing Equation 15 or producing an optimum.
"""
    (OUTPUT_DIR / "audit_report.md").write_text(report, encoding="utf-8")
    print(json.dumps(decision, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
