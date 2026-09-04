#!/usr/bin/env python3
"""Propagate approved slope-to-road boundary settings through network outcomes.

The central, strict, and permissive transfer scores come from the validated
Reviewer 2 Comment 5 sensitivity run. Every setting uses the canonical 85th
and 99.5th positive-score quantiles, five fixed seeds, and 1,000 draws per seed.
All caches and outputs are revision-only.
"""

from __future__ import annotations

import hashlib
import json
import os
from pathlib import Path

os.environ.pop("PROJ_LIB", None)
os.environ.pop("PROJ_DATA", None)

import numpy as np
import pandas as pd
from scipy.sparse import coo_matrix
from scipy.sparse.csgraph import connected_components
import shapely

import figure_basic_service_reachability_loss as service
import figure_community_isolation_frequency_and_exposed_population as isolation
import figure_road_disruption_exposure_and_observed_restriction_evidence as road_exposure
import revision_rainfall_parameter_sensitivity as rainfall_sensitivity


ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "data/exp/revision/reviewer-2-comment-5"
SCORE_ARCHIVE = OUT / "joint_boundary_scenario_scores.npz"
UPSTREAM_DECISION = OUT / "decision.json"
REFERENCE_TABLE = ROOT / "data/results/tables/Table_municipality_isolation_and_service_loss_summary.xlsx"
CURRENT_REFERENCE_DECISION = ROOT / "data/exp/revision/reviewer-2-comment-4/decision.json"
SETTINGS = ("central", "strict_joint", "permissive_joint")
SCENARIOS = ("Moderate", "Heavy", "Extreme")
DECISION_RECORD = "KILA-D-20260903-014"


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def setup_variant(
    roads: pd.DataFrame,
    scores: dict[str, np.ndarray],
    admin_union: object,
) -> dict[str, object]:
    heavy_lower = isolation.positive_score_quantile(
        scores["Heavy"], isolation.CANDIDATE_QUANTILE
    )
    heavy_upper = isolation.positive_score_quantile(
        scores["Heavy"], isolation.UPPER_MAPPING_QUANTILE
    )
    candidate_mask = np.isfinite(scores["Heavy"]) & (scores["Heavy"] >= heavy_lower)
    candidate_positions = np.flatnonzero(candidate_mask)
    candidate_ids = roads.loc[candidate_mask, "Road Section ID"].reset_index(drop=True)
    candidate_position = pd.Series(
        np.arange(len(candidate_ids), dtype="int32"), index=candidate_ids
    )

    nodes = pd.read_parquet(
        isolation.NODE_PATH,
        columns=["Network Node ID", "Network Component ID", "Geometry"],
    )
    node_geometry = road_exposure.decode_geometry(nodes.pop("Geometry"))
    node_index = pd.Index(nodes["Network Node ID"])
    edges = pd.read_parquet(
        isolation.EDGE_PATH,
        columns=[
            "Road Section ID",
            "From Node ID",
            "To Node ID",
            "Network Component ID",
            "Emergency Route Membership",
            "Baseline Edge Travel Time (min)",
            "Network Analysis Eligible",
        ],
    )
    edges = edges.loc[edges["Network Analysis Eligible"]].reset_index(drop=True)
    edge_u = node_index.get_indexer(edges["From Node ID"])
    edge_v = node_index.get_indexer(edges["To Node ID"])
    if np.any(edge_u < 0) or np.any(edge_v < 0):
        raise RuntimeError("Road edges reference missing network nodes.")
    edge_candidate = edges["Road Section ID"].isin(candidate_ids).to_numpy()
    stable_u, stable_v = edge_u[~edge_candidate], edge_v[~edge_candidate]
    stable_graph = coo_matrix(
        (
            np.ones(len(stable_u) * 2, dtype="uint8"),
            (np.concatenate([stable_u, stable_v]), np.concatenate([stable_v, stable_u])),
        ),
        shape=(len(nodes), len(nodes)),
    ).tocsr()
    root_count, stable_labels = connected_components(
        stable_graph, directed=False, return_labels=True
    )
    stable_labels = stable_labels.astype("int32")

    candidate_u = stable_labels[edge_u[edge_candidate]]
    candidate_v = stable_labels[edge_v[edge_candidate]]
    candidate_edge_section = (
        edges.loc[edge_candidate, "Road Section ID"]
        .map(candidate_position)
        .to_numpy(dtype="int32")
    )
    candidate_edge_time = edges.loc[
        edge_candidate, "Baseline Edge Travel Time (min)"
    ].to_numpy(dtype="float64")
    between_root = candidate_u != candidate_v
    network_u = candidate_u[between_root]
    network_v = candidate_v[between_root]
    network_edge_section = candidate_edge_section[between_root]
    pair_reduction = service.prepare_pair_reduction(
        network_u,
        network_v,
        network_edge_section,
        candidate_edge_time[between_root],
        root_count,
    )

    targets, target_network_components = isolation.external_target_definitions(
        nodes, node_geometry, stable_labels, edges, edge_u, edge_v, admin_union
    )
    (
        community,
        attachment_community,
        attachment_root,
        diagnostics,
        selected_mesh,
        _,
    ) = isolation.build_baseline_communities(
        nodes, node_geometry, stable_labels, target_network_components
    )
    service_geometry, service_source_counts = service.service_geometries()
    service_roots, _, service_attached_counts = service.attach_services_to_roots(
        service_geometry, node_geometry, stable_labels
    )
    for service_class in service.SERVICE_CLASSES:
        if len(service_roots[service_class]) == 0:
            raise RuntimeError(f"No {service_class} features attach under this setting.")
    return {
        "heavy_lower": heavy_lower,
        "heavy_upper": heavy_upper,
        "candidate_positions": candidate_positions,
        "candidate_ids": candidate_ids,
        "candidate_u": network_u,
        "candidate_v": network_v,
        "candidate_edge_section": network_edge_section,
        "pair_reduction": pair_reduction,
        "root_count": int(root_count),
        "target_roots": targets[isolation.PRIMARY_TARGET_NAME],
        "community": community,
        "attachment_community": attachment_community,
        "attachment_root": attachment_root,
        "diagnostics": diagnostics,
        "selected_mesh": selected_mesh,
        "service_roots": service_roots,
        "service_source_counts": service_source_counts,
        "service_attached_counts": service_attached_counts,
        "candidate_network_edges": int(len(network_u)),
    }


def reference_totals() -> dict[str, float]:
    from openpyxl import load_workbook

    worksheet = load_workbook(REFERENCE_TABLE, data_only=True, read_only=True).active
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=2, max_row=2))]
    rows = list(worksheet.iter_rows(min_row=3, values_only=True))
    wanted = {
        "Moderate": ("Moderate Expected Isolated Population",),
        "Heavy": ("Heavy Expected Isolated Population",),
        "Extreme": ("Extreme Expected Isolated Population",),
        "Heavy age 65+": ("Heavy Expected Isolated Population Age 65+",),
        "Shelter": (
            "Heavy Shelter Loss Population (Any Same-Class Facility)",
            "Heavy Shelter Loss Population (Baseline-Reachable)",
        ),
        "Emergency water": (
            "Heavy Emergency-Water Sensitivity Loss Population (Any of 10/36 Geolocated Facilities)",
            "Heavy Emergency-Water Sensitivity Loss Population (10/36 Geolocated)",
        ),
        "Fire service": (
            "Heavy Fire service Loss Population (Any Same-Class Facility)",
            "Heavy Fire service Loss Population (Baseline-Reachable)",
        ),
        "Municipal facility": (
            "Heavy Municipal facility Loss Population (Any Same-Class Facility)",
            "Heavy Municipal facility Loss Population (Baseline-Reachable)",
        ),
    }
    totals: dict[str, float] = {}
    for key, aliases in wanted.items():
        column = next((alias for alias in aliases if alias in headers), None)
        if column is None:
            raise ValueError(f"None of the reference-table columns are present for {key}: {aliases}")
        totals[key] = float(sum(float(row[headers.index(column)] or 0) for row in rows))
    return totals


def main() -> None:
    required = [
        SCORE_ARCHIVE,
        UPSTREAM_DECISION,
        REFERENCE_TABLE,
        CURRENT_REFERENCE_DECISION,
        isolation.ROAD_PATH,
        isolation.EDGE_PATH,
        isolation.NODE_PATH,
        isolation.MESH_PATH,
        isolation.GROUP_PATH,
        service.DESIGNATED_SHELTER_PATH,
        service.EVACUATION_SITE_PATH,
        service.CURRENT_SHELTER_PATH,
        service.WATER_PATH,
        service.FIRE_PATH,
        service.MUNICIPAL_PATH,
        Path(isolation.__file__),
        Path(service.__file__),
        Path(__file__),
    ]
    missing = [str(path.relative_to(ROOT)) for path in required if not path.exists()]
    if missing:
        raise FileNotFoundError(f"Missing frozen inputs: {missing}")
    upstream = json.loads(UPSTREAM_DECISION.read_text())
    if upstream.get("classification") != "material":
        raise RuntimeError("Downstream propagation is permitted only after material sensitivity.")

    isolation.SIMULATION_CACHE_DIR = OUT / "downstream_cache/community_isolation"
    service.SERVICE_LOSS_CACHE_DIR = OUT / "downstream_cache/service_loss"
    admin = pd.read_parquet(isolation.ADMIN_PATH, columns=["Geometry"])
    admin_geometry = road_exposure.decode_geometry(admin.pop("Geometry"))
    admin_union = shapely.union_all(admin_geometry)
    roads = pd.read_parquet(
        isolation.ROAD_PATH,
        columns=["Road Section ID", "Network Analysis Eligible"],
    )
    roads = roads.loc[roads["Network Analysis Eligible"]].reset_index(drop=True)
    archive = np.load(SCORE_ARCHIVE, allow_pickle=False)
    score_sets = {
        setting: {
            scenario: archive[f"{setting}__{scenario}"].astype("float32")
            for scenario in SCENARIOS
        }
        for setting in SETTINGS
    }
    if any(len(score_sets[s]["Heavy"]) != len(roads) for s in SETTINGS):
        raise RuntimeError("Stored road-score arrays do not match the eligible-road universe.")

    network_rows: list[dict[str, object]] = []
    service_rows: list[dict[str, object]] = []
    frequency_arrays: dict[str, np.ndarray] = {}
    service_arrays: dict[str, np.ndarray] = {}
    central_structure_equal = False
    for setting in SETTINGS:
        print(f"Preparing downstream network: {setting}", flush=True)
        scores = score_sets[setting]
        variant = setup_variant(roads, scores, admin_union)
        if setting == "central":
            canonical = rainfall_sensitivity.setup_network(
                roads, scores["Heavy"], admin_union
            )
            central_structure_equal = bool(
                variant["candidate_ids"].equals(canonical["candidate_ids"])
                and variant["root_count"] == canonical["root_count"]
                and np.array_equal(variant["candidate_u"], canonical["candidate_u"])
                and np.array_equal(variant["candidate_v"], canonical["candidate_v"])
                and np.array_equal(
                    variant["candidate_edge_section"],
                    canonical["candidate_edge_section"],
                )
                and np.array_equal(variant["target_roots"], canonical["target_roots"])
                and np.array_equal(
                    variant["attachment_community"],
                    canonical["attachment_community"],
                )
                and np.array_equal(
                    variant["attachment_root"], canonical["attachment_root"]
                )
                and np.allclose(
                    variant["community"]["Total_Population"],
                    canonical["community"]["Total_Population"],
                )
            )
            if not central_structure_equal:
                raise RuntimeError("Central network structure does not match the canonical helper.")
        candidate_positions = np.asarray(variant["candidate_positions"], dtype=int)
        total_population = variant["community"]["Total_Population"].to_numpy(dtype=float)
        older_population = variant["community"]["Population_Age_65"].to_numpy(dtype=float)
        heavy_propensity: np.ndarray | None = None
        for scenario in SCENARIOS:
            candidate_score = scores[scenario][candidate_positions]
            propensity = isolation.closure_propensity(
                candidate_score,
                float(variant["heavy_lower"]),
                float(variant["heavy_upper"]),
            )
            if scenario == "Heavy":
                heavy_propensity = propensity
            replicate = []
            for seed_index, seed in enumerate(isolation.REPLICATE_SEEDS):
                replicate.append(
                    isolation.cached_isolation(
                        f"r2c5_{setting}_{scenario.lower()}_seed_{seed}_m1000",
                        variant["candidate_u"],
                        variant["candidate_v"],
                        variant["candidate_edge_section"],
                        propensity,
                        variant["root_count"],
                        variant["target_roots"],
                        variant["attachment_community"],
                        variant["attachment_root"],
                        len(variant["community"]),
                        seed,
                        draws=1_000,
                        report_progress=seed_index == 0,
                    )
                )
            seed_total = np.asarray(
                [float(np.sum(total_population * value)) for value in replicate]
            )
            seed_older = np.asarray(
                [float(np.sum(older_population * value)) for value in replicate]
            )
            mean_frequency = np.mean(np.vstack(replicate), axis=0).astype("float32")
            frequency_arrays[f"{setting}__{scenario}"] = mean_frequency
            network_rows.append(
                {
                    "transfer_setting": setting,
                    "scenario": scenario,
                    "positive_road_sections": int(np.sum(scores["Heavy"] > 0)),
                    "candidate_road_sections": int(len(candidate_positions)),
                    "candidate_network_edges": int(variant["candidate_network_edges"]),
                    "heavy_candidate_lower_score": float(variant["heavy_lower"]),
                    "heavy_mapping_upper_score": float(variant["heavy_upper"]),
                    "nonzero_propensity_sections": int(np.count_nonzero(propensity)),
                    "community_count": int(len(variant["community"])),
                    "eligible_population": float(variant["diagnostics"]["Eligible Population"]),
                    "unresolved_population": float(variant["diagnostics"]["Unresolved Population"]),
                    "expected_isolated_population_mean": float(seed_total.mean()),
                    "expected_isolated_population_min": float(seed_total.min()),
                    "expected_isolated_population_max": float(seed_total.max()),
                    "expected_isolated_population_sd": float(seed_total.std(ddof=1)),
                    "expected_isolated_population_age65_mean": float(seed_older.mean()),
                }
            )
        if heavy_propensity is None:
            raise RuntimeError("Heavy propensity was not constructed.")

        print(f"Simulating Heavy service loss: {setting}", flush=True)
        loss, _, _ = service.cached_service_loss(
            heavy_propensity,
            variant["pair_reduction"],
            variant["root_count"],
            variant["service_roots"],
            variant["attachment_community"],
            variant["attachment_root"],
            len(variant["community"]),
            cache_tag=f"r2c5_{setting}",
        )
        seed_results = [
            service._cached_service_loss_seed(
                heavy_propensity,
                variant["pair_reduction"],
                variant["root_count"],
                variant["service_roots"],
                variant["attachment_community"],
                variant["attachment_root"],
                len(variant["community"]),
                seed,
                f"r2c5_{setting}",
            )
            for seed in isolation.REPLICATE_SEEDS
        ]
        for service_class in service.SERVICE_CLASSES:
            seed_loss = np.asarray(
                [
                    float(np.nansum(result[0][service_class] * total_population))
                    for result in seed_results
                ]
            )
            service_arrays[f"{setting}__{service_class}"] = loss[service_class]
            resolved, source_total = variant["service_source_counts"][service_class]
            service_rows.append(
                {
                    "transfer_setting": setting,
                    "service_class": service_class,
                    "resolved_source_features": int(resolved),
                    "source_features": int(source_total),
                    "attached_service_features": int(
                        variant["service_attached_counts"][service_class]
                    ),
                    "expected_service_loss_population_mean": float(seed_loss.mean()),
                    "expected_service_loss_population_min": float(seed_loss.min()),
                    "expected_service_loss_population_max": float(seed_loss.max()),
                    "expected_service_loss_population_sd": float(seed_loss.std(ddof=1)),
                    "interpretation": (
                        "conditional sensitivity for geolocated records"
                        if service_class == "Emergency water"
                        else "baseline-reachable service loss"
                    ),
                }
            )

    network_frame = pd.DataFrame(network_rows)
    service_frame = pd.DataFrame(service_rows)
    formal_references = reference_totals()
    central_network = network_frame.loc[network_frame["transfer_setting"].eq("central")]
    formal_table_differences = {
        scenario: abs(
            float(
                central_network.loc[
                    central_network["scenario"].eq(scenario),
                    "expected_isolated_population_mean",
                ].iloc[0]
            )
            - formal_references[scenario]
        )
        for scenario in SCENARIOS
    }
    formal_table_differences["Heavy age 65+"] = abs(
        float(
            central_network.loc[
                central_network["scenario"].eq("Heavy"),
                "expected_isolated_population_age65_mean",
            ].iloc[0]
        )
        - formal_references["Heavy age 65+"]
    )
    central_service = service_frame.loc[service_frame["transfer_setting"].eq("central")]
    for service_class in service.SERVICE_CLASSES:
        formal_table_differences[service_class] = abs(
            float(
                central_service.loc[
                    central_service["service_class"].eq(service_class),
                    "expected_service_loss_population_mean",
                ].iloc[0]
            )
            - formal_references[service_class]
        )
    current_reference = json.loads(CURRENT_REFERENCE_DECISION.read_text())
    current_heavy_reference = float(
        current_reference["corrected_central_expected_isolated_population"]
    )
    central_heavy = float(
        central_network.loc[
            central_network["scenario"].eq("Heavy"),
            "expected_isolated_population_mean",
        ].iloc[0]
    )
    current_reproduction_error = abs(central_heavy - current_heavy_reference)
    if current_reproduction_error > 1e-5 or not central_structure_equal:
        raise RuntimeError(
            "Central downstream reproduction failed: "
            + json.dumps(
                {
                    "current_heavy_error": current_reproduction_error,
                    "network_structure_equal": central_structure_equal,
                },
                ensure_ascii=False,
            )
        )

    network_frame.to_csv(OUT / "downstream_network_sensitivity.csv", index=False)
    service_frame.to_csv(OUT / "downstream_service_sensitivity.csv", index=False)
    pd.DataFrame(
        [
            {
                "outcome": outcome,
                "absolute_difference_people": difference,
                "status": "stale-formal-output-audit",
            }
            for outcome, difference in formal_table_differences.items()
        ]
    ).to_csv(OUT / "downstream_formal_table_difference_audit.csv", index=False)
    np.savez_compressed(
        OUT / "downstream_community_frequencies.npz", **frequency_arrays
    )
    np.savez_compressed(
        OUT / "downstream_service_loss_frequencies.npz", **service_arrays
    )
    input_paths = list(dict.fromkeys(required))
    pd.DataFrame(
        [
            {
                "path": str(path.relative_to(ROOT)),
                "sha256": sha256(path),
                "bytes": path.stat().st_size,
            }
            for path in input_paths
        ]
    ).to_csv(OUT / "downstream_input_hashes.csv", index=False)

    heavy = network_frame.loc[network_frame["scenario"].eq("Heavy")].set_index(
        "transfer_setting"
    )
    decision = {
        "reviewer_unit": "reviewer-2/comment-5",
        "decision_record": DECISION_RECORD,
        "settings": list(SETTINGS),
        "scenarios": list(SCENARIOS),
        "draws_per_seed": 1_000,
        "seeds": list(isolation.REPLICATE_SEEDS),
        "candidate_quantile": isolation.CANDIDATE_QUANTILE,
        "mapping_upper_quantile": isolation.UPPER_MAPPING_QUANTILE,
        "central_network_structure_matches_canonical_helper": central_structure_equal,
        "current_central_heavy_reference": current_heavy_reference,
        "current_central_heavy_reproduction_abs_error_people": current_reproduction_error,
        "formal_table_max_abs_difference_people": max(formal_table_differences.values()),
        "formal_table_difference_status": (
            "The formal municipality workbook predates the current corrected central "
            "network result and is retained as a discrepancy audit, not a reproduction target."
        ),
        "central_heavy_expected_isolated_population": float(
            heavy.loc["central", "expected_isolated_population_mean"]
        ),
        "strict_heavy_expected_isolated_population": float(
            heavy.loc["strict_joint", "expected_isolated_population_mean"]
        ),
        "permissive_heavy_expected_isolated_population": float(
            heavy.loc["permissive_joint", "expected_isolated_population_mean"]
        ),
        "strict_to_central_heavy_ratio": float(
            heavy.loc["strict_joint", "expected_isolated_population_mean"]
            / heavy.loc["central", "expected_isolated_population_mean"]
        ),
        "permissive_to_central_heavy_ratio": float(
            heavy.loc["permissive_joint", "expected_isolated_population_mean"]
            / heavy.loc["central", "expected_isolated_population_mean"]
        ),
        "interpretation": (
            "Downstream consequences are conditional on the transfer specification; "
            "the central setting remains a transparent reference, not a calibrated optimum."
        ),
    }
    (OUT / "downstream_decision.json").write_text(
        json.dumps(decision, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    report = [
        "# Downstream slope-to-road boundary propagation",
        "",
        f"- Decision record: `{DECISION_RECORD}`",
        f"- Central network structure matches the canonical helper: {'yes' if central_structure_equal else 'no'}",
        f"- Current corrected Heavy central-result reproduction error: {current_reproduction_error:.3g} people",
        f"- Maximum difference from the stale formal municipality workbook: {max(formal_table_differences.values()):.3f} people; this is carried forward to the later whole-manuscript numerical-consistency correction.",
        "- Each transfer setting redefines its own Heavy candidate set from the 85th positive-score quantile and uses its own 99.5th-quantile closure-mapping upper bound.",
        "- Every outcome is the mean of five fixed 1,000-draw seed sets.",
        "",
        "## Community isolation",
        "",
        network_frame.to_markdown(index=False),
        "",
        "## Heavy service loss",
        "",
        service_frame.to_markdown(index=False),
        "",
        "## Interpretation boundary",
        "",
        decision["interpretation"],
        "Emergency-water results remain a conditional sensitivity because only geolocated records enter the network analysis.",
        "",
    ]
    (OUT / "downstream_sensitivity_report.md").write_text(
        "\n".join(report), encoding="utf-8"
    )
    print(json.dumps(decision, ensure_ascii=False, indent=2), flush=True)


if __name__ == "__main__":
    main()
