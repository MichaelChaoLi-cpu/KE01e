#!/usr/bin/env python3
"""Analytical Data Coverage and Quality.

Plan: Document completeness, temporal coverage, spatial support, and interpretation
limits for each analytical layer.
Framework: Section 5 baseline quality gates; Section 6 support definitions; Section 7
coverage, missingness, geometry validity, temporal range, and calibration-role audit.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Callable

import numpy as np
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd
import rasterio


ROOT = Path(__file__).resolve().parents[2]
PROCESSED = ROOT / "data/processed"
OUT = ROOT / "data/results/tables/Table_analytical_data_coverage_and_quality.xlsx"
SHEET_NAME = "Data Coverage"
TABLE_TITLE = "Analytical Data Coverage and Quality"


@dataclass(frozen=True)
class LayerSpec:
    label: str
    file_name: str
    spatial_type: str
    spatial_support: str
    time_columns: tuple[str, ...]
    key_identifier: str
    required_fields: tuple[str, ...]
    location_fields: tuple[str, ...]
    analytical_role: str
    interpretation_boundary: str
    location_rule: Callable[[pd.DataFrame], pd.Series] | None = None


LAYER_SPECS = [
    LayerSpec("Administrative areas", "administrative_areas_preprocessed.parquet", "Polygon", "Municipality or ward", tuple(), "Municipality Code", ("Municipality Code", "Municipality Name"), ("Geometry",), "Spatial reporting and threshold assignment", "Administrative boundaries do not represent within-area hazard variation."),
    LayerSpec("Current shelters", "current_shelters_preprocessed.parquet", "Point when resolved", "Facility record", ("Snapshot Time",), "Shelter Number", ("Shelter Number", "Shelter Name", "Snapshot Time"), ("Latitude", "Longitude"), "Current shelter status and service context", "Unmatched current shelters remain unresolved and are not treated as unreachable."),
    LayerSpec("Designated shelters", "designated_shelters_preprocessed.parquet", "Point", "Facility location", tuple(), "Shelter ID", ("Shelter ID", "Shelter Name"), ("Geometry",), "Shelter reachability target", "Designation does not establish current operational status or capacity."),
    LayerSpec("Earthquake damage evidence", "earthquake_damage_evidence_preprocessed.parquet", "Point", "Reported damage location", ("Observation Time",), "Evidence ID", ("Evidence ID", "Observation Time", "Observed Damage Type", "Evidence Tier"), ("Latitude", "Longitude"), "Earthquake-context evidence", "Evidence is selectively reported and cannot be interpreted as a complete damage inventory."),
    LayerSpec("Emergency evacuation sites", "emergency_evacuation_sites_preprocessed.parquet", "Point", "Facility location", tuple(), "Evacuation Site ID", ("Evacuation Site ID", "Evacuation Site Name"), ("Geometry",), "Evacuation reachability target", "Designation categories do not establish real-time opening or accessibility."),
    LayerSpec("Emergency transport roads", "emergency_transport_roads_preprocessed.parquet", "Line", "Declared route segment", ("Source Date",), "Route ID + Branch ID", ("Route ID", "Branch ID", "Emergency Road Class"), ("Geometry",), "External-road targets and intervention stratification", "Plan membership is a functional designation, not observed passability."),
    LayerSpec("Emergency water points", "emergency_water_points_preprocessed.parquet", "Point when resolved", "Current supply point", ("Valid From Date", "Valid To Date", "Source Status Time"), "Water Point Name", ("Water Point Name", "Source Status Time"), ("Latitude", "Longitude"), "Emergency-water reachability target", "Only resolved current points enter routing; unmatched records make results a lower bound."),
    LayerSpec("Evacuation facilities", "evacuation_facilities_preprocessed.parquet", "Point", "Facility location", tuple(), "Source Record ID", ("Source Record ID", "Facility Name", "Facility Type"), ("Geometry",), "Supplementary evacuation-facility context", "Source vintages and facility labels may not describe current operating conditions."),
    LayerSpec("Fire stations", "fire_stations_preprocessed.parquet", "Point", "Fire facility", tuple(), "Fire Facility Name", ("Fire Facility Name", "Fire Facility Type"), ("Geometry",), "Fire-service reachability target", "Facility presence does not imply vehicle, staff, or dispatch availability."),
    LayerSpec("2016 landslide inventory", "gsi_2016_landslide_inventory_preprocessed.parquet", "Point", "Interpreted landslide placemark", ("Observation Date",), "Landslide Inventory ID", ("Landslide Inventory ID", "Observation Date"), ("Geometry",), "Presence-background calibration evidence", "Incomplete presence evidence; absence of a point is not a confirmed non-event."),
    LayerSpec("GSI DEM 10B elevation", "gsi_dem10b_elevation_preprocessed.tif", "Raster", "Approximately 10 m terrain cell", tuple(), "Raster cell", tuple(), tuple(), "Elevation, slope, and curvature derivation", "DEM support does not imply rainfall or hazard information at 10 m resolution."),
    LayerSpec("JMA hourly rainfall", "jma_hourly_rainfall_preprocessed.parquet", "Station time series", "Station-hour", ("Observation Time",), "Station ID + Observation Time", ("Station ID", "Observation Time", "Hourly Rainfall"), ("Station ID",), "Rainfall history and scenario quantiles", "Station support is not fine-resolution rainfall interpolation."),
    LayerSpec("Landslide warning zones", "landslide_warning_zones_preprocessed.parquet", "Polygon", "Official warning-zone polygon", ("Designation Date",), "Zone ID", ("Zone ID", "Hazard Type", "Warning Zone Class"), ("Geometry",), "Warning-zone exposure baseline", "Official zoning is a screening layer and does not represent event occurrence."),
    LayerSpec("Official threshold factors", "official_threshold_factors_preprocessed.parquet", "Area-linked table", "Municipality or named subarea", tuple(), "Municipality or Subarea (Japanese)", ("Municipality or Subarea (Japanese)", "Rainfall Threshold Retention Factor"), ("Municipality or Subarea (Japanese)",), "Post-earthquake threshold scenario adjustment", "Factors are area-level official settings, not a continuous shaking surface."),
    LayerSpec("Population disclosure groups", "population_disclosure_groups_preprocessed.parquet", "Polygon", "Disclosure group of 125 m meshes", tuple(), "Disclosure Group Code", ("Disclosure Group Code", "Total Population", "Population Age 65+"), ("Geometry",), "Age-structured vulnerability allocation", "Suppressed meshes require group-level allocation and do not reveal exact small-cell age counts."),
    LayerSpec("Population mesh 125 m", "population_mesh_125m_preprocessed.parquet", "Polygon", "125 m populated mesh", tuple(), "Mesh Code", ("Mesh Code", "Disclosure Group Code", "Total Population"), ("Geometry",), "Community construction and population weighting", "Mesh population is census-based and may differ from post-earthquake occupancy."),
    LayerSpec("Public offices and halls", "public_offices_halls_preprocessed.parquet", "Point", "Support facility", ("Source Reference Year",), "Support Facility ID", ("Support Facility ID", "Facility Name", "Support Facility Type"), ("Geometry",), "Municipal-service reachability target", "Reference-year facility presence does not establish current service capacity."),
    LayerSpec("Road edges", "road_edges_preprocessed.parquet", "Line", "Routable directed/undirected edge", tuple(), "Road Edge ID", ("Road Edge ID", "Road Section ID", "From Node ID", "To Node ID", "Baseline Edge Travel Time (min)"), ("Geometry",), "Routing, travel time, and disruption simulation", "Travel times use assumed speeds and omit real-time traffic and repair duration."),
    LayerSpec("Road nodes", "road_nodes_preprocessed.parquet", "Point", "Network intersection or endpoint", tuple(), "Network Node ID", ("Network Node ID", "Network Component ID"), ("Geometry",), "Community and service attachment", "Nearest-node attachment does not establish field-level entrance accessibility."),
    LayerSpec("Restriction-edge matches", "road_restriction_edge_matches_preprocessed.parquet", "Line-to-network match", "Restriction observation × matched edge", ("Snapshot Time",), "Restriction Observation ID + Matched Road Edge ID", ("Restriction Observation ID", "Snapshot Time", "Matched Road Edge ID", "Road Edge Match Status"), ("Geometry",), "Road disruption ranking validation", "Unreported roads are not non-events; uncertain matches cannot calibrate failure probability."),
    LayerSpec("Road restrictions", "road_restrictions_preprocessed.parquet", "Line or reported place", "Restriction record snapshot", ("Snapshot Time", "Restriction Start Time", "Restriction Change Time"), "Feature Index + Snapshot Time", ("Snapshot Time", "Feature Index", "Restriction Reason"), ("Geometry JSON",), "Observed restriction evidence", "Snapshots contain repeats and causes beyond landslide-related disruption."),
    LayerSpec("Road sections", "road_sections_preprocessed.parquet", "Line", "Analysis road section", tuple(), "Road Section ID", ("Road Section ID", "Section From Node ID", "Section To Node ID", "Road Section Length (m)"), ("Geometry",), "Road disruption and intervention unit", "Section scores are screening indices, not calibrated closure probabilities."),
]


def missing_fraction(frame: pd.DataFrame, columns: tuple[str, ...]) -> float:
    """Return cell-level missingness across declared required columns."""
    if not columns:
        return 0.0
    missing = frame.loc[:, list(columns)].isna().to_numpy()
    return float(missing.mean())


def location_completeness(frame: pd.DataFrame, spec: LayerSpec) -> float:
    """Return the share of records with complete declared spatial support."""
    if spec.location_rule is not None:
        return float(spec.location_rule(frame).mean())
    if not spec.location_fields:
        return 1.0
    return float(frame.loc[:, list(spec.location_fields)].notna().all(axis=1).mean())


def temporal_coverage(frame: pd.DataFrame, columns: tuple[str, ...]) -> str:
    """Summarize minimum and maximum date/time across declared temporal columns."""
    values: list[pd.Series] = []
    years: list[pd.Series] = []
    for column in columns:
        if column not in frame:
            continue
        if "Year" in column:
            series = pd.to_numeric(frame[column], errors="coerce").dropna().astype(int)
            if len(series):
                years.append(series)
            continue
        series = pd.to_datetime(frame[column], errors="coerce", utc=True).dropna()
        if len(series):
            values.append(series)
    if years and not values:
        combined_years = pd.concat(years, ignore_index=True)
        minimum_year = int(combined_years.min())
        maximum_year = int(combined_years.max())
        return str(minimum_year) if minimum_year == maximum_year else f"{minimum_year} to {maximum_year}"
    if not values:
        return "Static / reference layer"
    combined = pd.concat(values, ignore_index=True)
    minimum = combined.min().date().isoformat()
    maximum = combined.max().date().isoformat()
    return minimum if minimum == maximum else f"{minimum} to {maximum}"


def raster_row(spec: LayerSpec) -> dict[str, object]:
    """Build the DEM row from raster metadata and the precomputed tile audit."""
    path = PROCESSED / spec.file_name
    with rasterio.open(path) as source:
        total_cells = int(source.width * source.height)
    tile_summary = pd.read_parquet(
        PROCESSED / "gsi_dem10b_tile_summary_preprocessed.parquet",
        columns=["Valid Pixel Count"],
    )
    valid_cells = int(tile_summary["Valid Pixel Count"].sum())
    return {
        "Analytical Data Layer": spec.label,
        "Record Count": total_cells,
        "Spatial Type": spec.spatial_type,
        "Spatial Resolution / Support": spec.spatial_support,
        "Temporal Coverage": "Static terrain surface",
        "Key Identifier": spec.key_identifier,
        "Required-Field Missingness": 1.0 - valid_cells / total_cells,
        "Location Completeness": valid_cells / total_cells,
        "Analytical Role": spec.analytical_role,
        "Interpretation Boundary": spec.interpretation_boundary,
    }


def build_table() -> pd.DataFrame:
    """Assemble one quality-audit row per planned analytical layer."""
    rows: list[dict[str, object]] = []
    for spec in LAYER_SPECS:
        if spec.file_name.endswith(".tif"):
            rows.append(raster_row(spec))
            continue
        columns = tuple(dict.fromkeys((*spec.required_fields, *spec.location_fields, *spec.time_columns)))
        frame = pd.read_parquet(PROCESSED / spec.file_name, columns=list(columns))
        rows.append(
            {
                "Analytical Data Layer": spec.label,
                "Record Count": int(len(frame)),
                "Spatial Type": spec.spatial_type,
                "Spatial Resolution / Support": spec.spatial_support,
                "Temporal Coverage": temporal_coverage(frame, spec.time_columns),
                "Key Identifier": spec.key_identifier,
                "Required-Field Missingness": missing_fraction(frame, spec.required_fields),
                "Location Completeness": location_completeness(frame, spec),
                "Analytical Role": spec.analytical_role,
                "Interpretation Boundary": spec.interpretation_boundary,
            }
        )
    table = pd.DataFrame(rows)
    if table.shape != (22, 10):
        raise RuntimeError(f"Expected a 22 × 10 table, found {table.shape}.")
    if table["Analytical Data Layer"].duplicated().any():
        raise RuntimeError("Analytical data-layer labels must be unique.")
    for column in ("Required-Field Missingness", "Location Completeness"):
        if not table[column].between(0, 1).all():
            raise RuntimeError(f"{column} contains a value outside [0, 1].")
    return table


def style_workbook(path: Path) -> None:
    """Apply compact scientific-table formatting and validation-friendly styling."""
    workbook = load_workbook(path)
    worksheet = workbook[SHEET_NAME]
    worksheet.insert_rows(1)
    worksheet.merge_cells("A1:J1")
    worksheet["A1"] = TABLE_TITLE
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A3"
    worksheet.auto_filter.ref = f"A2:J{worksheet.max_row}"
    worksheet.sheet_view.zoomScale = 85
    worksheet.print_area = f"A1:J{worksheet.max_row}"
    worksheet.print_title_rows = "1:2"
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A3
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 1
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_margins = PageMargins(
        left=0.25,
        right=0.25,
        top=0.35,
        bottom=0.35,
        header=0.15,
        footer=0.15,
    )

    header_fill = PatternFill("solid", fgColor="17365D")
    header_font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor="D9EAF7")
    title_font = Font(name="Aptos Display", size=14, bold=True, color="17365D")
    body_font = Font(name="Aptos", size=9, color="172033")
    subtle_border = Border(bottom=Side(style="thin", color="D0D5DD"))
    worksheet["A1"].fill = title_fill
    worksheet["A1"].font = title_font
    worksheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    worksheet.row_dimensions[1].height = 30
    for cell in worksheet[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[2].height = 34

    for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = subtle_border
        row[1].number_format = "#,##0"
        row[6].number_format = "0.0%"
        row[7].number_format = "0.0%"
        row[1].alignment = Alignment(horizontal="right", vertical="top")
        row[6].alignment = Alignment(horizontal="right", vertical="top")
        row[7].alignment = Alignment(horizontal="right", vertical="top")
        worksheet.row_dimensions[row[0].row].height = 45

    widths = {
        "A": 31,
        "B": 14,
        "C": 19,
        "D": 27,
        "E": 23,
        "F": 28,
        "G": 20,
        "H": 18,
        "I": 37,
        "J": 54,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    worksheet.conditional_formatting.add(
        f"G3:G{worksheet.max_row}",
        ColorScaleRule(
            start_type="num",
            start_value=0,
            start_color="63BE7B",
            mid_type="num",
            mid_value=0.10,
            mid_color="FFEB84",
            end_type="num",
            end_value=1,
            end_color="F8696B",
        ),
    )
    worksheet.conditional_formatting.add(
        f"H3:H{worksheet.max_row}",
        ColorScaleRule(
            start_type="num",
            start_value=0,
            start_color="F8696B",
            mid_type="num",
            mid_value=0.80,
            mid_color="FFEB84",
            end_type="num",
            end_value=1,
            end_color="63BE7B",
        ),
    )
    excel_table = Table(displayName="AnalyticalDataCoverage", ref=f"A2:J{worksheet.max_row}")
    excel_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(excel_table)
    workbook.save(path)


def verify_workbook(path: Path) -> None:
    """Reopen the final workbook and verify structure, types, and error-free cells."""
    workbook = load_workbook(path, data_only=False)
    if workbook.sheetnames != [SHEET_NAME]:
        raise RuntimeError(f"Unexpected workbook sheets: {workbook.sheetnames}")
    worksheet = workbook[SHEET_NAME]
    if worksheet.max_row != 24 or worksheet.max_column != 10:
        raise RuntimeError(
            f"Expected 24 worksheet rows including title and header and 10 columns; found "
            f"{worksheet.max_row} × {worksheet.max_column}."
        )
    if worksheet["A1"].value != TABLE_TITLE:
        raise RuntimeError("Workbook title row is missing or incorrect.")
    error_tokens = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"}
    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value in error_tokens:
                raise RuntimeError(f"Spreadsheet error token in {cell.coordinate}: {cell.value}")
    for row in range(3, worksheet.max_row + 1):
        if not isinstance(worksheet.cell(row, 2).value, int):
            raise RuntimeError(f"Record Count must be integer at B{row}.")
        for column in (7, 8):
            value = worksheet.cell(row, column).value
            if not isinstance(value, (int, float)) or not 0 <= value <= 1:
                raise RuntimeError(f"Invalid percentage at row {row}, column {column}.")


def main() -> None:
    table = build_table()
    OUT.parent.mkdir(parents=True, exist_ok=True)
    table.to_excel(OUT, index=False, sheet_name=SHEET_NAME, engine="openpyxl")
    style_workbook(OUT)
    verify_workbook(OUT)
    print(f"Saved: {OUT.relative_to(ROOT)}")
    print(f"Rows: {len(table):,}; columns: {len(table.columns):,}")
    print(
        "Lowest location completeness: "
        f"{table.loc[table['Location Completeness'].idxmin(), 'Analytical Data Layer']} "
        f"({table['Location Completeness'].min():.1%})"
    )
    print(
        "Highest required-field missingness: "
        f"{table.loc[table['Required-Field Missingness'].idxmax(), 'Analytical Data Layer']} "
        f"({table['Required-Field Missingness'].max():.1%})"
    )
    print("Workbook verification: passed")


if __name__ == "__main__":
    main()
