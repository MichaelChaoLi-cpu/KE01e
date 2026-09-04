#!/usr/bin/env python3
"""Build Appendix Table B10 from the verified R3C6 target-sensitivity output."""

from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


TITLE = "Table B10. Emergency-road-backbone target sensitivity"
HEADERS = [
    "Target definition",
    "Rainfall scenario",
    "Target roots",
    "Eligible communities",
    "Eligible meshes",
    "Eligible population",
    "Expected disconnected population, mean [five-seed range]",
    "Expected disconnected age 65+ population, mean [five-seed range]",
    "Community-frequency Spearman vs primary",
    "Top-30 population-burden overlap vs primary",
]


def one_decimal(value: float) -> str:
    return f"{value:,.1f}"


def target_label(value: str) -> str:
    if value == "Legacy coast-inclusive boundary proxy":
        return f"{value} (audit only)"
    return value


def comparison_value(row: pd.Series, field: str, *, percentage: bool = False) -> str:
    if "Legacy coast-inclusive" in str(row["target_definition"]):
        return "Not comparable"
    value = row[field]
    if pd.isna(value):
        raise ValueError(f"Missing comparison value for {row['target_definition']} / {row['scenario']}")
    return f"{100 * float(value):.1f}%" if percentage else f"{float(value):.3f}"


def rows_from_csv(path: Path) -> list[list[object]]:
    frame = pd.read_csv(path)
    required = {
        "target_definition",
        "scenario",
        "target_root_count",
        "eligible_community_count",
        "eligible_mesh_count",
        "eligible_population",
        "expected_isolated_population_mean",
        "expected_isolated_population_min",
        "expected_isolated_population_max",
        "expected_isolated_age65_mean",
        "expected_isolated_age65_min",
        "expected_isolated_age65_max",
        "frequency_spearman_vs_primary",
        "top30_population_burden_overlap_vs_primary",
    }
    missing = required.difference(frame.columns)
    if missing:
        raise ValueError(f"Missing required columns: {sorted(missing)}")
    if len(frame) != 9:
        raise ValueError(f"Expected 9 target-scenario rows, found {len(frame)}")

    target_order = {
        "Primary emergency-road backbone": 0,
        "All emergency-road backbone": 1,
        "Legacy coast-inclusive boundary proxy": 2,
    }
    scenario_order = {"Moderate": 0, "Heavy": 1, "Extreme": 2}
    frame = frame.assign(
        _scenario=frame["scenario"].map(scenario_order),
        _target=frame["target_definition"].map(target_order),
    ).sort_values(["_scenario", "_target"])
    if frame[["_scenario", "_target"]].isna().any().any():
        raise ValueError("Unexpected target definition or rainfall scenario")

    rows: list[list[object]] = []
    for _, row in frame.iterrows():
        total = (
            f"{one_decimal(row['expected_isolated_population_mean'])} "
            f"[{one_decimal(row['expected_isolated_population_min'])}–"
            f"{one_decimal(row['expected_isolated_population_max'])}]"
        )
        older = (
            f"{one_decimal(row['expected_isolated_age65_mean'])} "
            f"[{one_decimal(row['expected_isolated_age65_min'])}–"
            f"{one_decimal(row['expected_isolated_age65_max'])}]"
        )
        rows.append(
            [
                target_label(str(row["target_definition"])),
                str(row["scenario"]),
                int(row["target_root_count"]),
                int(row["eligible_community_count"]),
                int(row["eligible_mesh_count"]),
                int(row["eligible_population"]),
                total,
                older,
                comparison_value(row, "frequency_spearman_vs_primary"),
                comparison_value(
                    row,
                    "top30_population_burden_overlap_vs_primary",
                    percentage=True,
                ),
            ]
        )
    return rows


def build_workbook(csv_path: Path, output_path: Path) -> None:
    rows = rows_from_csv(csv_path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Table B10"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A3"

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    title_cell = sheet.cell(1, 1, TITLE)
    title_cell.font = Font(name="Arial", size=12, bold=True, color="1F4E78")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 24

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_side = Side(style="thin", color="FFFFFF")
    for column, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(2, column, header)
        cell.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=header_side, right=header_side)
    sheet.row_dimensions[2].height = 58

    body_side = Side(style="thin", color="D9E2F3")
    for row_index, values in enumerate(rows, start=3):
        fill = PatternFill("solid", fgColor="F4F7F9" if row_index % 2 else "FFFFFF")
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, column, value)
            cell.font = Font(name="Arial", size=9, color="222222")
            cell.fill = fill
            cell.alignment = Alignment(
                horizontal="left" if column == 1 else "center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = Border(bottom=body_side)
            if column in {3, 4, 5, 6}:
                cell.number_format = "#,##0"
        sheet.row_dimensions[row_index].height = 31

    widths = [36, 15, 13, 16, 14, 16, 28, 30, 22, 22]
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width

    sheet.auto_filter.ref = f"A2:{get_column_letter(len(HEADERS))}{2 + len(rows)}"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.print_title_rows = "1:2"
    sheet.sheet_properties.outlinePr.summaryBelow = True
    sheet.page_margins.left = 0.2
    sheet.page_margins.right = 0.2
    sheet.page_margins.top = 0.35
    sheet.page_margins.bottom = 0.35

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def verify_workbook(path: Path) -> dict[str, object]:
    workbook = load_workbook(path, data_only=False)
    if workbook.sheetnames != ["Table B10"]:
        raise RuntimeError(f"Unexpected sheets: {workbook.sheetnames}")
    sheet = workbook["Table B10"]
    if (sheet.max_row, sheet.max_column) != (11, 10):
        raise RuntimeError(f"Unexpected shape: {(sheet.max_row, sheet.max_column)}")
    if sheet["A1"].value != TITLE:
        raise RuntimeError("Title verification failed")
    if [sheet.cell(2, column).value for column in range(1, 11)] != HEADERS:
        raise RuntimeError("Header verification failed")
    legacy_rows = [row for row in range(3, 12) if "Legacy" in str(sheet.cell(row, 1).value)]
    if legacy_rows != [5, 8, 11]:
        raise RuntimeError(f"Unexpected legacy rows: {legacy_rows}")
    for row in legacy_rows:
        if sheet.cell(row, 9).value != "Not comparable" or sheet.cell(row, 10).value != "Not comparable":
            raise RuntimeError("Legacy comparison labeling verification failed")
    return {
        "path": str(path),
        "sheet": sheet.title,
        "shape": [sheet.max_row, sheet.max_column],
        "legacy_rows": legacy_rows,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--input",
        type=Path,
        default=Path("data/exp/revision/reviewer-3-comment-6/isolation_target_sensitivity_summary.csv"),
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("data/results/tables/Table_emergency_backbone_target_sensitivity.xlsx"),
    )
    args = parser.parse_args()
    build_workbook(args.input, args.output)
    print(verify_workbook(args.output))


if __name__ == "__main__":
    main()
