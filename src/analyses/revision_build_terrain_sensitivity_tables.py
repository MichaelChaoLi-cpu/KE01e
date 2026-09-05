"""Build approved Tables B14/B15 from the frozen proposal and verified CSV evidence."""
from __future__ import annotations

import csv
import hashlib
import html
import json
import re
import subprocess
import textwrap
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "data/exp/revision/reviewer-3-comment-2"
TABLES = ROOT / "data/results/tables"
PROPOSAL = ROOT / "Rev/docs/proposal-reviewer-3-comment-2.md"
STEMS = {14: "Table_terrain_weight_sensitivity", 15: "Table_staged_parameter_sensitivity_coverage"}


def approved_table(number):
    assert hashlib.sha256(PROPOSAL.read_bytes()).hexdigest() == "cb69fbccec0167fa212560dd541b82d362cb586da720d716aa5df96369777c54"
    text = PROPOSAL.read_text().split(f"## Part {number - 9} —", 1)[1].split("\n## ", 1)[0]
    title = re.search(r'Title \(merged first row\): "([^"]+)"', text)[1]
    note = re.search(r'Complete new notes: "([^"]+)"', text)[1]
    lines = [line for line in text.splitlines() if line.startswith("| ")]
    rows = [[cell.strip() for cell in line.strip("|").split("|")] for line in lines]
    return title, rows[0], rows[1:], note


def render_svg(title, headers, rows, note, path, widths):
    # Source and preview share the exact verified table content; natural text wrapping.
    unit = 95
    widths = [w * unit for w in widths]
    margin = 24
    total = sum(widths)
    font = 13
    line_h = 19
    parts = []
    y = margin
    def line(text, x, baseline, size=font, bold=False, color="#222222", anchor="start"):
        parts.append(f'<text x="{x}" y="{baseline}" font-family="Arial" font-size="{size}" font-weight="{700 if bold else 400}" fill="{color}" text-anchor="{anchor}">{html.escape(text)}</text>')
    line(title, margin, y + 20, 18, True, "#1F4E78")
    y += 38
    for index, row in enumerate([headers, *rows]):
        wrapped = [textwrap.wrap(value, max(8, int((width-18)/(font*.53))), break_long_words=False)
                   for value, width in zip(row, widths)]
        height = max(len(value) for value in wrapped) * line_h + 18
        fill = "#1F4E78" if index == 0 else ("#F4F7F9" if index % 2 else "#FFFFFF")
        parts.append(f'<rect x="{margin}" y="{y}" width="{total}" height="{height}" fill="{fill}"/>')
        x = margin
        for col, (value, width) in enumerate(zip(wrapped, widths)):
            numeric = len(headers) == 7 and col > 0 and index > 0
            for offset, text in enumerate(value):
                line(text, x + width - 10 if numeric else x + 9,
                     y + (height-len(value)*line_h)/2 + 14 + offset*line_h,
                     bold=index == 0, color="#FFFFFF" if index == 0 else "#222222",
                     anchor="end" if numeric else "start")
            x += width
        y += height
        parts.append(f'<path d="M {margin} {y} H {margin+total}" stroke="#D9D9D9" stroke-width="0.7"/>')
    y += 24
    for text in textwrap.wrap(note, int(total/6.1), break_long_words=False):
        line(text, margin, y, 12)
        y += 18
    height = y + margin
    path.write_text(f'<svg xmlns="http://www.w3.org/2000/svg" width="{total+2*margin}" height="{height}" viewBox="0 0 {total+2*margin} {height}"><rect width="100%" height="100%" fill="white"/>{"".join(parts)}</svg>')
    node = "/Users/lichao/.cache/codex-runtimes/codex-primary-runtime/dependencies/node/bin/node"
    sharp = "/Users/lichao/.cache/codex-runtimes/codex-primary-runtime/dependencies/node/node_modules/sharp"
    subprocess.run([node, "-e", f'const sharp=require({json.dumps(sharp)}); sharp(process.argv[1],{{density:300}}).withMetadata({{density:300}}).png().toFile(process.argv[2]);', str(path), str(path.with_suffix('.png'))], check=True)


def build():
    summary = list(csv.DictReader((OUT / "terrain_weight_summary.csv").open()))
    receipts = []
    for number in (14, 15):
        title, headers, rows, note = approved_table(number)
        assert len(rows) == (9 if number == 14 else 8)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = f"Table B{number}"
        sheet.sheet_view.showGridLines = False
        ncol = len(headers)
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
        sheet.cell(1, 1, title)
        sheet.append(headers)
        for index, row in enumerate(rows):
            if number == 14:
                source = summary[index]
                values = [float(source[key]) for key in
                          ('isolated_population', 'older_isolated_population', 'protected_population',
                           'community_top30_overlap', 'intervention_top30_overlap', 'portfolio_overlap_central')]
                means = f"{values[2]:.1f} [{float(source['protected_seed_min']):.1f}–{float(source['protected_seed_max']):.1f}]"
                display = [row[0], f"{values[0]:,.1f}", f"{values[1]:,.1f}", means,
                           *[f"{value*100:.1f}" for value in values[3:]]]
                assert display == row, (display, row)
                sheet.append([row[0], *values[:3], *[value*100 for value in values[3:]]])
                for col in range(2, 8):
                    sheet.cell(index+3, col).number_format = '#,##0.0'
                sheet.cell(index+3, 4).number_format = f'0.0" [{float(source["protected_seed_min"]):.1f}–{float(source["protected_seed_max"]):.1f}]"'
            else:
                sheet.append(row)
        widths = [27, 20, 14, 30, 20, 22, 20] if number == 14 else [27, 39, 38, 80]
        for col, width in enumerate(widths, 1):
            sheet.column_dimensions[get_column_letter(col)].width = width
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = Font(name='Arial', size=11, color='222222')
                cell.alignment = Alignment(vertical='center', horizontal='left', wrap_text=True)
                if cell.row == 2:
                    cell.fill = PatternFill('solid', fgColor='1F4E78')
                    cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
                    cell.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
                elif cell.row > 2:
                    if cell.row % 2:
                        cell.fill = PatternFill('solid', fgColor='F4F7F9')
                    cell.border = Border(bottom=Side(style='hair', color='D9D9D9'))
                    if cell.data_type == 'n':
                        cell.alignment = Alignment(horizontal='right', vertical='center')
        sheet['A1'].font = Font(name='Arial', size=14, bold=True, color='1F4E78')
        sheet.row_dimensions[1].height = 28
        sheet.row_dimensions[2].height = 46
        for r in range(3, 3+len(rows)):
            sheet.row_dimensions[r].height = 32 if number == 14 else 64
        note_row = len(rows)+4
        sheet.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=ncol)
        sheet.cell(note_row, 1, note).alignment = Alignment(wrap_text=True, vertical='top')
        sheet.cell(note_row, 1).font = Font(name='Arial', size=10, italic=True)
        sheet.row_dimensions[note_row].height = 110 if number == 14 else 85
        source_row = note_row+2
        sheet.merge_cells(start_row=source_row, start_column=1, end_row=source_row, end_column=ncol)
        sheet.cell(source_row, 1, 'Source: '+('terrain_weight_summary.csv; score_diagnostics.csv' if number == 14 else 'staged_uncertainty_coverage.csv; Appendix Tables B5–B8 and B12–B14'))
        sheet.cell(source_row,1).font = Font(name='Arial',size=10)
        if number == 14:
            sheet.cell(source_row+2, 1, 'Protected-population seed bounds (unrounded source values)')
            for col, label in enumerate(['Weight stress','Mean','Seed minimum','Seed maximum'],1):
                sheet.cell(source_row+3,col,label).font=Font(name='Arial',size=11,bold=True)
            for i, source in enumerate(summary):
                for col,value in enumerate([rows[i][0],*[float(source[key]) for key in ('protected_population','protected_seed_min','protected_seed_max')]],1):
                    cell=sheet.cell(source_row+4+i,col,value)
                    cell.font=Font(name='Arial',size=11)
                    cell.number_format='#,##0.000000'
        sheet.print_options.horizontalCentered = True
        sheet.page_setup.orientation='landscape'
        sheet.page_setup.paperSize=sheet.PAPERSIZE_A3
        sheet.page_setup.fitToWidth=1
        sheet.page_setup.fitToHeight=1
        sheet.sheet_properties.pageSetUpPr.fitToPage=True
        sheet.print_area=f'A1:{get_column_letter(ncol)}{note_row}'
        path=TABLES / (STEMS[number]+'.xlsx')
        assert not path.exists(), f'Refusing to overwrite {path}'
        workbook.save(path)
        check=load_workbook(path,data_only=True)
        assert check.active['A1'].value==title
        assert [check.active.cell(2,c).value for c in range(1,ncol+1)]==headers
        assert not any(c.data_type=='e' for row in check.active for c in row)
        if number==14:
            assert abs(check.active['B3'].value-float(summary[0]['isolated_population']))<1e-10
        render_svg(title,headers,rows,note,OUT/(STEMS[number]+'.svg'),
                   [1.65,1.05,.75,1.5,1.1,1.2,1.1] if number==14 else [1.4,2.1,1.9,3.2])
        receipts.append(dict(table=number,path=str(path),shape=[len(rows)+2,ncol],
                             exact_approved_display=True,raw_numeric_values_preserved=True,
                             sha256=hashlib.sha256(path.read_bytes()).hexdigest()))
    (OUT/'table_asset_verification.json').write_text(json.dumps(receipts,indent=2)+'\n')
    print(json.dumps(receipts,indent=2))


if __name__=='__main__':
    build()
