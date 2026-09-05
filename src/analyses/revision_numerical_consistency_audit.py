"""Read-only checks of revision headline numbers; write audit evidence only."""
from pathlib import Path
from zipfile import ZipFile
import csv
import hashlib
import json
import re
from lxml import etree
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / 'data/exp/revision/reviewer-2-comment-10'
NS = {'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main',
      'm': 'http://schemas.openxmlformats.org/officeDocument/2006/math'}
INPUTS = {}


def source(rel):
    p = ROOT / rel
    INPUTS[rel] = hashlib.sha256(p.read_bytes()).hexdigest()
    return p


def doc(rel):
    with ZipFile(source(rel)) as z:
        return etree.fromstring(z.read('word/document.xml'))


def text(node):
    return ''.join(node.xpath('.//w:t/text()|.//m:t/text()', namespaces=NS))


def rows(rel):
    with source(rel).open(newline='') as f:
        return list(csv.DictReader(f))


def workbook(rel):
    w = load_workbook(source(rel), read_only=True, data_only=True)
    vals = list(w.active.values)
    w.close()
    return vals


def same_display(display, value):
    if value is None:
        return display.strip() == ''
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        raw = display.strip().replace(',', '').replace('−', '-')
        percent = raw.endswith('%')
        raw = raw.rstrip('%')
        if not re.fullmatch(r'[+-]?\d+(?:\.\d+)?', raw):
            return False
        decimals = len(raw.split('.')[1]) if '.' in raw else 0
        return abs(float(raw) - value*(100 if percent else 1)) <= 0.5*10**(-decimals)+1e-7
    return re.sub(r'\s+', '', display) == re.sub(r'\s+', '', str(value))


def check_b5_narrative(appendix, episode_values):
    """Check the narrative independently of the B5 table's correct cells."""
    candidates = [text(p) for p in appendix.xpath('.//w:body/w:p', namespaces=NS)
                  if text(p).startswith('Road rankings remained similar between the all-area baseline')]
    assert len(candidates) == 1, 'B5 explanatory paragraph missing or ambiguous'
    expected = ('episode-weighted road-restriction correspondence ranged from '
                f'{min(episode_values):.3f} to {max(episode_values):.3f}')
    assert expected in candidates[0], 'B5 narrative correspondence disagrees with episode-weighted source'
    assert not re.search(r'(?<![\d.])0\.(?:632|652)(?![\d.])', candidates[0]), 'Legacy B5 narrative range'
    return dict(status='pass', expected_phrase=expected, paragraph=candidates[0])


def main():
    clean = doc('Rev/revision/KE01e.rev.clean.docx')
    appendix = doc('Rev/revision/Appendix.docx')
    highlights_rel = ('Rev/revision/Highlights.docx' if (ROOT/'Rev/revision/Highlights.docx').exists()
                      else 'article/deliverables/Highlights.docx')
    highlights = doc(highlights_rel)
    paragraphs = [text(p) for p in clean.xpath('.//w:body/w:p', namespaces=NS)]
    # Restrict retired-number scan to scientific body, excluding bibliography.
    stop = next(i for i, t in enumerate(paragraphs) if t.strip() in ('References', 'Reference:', 'References:'))
    body = '\n'.join(paragraphs[:stop])
    tabs = [[[text(c) for c in tr.findall('w:tc', NS)] for tr in tb.findall('w:tr', NS)]
            for tb in appendix.xpath('.//w:tbl', namespaces=NS)]
    main_tabs = [[[text(c) for c in tr.findall('w:tc', NS)] for tr in tb.findall('w:tr', NS)]
                 for tb in clean.xpath('.//w:tbl', namespaces=NS)]
    # Compare every matching canonical workbook table with its embedded Word table.
    table_checks = []
    norm = lambda r: tuple(re.sub(r'\s+', '', str(v or '')) for v in r)
    for file in sorted((ROOT/'data/results/tables').glob('*.xlsx')):
        rel = str(file.relative_to(ROOT))
        w = load_workbook(source(rel), read_only=True, data_only=True)
        formulas = load_workbook(file, read_only=True, data_only=False)
        for sheet in w:
            vals = list(sheet.values)
            if len(vals)<2:
                continue
            header = norm(vals[1])
            matches = [(index, table[start:]) for index,table in enumerate(main_tabs+tabs)
                       for start in range(min(2,len(table))) if norm(table[start]) == header]
            for index,table in matches:
                mismatches = []
                for ri,(actual,expected) in enumerate(zip(table[1:], vals[2:]),1):
                    for ci,(a,e) in enumerate(zip(actual,expected)):
                        if not same_display(a,e):
                            formula = formulas[sheet.title].cell(ri+2,ci+1)
                            if formula.data_type == 'f' and e is None:
                                category = 'uncached_formula_requires_source_check'
                            elif isinstance(e,(int,float)) and ' [' in a and same_display(a.split(' [')[0],e):
                                category = 'matching_mean_with_additional_word_interval'
                            elif isinstance(e,str):
                                category = 'text_or_formatted_value_requires_semantic_check'
                            else:
                                category = 'unresolved_numeric_display_difference'
                            mismatches.append(dict(row=ri,column=ci,word=a,xlsx=e,category=category))
                table_checks.append(dict(source=rel,sheet=sheet.title,combined_table_index=index,
                                         word_data_rows=len(table)-1,xlsx_data_rows=len(vals)-2,
                                         mismatches=mismatches))
        w.close()
        formulas.close()
    c1 = next(t for t in tabs if t[0][0] == 'Admin Area Code')
    municipal = workbook('data/results/tables/Table_municipality_isolation_and_service_loss_summary.xlsx')
    assert list(municipal[1]) == c1[0]
    assert len(municipal[2:]) == len(c1[1:]) == 49
    assert [r[0] for r in municipal[2:]] == [r[0] for r in c1[1:]]
    isolation = rows('data/exp/revision/reviewer-3-comment-6/isolation_target_sensitivity_summary.csv')
    primary = {r['scenario']: r for r in isolation if r['target_definition'] == 'Primary emergency-road backbone'}
    services = {r['Service Class']: r for r in rows('data/exp/revision/reviewer-2-comment-8/service_destination_estimand_summary.csv')}
    checks = []
    for scenario, col in [('Moderate', 5), ('Heavy', 7), ('Extreme', 9)]:
        checks.append((scenario + ' disconnection', col, float(primary[scenario]['expected_isolated_population_mean'])))
    checks.append(('Heavy age65 disconnection', 10, float(primary['Heavy']['expected_isolated_age65_mean'])))
    for service, col in [('Shelter', 11), ('Emergency water', 12), ('Fire service', 13), ('Municipal facility', 14)]:
        checks.append((service + ' loss', col, float(services[service]['Any-Same-Class Loss Population Mean'])))
    totals = []
    for label, col, expected in checks:
        exact = sum(float(r[col]) for r in municipal[2:])
        printed = sum(float(r[col].replace(',', '')) for r in c1[1:])
        row_errors = [abs(float(a[col].replace(',', '')) - float(b[col])) for a,b in zip(c1[1:], municipal[2:])]
        assert max(row_errors) <= 0.050001, (label, max(row_errors))
        assert abs(exact - expected) < 0.002, (label, exact, expected)
        assert f'{expected:,.1f}' in body, (label, 'missing body value')
        totals.append(dict(metric=label, source_estimate=expected, workbook_unrounded_sum=exact,
                           appendix_printed_sum=printed, printed_sum_difference=printed-exact,
                           max_cell_rounding_error=max(row_errors)))
    assert f"{float(primary['Heavy']['expected_isolated_population_mean']):,.0f}" in paragraphs[3]
    assert f"{float(primary['Heavy']['expected_isolated_age65_mean']):,.0f}" in paragraphs[3]
    # Independently compare all printed protected-population cells, not just the maximum.
    portfolios = workbook('data/results/tables/Table_intervention_portfolios.xlsx')
    comparators = workbook('data/results/tables/Table_comparator_robustness.xlsx')
    b3 = next(t for t in tabs if t[0][0]=='Sensitivity Setting')
    b4 = next(t for t in tabs if t[0][0]=='Comparator')
    assert len(b3)-1 == len(portfolios)-2 == 21
    assert len(b4)-1 == len(comparators)-2 == 84
    assert [r[6] for r in b3[1:]] == [r[6] for r in portfolios[2:]]
    assert [r[5] for r in b4[1:]] == [r[5] for r in comparators[2:]]
    maxbudget = max(float(r[1]) for r in portfolios[2:])
    maximum = [r for r in portfolios[2:] if abs(float(r[1])-maxbudget)<1e-8]
    for row in maximum:
        value = float(row[6].split(' ')[0].replace(',',''))
        assert f'{value:,.1f}' in body
    central = next(r for r in maximum if r[0]=='Central')
    equal = next(r for r in comparators[2:] if r[0]=='Equal-cost consequence' and r[1]=='Central' and abs(float(r[2])-maxbudget)<1e-8)
    assert central[6] == equal[5]
    # All eight manuscript raster figures must be the current production assets.
    with ZipFile(ROOT/'Rev/revision/KE01e.rev.clean.docx') as z:
        media = {hashlib.sha256(z.read(n)).hexdigest():n for n in z.namelist() if n.startswith('word/media/')}
    figures = []
    for p in sorted((ROOT/'data/results/figures').glob('*.png')):
        rel = str(p.relative_to(ROOT)); source(rel)
        h = INPUTS[rel]
        assert h in media, rel
        figures.append(dict(asset=rel, embedded=media[h], sha256=h))
    assert len(figures)==8
    retired = ['12,000','3,000','1,107','603','4,723','2,610','2,993','8,266','2,673','28,968','4,920','2,788','259.7']
    matches = {value:[i for i,t in enumerate(paragraphs[:stop]) if re.search(r'(?<![\d.,])'+re.escape(value)+r'(?![\d.,])',t)] for value in retired}
    episode = rows('data/exp/revision/reviewer-2-comment-7/rainfall_parameter_event_clustered_validation.csv')
    assert len(episode) == 15
    central_episode = next(r for r in episode if r['Specification']=='equal__g1.00')
    values = [float(r['Episode-Weighted Concordance']) for r in episode]
    b5_narrative = check_b5_narrative(appendix, values)
    b5_expected = ['Rainfall parameters', 'Episode-weighted road-restriction correspondence',
                   f"{float(central_episode['Episode-Weighted Concordance']):.3f}",
                   f'{min(values):.3f}–{max(values):.3f}', 'Range across 15 combinations',
                   '10 physical episodes',
                   'Supplementary dry-event correspondence; not rainfall-trigger validation']
    b5_current = next(r for t in tabs for r in t if len(r)==7 and r[1] in
                      ('Matched road-evidence concordance',b5_expected[1]))
    report = dict(scope='Headline outcomes and their repeated displays, not full model re-estimation or submission clearance',
                  source_hashes=INPUTS, municipality_reconciliation=totals, table_display_checks=table_checks,
                  portfolio_and_comparator_protected_population_rows_checked=105,
                  maximum_budget=maxbudget, maximum_portfolios=maximum,
                  all_eight_figures_match_embedded_bytes=figures,
                  retired_number_hits_in_scientific_body={k:v for k,v in matches.items() if v},
                  current_highlights=[text(p) for p in highlights.xpath('.//w:body/w:p',namespaces=NS)],
                  headline_result_checks='pass',
                  highlights_source=highlights_rel,
                  highlights_correction_required=('12,000' in text(highlights) or '3,000' in text(highlights) or 'strong concordance' in text(highlights)),
                  b5_correspondence=dict(current=b5_current,expected=b5_expected,
                                         correction_required=b5_current!=b5_expected),
                  b5_narrative=b5_narrative,
                  table_check_caveat='Header-matched table displays only. Missing formula caches are not missing results; text and additional intervals need separate semantic review.')
    OUT.mkdir(parents=True, exist_ok=True)
    (OUT/'headline_audit.json').write_text(json.dumps(report, indent=2, ensure_ascii=False)+'\n')
    print(json.dumps({k:v for k,v in report.items() if k not in ('source_hashes','all_eight_figures_match_embedded_bytes','municipality_reconciliation','table_display_checks')},indent=2,ensure_ascii=False))


if __name__ == '__main__':
    main()
