#!/usr/bin/env python3
"""Road-Disruption Validation.

Plan: Compare Moderate, Heavy, and Extreme road-disruption rankings with a
warning-zone baseline, road-length baseline, and municipality-wide Yatsushiro
0.70 and 0.80 bounding assignments.
Framework: AnaSOP Sections 5-7 require matching on municipality, road category,
emergency-route membership, and road-length decile. Physical restriction episodes
are weighted equally and bootstrap uncertainty is clustered by episode. The dry,
earthquake-proximate records support supplementary correspondence only.
"""
from __future__ import annotations

import os
from pathlib import Path

os.environ.pop("PROJ_LIB", None)
os.environ.pop("PROJ_DATA", None)

import numpy as np
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd
from rasterio.features import rasterize
from rasterio.transform import from_bounds
import shapely

import figure_road_disruption_exposure_and_observed_restriction_evidence as road_validation
from cache_fingerprint import cache_matches, content_signature
from road_restriction_event_validation import (
    build_matched_design,
    event_weighted_concordance,
    load_restriction_evidence,
    paired_event_contrast,
)


ROOT = Path(__file__).resolve().parents[2]
PROCESSED = ROOT / "data/processed"
OUT = ROOT / "data/results/tables/Table_road_disruption_validation.xlsx"
CORRESPONDENCE_SHEET = "Correspondence"
FUNNEL_SHEET = "Evidence Funnel"
EVENT_SHEET = "Episode Audit"
TABLE_TITLE = "Road-Restriction Correspondence"
TRIGGER_AUDIT_PATH = (
    ROOT
    / "data/exp/revision/reviewer-2-comment-7/restriction_episode_trigger_audit.csv"
)
WARNING_SCORE_CACHE = (
    ROOT / "data/results/intermediate/road_warning_zone_scores_normalized_v3.npz"
)
MUNICIPALITY_ENGLISH = {
    "八代市": "Yatsushiro City",
    "山江村": "Yamae Village",
    "山都町": "Yamato Town",
    "御船町": "Mifune Town",
    "球磨村": "Kuma Village",
    "甲佐町": "Kosa Town",
    "美里町": "Misato Town",
}
PROCESS_REASON_ENGLISH = {
    "落石": "Rockfall",
    "法面崩落": "Slope collapse",
    "土砂流入": "Sediment inflow",
}
STATION_ENGLISH = {"人吉": "Hitoyoshi", "甲佐": "Kosa"}


def warning_zone_road_score(
    road_geometry: np.ndarray,
    elevation_grid: np.ndarray,
    extent: tuple[float, float, float, float],
    display_shape: tuple[int, int],
    display_transform: object,
) -> np.ndarray:
    """Build or load normalized directional warning-zone exposure by road section."""
    signature = content_signature(
        "road-warning-zone-score-v3",
        files=(road_validation.WARNING_PATH, road_validation.ROAD_PATH, Path(__file__)),
        arrays={"elevation_grid": elevation_grid},
        parameters={
            "road_count": len(road_geometry),
            "extent": tuple(float(value) for value in extent),
            "display_shape": tuple(int(value) for value in display_shape),
            "sample_fractions": road_validation.SAMPLE_FRACTIONS,
            "upslope_radius_cells": road_validation.UPSLOPE_RADIUS_CELLS,
        },
    )
    if WARNING_SCORE_CACHE.exists():
        cached = np.load(WARNING_SCORE_CACHE, allow_pickle=False)
        if cache_matches(cached, signature):
            return cached["score"].astype("float32")
    warning = pd.read_parquet(road_validation.WARNING_PATH, columns=["Geometry"])
    warning_geometry = road_validation.decode_geometry(warning["Geometry"])
    warning_grid = rasterize(
        ((geometry, 1.0) for geometry in warning_geometry),
        out_shape=display_shape,
        transform=display_transform,
        fill=0.0,
        all_touched=True,
        dtype="float32",
    )
    score = road_validation.road_scores(
        road_geometry,
        {"Warning-zone baseline": warning_grid},
        extent,
        elevation_grid,
    )["Warning-zone baseline"]
    WARNING_SCORE_CACHE.parent.mkdir(parents=True, exist_ok=True)
    np.savez_compressed(
        WARNING_SCORE_CACHE,
        signature=np.asarray(signature),
        score=score.astype("float32"),
    )
    return score


def build_tables() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Build the trigger audit, evidence funnel, and correspondence tables."""
    admin = pd.read_parquet(
        road_validation.ADMIN_PATH,
        columns=["Municipality Name", "Geometry"],
    )
    admin_geometry = road_validation.decode_geometry(admin.pop("Geometry"))
    admin_union = shapely.union_all(admin_geometry)
    min_x, min_y, max_x, max_y = shapely.bounds(admin_union)
    pad_x = (max_x - min_x) * 0.025
    pad_y = (max_y - min_y) * 0.025
    extent = (min_x - pad_x, max_x + pad_x, min_y - pad_y, max_y + pad_y)
    west, east, south, north = extent
    display_height = max(
        650,
        round(road_validation.DISPLAY_WIDTH * (north - south) / (east - west)),
    )
    display_shape = (display_height, road_validation.DISPLAY_WIDTH)
    display_transform = from_bounds(
        west,
        south,
        east,
        north,
        road_validation.DISPLAY_WIDTH,
        display_height,
    )

    landslide_scores, _, _, elevation_grid = (
        road_validation.load_or_build_landslide_scores(
            admin,
            admin_geometry,
            admin_union,
            extent,
            display_shape,
            display_transform,
        )
    )
    roads = pd.read_parquet(
        road_validation.ROAD_PATH,
        columns=[
            "Road Section ID",
            "Road Category",
            "Road Section Length (m)",
            "Emergency Route Membership",
            "Network Analysis Eligible",
            "Geometry",
        ],
    )
    roads = roads.loc[roads["Network Analysis Eligible"]].reset_index(drop=True)
    road_geometry = road_validation.decode_geometry(roads.pop("Geometry"))
    scenario_scores = road_validation.load_or_build_road_scores(
        road_geometry,
        landslide_scores,
        extent,
        elevation_grid,
    )
    warning_score = warning_zone_road_score(
        road_geometry,
        elevation_grid,
        extent,
        display_shape,
        display_transform,
    )
    restriction_evidence = load_restriction_evidence(
        road_validation.RESTRICTION_PATH,
        road_validation.MATCH_PATH,
        road_validation.EDGE_PATH,
        roads["Road Section ID"],
    )
    matched_design = build_matched_design(
        roads,
        road_geometry,
        restriction_evidence.event_section_pairs,
        admin_geometry,
        display_shape,
        display_transform,
        extent,
        road_validation.sample_grid,
    )
    heavy_score = scenario_scores["Heavy"]
    yatsushiro_bound_scores: dict[str, np.ndarray] = {}
    for bound_factor in (0.70, 0.80):
        bound_landslide_scores, _, _, _ = road_validation.load_or_build_landslide_scores(
            admin,
            admin_geometry,
            admin_union,
            extent,
            display_shape,
            display_transform,
            yatsushiro_factor=bound_factor,
        )
        yatsushiro_bound_scores[f"{bound_factor:.2f}"] = (
            road_validation.load_or_build_road_scores(
                road_geometry,
                bound_landslide_scores,
                extent,
                elevation_grid,
                yatsushiro_factor=bound_factor,
            )["Heavy"]
        )
    specifications = [
        (
            "Road score — Moderate rainfall",
            scenario_scores["Moderate"],
            "Scenario ranking; near-unity rank agreement limits spatial reprioritization claims.",
        ),
        (
            "Road score — Heavy rainfall",
            heavy_score,
            "Supplementary dry-event correspondence; not rainfall-trigger validation or a closure probability.",
        ),
        (
            "Road score — Extreme rainfall",
            scenario_scores["Extreme"],
            "Scenario ranking; near-unity rank agreement limits spatial reprioritization claims.",
        ),
        (
            "Road warning-zone exposure baseline",
            warning_score,
            "Official-zone comparator under the same event-weighted matched-control design.",
        ),
        (
            "Road-length baseline",
            roads["Road Section Length (m)"].to_numpy(dtype=float),
            "Length-only comparator under the same event-weighted matched-control design.",
        ),
    ]
    bound_specifications = [
        (
            "Heavy road score — Yatsushiro bound 0.70",
            yatsushiro_bound_scores["0.70"],
            "Municipality-wide analytical bound for unresolved Yatsushiro subareas; not an official municipality value.",
        ),
        (
            "Heavy road score — Yatsushiro bound 0.80",
            yatsushiro_bound_scores["0.80"],
            "Municipality-wide analytical bound for unresolved Yatsushiro subareas; not an official municipality value.",
        ),
    ]

    rows: list[dict[str, object]] = []
    metrics_by_specification: dict[str, dict[str, object]] = {}
    bootstrap_random = np.random.default_rng(20260903)
    for specification, score, interpretation in specifications:
        metrics = event_weighted_concordance(
            np.asarray(score, dtype=float),
            matched_design,
            restriction_evidence.event_section_pairs,
            bootstrap_random=bootstrap_random,
        )
        metrics_by_specification[specification] = metrics
        rank_correlation = float(
            pd.Series(score).corr(pd.Series(heavy_score), method="spearman")
        )
        rows.append(
            {
                "Specification": specification,
                "Evidence Sample (Episodes / Matched Sections / Controls)": (
                    f"{int(metrics['Physical Episodes']):,} / "
                    f"{int(metrics['Matched Evidence Sections']):,} / "
                    f"{int(metrics['Matched Controls']):,}"
                ),
                "Episode-Weighted Concordance": metrics["Road Score Concordance"],
                "Episode-Cluster Bootstrap 95% CI": (
                    f"{metrics['Road Score CI Low']:.3f}–"
                    f"{metrics['Road Score CI High']:.3f}"
                ),
                "Rank Correlation vs Heavy": rank_correlation,
                "Permitted Interpretation": interpretation,
            }
        )
    heavy_events = metrics_by_specification["Road score — Heavy rainfall"][
        "Event Concordance Values"
    ]
    for comparator, label in (
        ("Road warning-zone exposure baseline", "Heavy minus warning-zone baseline"),
        ("Road-length baseline", "Heavy minus road-length baseline"),
    ):
        estimate, low, high = paired_event_contrast(
            heavy_events,
            metrics_by_specification[comparator]["Event Concordance Values"],
            bootstrap_random=bootstrap_random,
        )
        rows.append(
            {
                "Specification": label,
                "Evidence Sample (Episodes / Matched Sections / Controls)": (
                    f"{int(metrics_by_specification[comparator]['Physical Episodes']):,} / "
                    f"{int(metrics_by_specification[comparator]['Matched Evidence Sections']):,} / "
                    f"{int(metrics_by_specification[comparator]['Matched Controls']):,}"
                ),
                "Episode-Weighted Concordance": estimate,
                "Episode-Cluster Bootstrap 95% CI": f"{low:.3f}–{high:.3f}",
                "Rank Correlation vs Heavy": np.nan,
                "Permitted Interpretation": (
                    "Paired episode-level contrast; an interval crossing zero does not establish "
                    "comparative superiority."
                ),
            }
        )
    for specification, score, interpretation in bound_specifications:
        metrics = event_weighted_concordance(
            np.asarray(score, dtype=float),
            matched_design,
            restriction_evidence.event_section_pairs,
            bootstrap_random=bootstrap_random,
        )
        rank_correlation = float(
            pd.Series(score).corr(pd.Series(heavy_score), method="spearman")
        )
        rows.append(
            {
                "Specification": specification,
                "Evidence Sample (Episodes / Matched Sections / Controls)": (
                    f"{int(metrics['Physical Episodes']):,} / "
                    f"{int(metrics['Matched Evidence Sections']):,} / "
                    f"{int(metrics['Matched Controls']):,}"
                ),
                "Episode-Weighted Concordance": metrics["Road Score Concordance"],
                "Episode-Cluster Bootstrap 95% CI": (
                    f"{metrics['Road Score CI Low']:.3f}–"
                    f"{metrics['Road Score CI High']:.3f}"
                ),
                "Rank Correlation vs Heavy": rank_correlation,
                "Permitted Interpretation": interpretation,
            }
        )
    table = pd.DataFrame(rows)
    if table.shape != (9, 6):
        raise RuntimeError(f"Expected a 9 × 6 table, found {table.shape}.")

    episode_audit = pd.read_csv(TRIGGER_AUDIT_PATH)
    episode_table = episode_audit[
        [
            "Municipality",
            "Route Name",
            "Restriction Process Reason",
            "Restriction Start Time",
            "Hours After Earthquake",
            "Repeated Snapshot Rows",
            "Matched Section Count",
            "Nearest JMA Station",
            "Nearest-Station Distance (km)",
            "Nearest 72 h Rainfall (mm)",
            "Trigger Classification",
        ]
    ].copy()
    episode_table.columns = [
        "Municipality",
        "Route",
        "Process Reason",
        "Restriction Start Time",
        "Hours After Earthquake",
        "Repeated Snapshots",
        "Matched Sections",
        "Nearest JMA Station",
        "Nearest-Station Distance (km)",
        "Preceding 72 h Rainfall (mm)",
        "Trigger Classification",
    ]
    episode_table["Municipality"] = episode_table["Municipality"].replace(
        MUNICIPALITY_ENGLISH
    )
    episode_table["Process Reason"] = episode_table["Process Reason"].replace(
        PROCESS_REASON_ENGLISH
    )
    episode_table["Nearest JMA Station"] = (
        episode_table["Nearest JMA Station"]
        .astype(str)
        .str.replace(r"^\d+\s*", "", regex=True)
        .replace(STATION_ENGLISH)
    )
    funnel = restriction_evidence.funnel.copy()
    funnel = pd.concat(
        [
            funnel,
            pd.DataFrame(
                [
                    ("Matched sections with eligible controls", len(matched_design)),
                    ("Retained episodes with explicit rainfall trigger", 0),
                    (
                        "Direct-earthquake-consistent episodes; trigger source-unconfirmed",
                        int(
                            episode_audit["Trigger Classification"]
                            .str.startswith("Direct-earthquake consistent")
                            .sum()
                        ),
                    ),
                ],
                columns=["Stage", "Count"],
            ),
        ],
        ignore_index=True,
    )
    print(
        f"Restriction linkage: {len(restriction_evidence.retained_episodes):,} physical "
        f"episodes; {restriction_evidence.episode_matches['Road Edge ID'].nunique():,} "
        "matched edges"
    )
    return table, funnel, episode_table


def style_workbook(path: Path) -> None:
    """Apply the accepted compact scientific-table style."""
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="17365D")
    header_font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
    body_font = Font(name="Aptos", size=9.5, color="172033")
    subtle_border = Border(bottom=Side(style="thin", color="D0D5DD"))
    settings = {
        CORRESPONDENCE_SHEET: {
            "title": TABLE_TITLE,
            "widths": [39, 31, 24, 28, 23, 55],
            "table": "RoadRestrictionCorrespondence",
        },
        FUNNEL_SHEET: {
            "title": "Road-Restriction Evidence Funnel",
            "widths": [70, 16],
            "table": "RoadRestrictionEvidenceFunnel",
        },
        EVENT_SHEET: {
            "title": "Restriction-Episode Trigger Audit",
            "widths": [22, 22, 18, 24, 18, 18, 17, 21, 18, 22, 48],
            "table": "RestrictionEpisodeTriggerAudit",
        },
    }
    for sheet_name, config in settings.items():
        worksheet = workbook[sheet_name]
        max_column = worksheet.max_column
        last_column = get_column_letter(max_column)
        worksheet.sheet_view.showGridLines = False
        worksheet.freeze_panes = "A3"
        worksheet.sheet_view.zoomScale = 90
        worksheet.print_area = f"A1:{last_column}{worksheet.max_row}"
        worksheet.print_title_rows = "1:2"
        worksheet.page_setup.orientation = "landscape"
        worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A3
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 1
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.page_margins = PageMargins(
            left=0.25,
            right=0.25,
            top=0.35,
            bottom=0.35,
            header=0.10,
            footer=0.10,
        )
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_column)
        title_cell = worksheet["A1"]
        title_cell.value = config["title"]
        title_cell.fill = PatternFill("solid", fgColor="D9EAF7")
        title_cell.font = Font(name="Aptos Display", size=15, bold=True, color="17365D")
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        worksheet.row_dimensions[1].height = 28
        for cell in worksheet[2]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        worksheet.row_dimensions[2].height = 46
        for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row):
            for cell in row:
                cell.font = body_font
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = subtle_border
            row[0].fill = PatternFill("solid", fgColor="FCE8D7")
            if row[0].value == "Road score — Heavy rainfall":
                row[0].fill = PatternFill("solid", fgColor="D9EDE9")
            worksheet.row_dimensions[row[0].row].height = 46
        for index, width in enumerate(config["widths"], start=1):
            worksheet.column_dimensions[get_column_letter(index)].width = width
        excel_table = Table(
            displayName=config["table"],
            ref=f"A2:{last_column}{worksheet.max_row}",
        )
        excel_table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(excel_table)

    correspondence = workbook[CORRESPONDENCE_SHEET]
    for row in range(3, correspondence.max_row + 1):
        correspondence.cell(row, 3).number_format = "0.0%"
        correspondence.cell(row, 5).number_format = "0.000"
    correspondence.conditional_formatting.add(
        f"C3:C{correspondence.max_row}",
        ColorScaleRule(
            start_type="min",
            start_color="F8696B",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFEB84",
            end_type="max",
            end_color="63BE7B",
        ),
    )
    episode = workbook[EVENT_SHEET]
    for row in range(3, episode.max_row + 1):
        episode.cell(row, 5).number_format = "0.00"
        episode.cell(row, 9).number_format = "0.0"
    workbook.save(path)


def verify_workbook(path: Path) -> None:
    """Verify dimensions, numeric fields, and absence of spreadsheet errors."""
    workbook = load_workbook(path, data_only=False)
    expected_sheets = [CORRESPONDENCE_SHEET, FUNNEL_SHEET, EVENT_SHEET]
    if workbook.sheetnames != expected_sheets:
        raise RuntimeError(f"Unexpected workbook sheets: {workbook.sheetnames}")
    correspondence = workbook[CORRESPONDENCE_SHEET]
    if correspondence.max_row != 11 or correspondence.max_column != 6:
        raise RuntimeError(
            "Expected 11 correspondence rows including title and header and 6 columns; "
            f"found {correspondence.max_row} × {correspondence.max_column}."
        )
    if correspondence["A1"].value != TABLE_TITLE:
        raise RuntimeError("Workbook title row is missing or incorrect.")
    error_tokens = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"}
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value in error_tokens:
                    raise RuntimeError(
                        f"Spreadsheet error token in {worksheet.title}!{cell.coordinate}: "
                        f"{cell.value}"
                    )
    for row in range(3, 12):
        if not isinstance(correspondence.cell(row, 3).value, (int, float)):
            raise RuntimeError(f"Expected numeric concordance or contrast at row {row}.")
    if workbook[FUNNEL_SHEET].max_row != 13:
        raise RuntimeError("Evidence funnel must contain 11 audited stages.")
    if workbook[EVENT_SHEET].max_row != 12 or workbook[EVENT_SHEET].max_column != 11:
        raise RuntimeError("Episode audit must contain 10 physical episodes.")


def main() -> None:
    table, funnel, episodes = build_tables()
    OUT.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
        table.to_excel(
            writer,
            index=False,
            sheet_name=CORRESPONDENCE_SHEET,
            startrow=1,
        )
        funnel.to_excel(
            writer,
            index=False,
            sheet_name=FUNNEL_SHEET,
            startrow=1,
        )
        episodes.to_excel(
            writer,
            index=False,
            sheet_name=EVENT_SHEET,
            startrow=1,
        )
    style_workbook(OUT)
    verify_workbook(OUT)
    print(f"Saved: {OUT.relative_to(ROOT)}")
    print(
        f"Correspondence rows: {len(table):,}; funnel rows: {len(funnel):,}; "
        f"episode rows: {len(episodes):,}"
    )
    for row in table.itertuples(index=False):
        print(
            f"{row[0]}: concordance={row[2]:.4f}; CI={row[3]}; "
            f"rho={row[4] if pd.isna(row[4]) else f'{row[4]:.6f}'}"
        )
    print("Workbook verification: passed")


if __name__ == "__main__":
    main()
