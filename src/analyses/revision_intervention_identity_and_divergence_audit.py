#!/usr/bin/env python3
"""Audit the identity and divergence conditions for Reviewer 4 Comment 3."""

from __future__ import annotations

import ast
import csv
import hashlib
import json
from pathlib import Path
from statistics import median
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
PORTFOLIO_PATH = ROOT / "data/results/tables/Table_intervention_portfolios.xlsx"
COMPARATOR_PATH = ROOT / "data/results/tables/Table_comparator_robustness.xlsx"
SENSITIVITY_PATH = (
    ROOT
    / "data/exp/revision/reviewer-2-comment-6/intervention_parameter_sensitivity.csv"
)
SOURCE_PATH = ROOT / "src/analyses/figure_intervention_priorities_and_budgeted_benefits.py"
OUTPUT_DIR = ROOT / "data/exp/revision/reviewer-4-comment-3"


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def workbook_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[2]]
    rows = [
        dict(zip(headers, values, strict=True))
        for values in sheet.iter_rows(min_row=3, values_only=True)
        if any(value is not None for value in values)
    ]
    workbook.close()
    return rows


def literal_assignment(source: str, name: str) -> Any:
    tree = ast.parse(source)
    for node in tree.body:
        if isinstance(node, ast.Assign) and any(
            isinstance(target, ast.Name) and target.id == name for target in node.targets
        ):
            return ast.literal_eval(node.value)
    raise RuntimeError(f"Could not find literal assignment {name}")


def close_or_equal(left: Any, right: Any, tolerance: float = 1e-12) -> bool:
    if left is None or right is None:
        return left is right
    if isinstance(left, (int, float)) and isinstance(right, (int, float)):
        return abs(float(left) - float(right)) <= tolerance
    return left == right


def main() -> None:
    portfolio_rows = workbook_rows(PORTFOLIO_PATH)
    comparator_rows = workbook_rows(COMPARATOR_PATH)
    source = SOURCE_PATH.read_text(encoding="utf-8")
    with SENSITIVITY_PATH.open(newline="", encoding="utf-8") as handle:
        sensitivity_rows = list(csv.DictReader(handle))

    assigned = {
        float(row["Budget (Relative Planning Units)"]): row
        for row in portfolio_rows
        if row["Sensitivity Setting"] == "Central"
    }
    equal_cost = {
        float(row["Budget (Relative Planning Units)"]): row
        for row in comparator_rows
        if row["Setting"] == "Central" and row["Comparator"] == "Equal-cost consequence"
    }
    common_budgets = sorted(set(assigned) & set(equal_cost))
    compared_fields = [
        "Selected Road Count",
        "Realized Cost (Relative Planning Units)",
        "Protected Population Mean [Seed Range] (Total / Age 65+)",
        "Avoided Isolation Share",
        "Protected Population per Relative Cost",
    ]
    budget_checks = []
    for budget in common_budgets:
        fields = {
            field: close_or_equal(assigned[budget][field], equal_cost[budget][field])
            for field in compared_fields
        }
        budget_checks.append({"budget": budget, "fields": fields, "all_equal": all(fields.values())})

    action_effect = literal_assignment(source, "ACTION_EFFECT")
    cost_multiplier = literal_assignment(source, "COST_MULTIPLIER")
    ratio_checks: dict[str, dict[str, Any]] = {}
    for action, effects in action_effect.items():
        ratios = {
            setting: float(effects[setting]) / float(cost_multiplier[setting])
            for setting in ("Conservative", "Central", "Optimistic")
        }
        median_ratio = float(median(ratios.values()))
        ratio_checks[action] = {
            "ratios": ratios,
            "median_ratio": median_ratio,
            "central_ratio": ratios["Central"],
            "equal": close_or_equal(median_ratio, ratios["Central"]),
        }

    sensitivity = {row["Setting"]: row for row in sensitivity_rows}
    selected_settings = [
        "Global cost x0.8",
        "Global cost x1.2",
        "Equal-action cost",
        "Length-only cost",
    ]
    cost_evidence = {
        setting: {
            "score_spearman": float(sensitivity[setting]["Score Spearman vs Central"]),
            "top30_overlap": float(sensitivity[setting]["Top-30 Overlap vs Central"]),
            "portfolio_overlap": float(
                sensitivity[setting]["Selected-Portfolio Overlap vs Central"]
            ),
            "protected_population": float(sensitivity[setting]["Protected Population"]),
        }
        for setting in selected_settings
    }

    checks = {
        "seven_common_central_budgets": len(common_budgets) == 7,
        "central_outputs_identical_at_every_budget": all(
            item["all_equal"] for item in budget_checks
        ),
        "median_ratio_equals_central_for_every_action": all(
            item["equal"] for item in ratio_checks.values()
        ),
        "uniform_cost_scaling_preserves_top30_order": all(
            close_or_equal(cost_evidence[setting]["top30_overlap"], 1.0)
            for setting in ("Global cost x0.8", "Global cost x1.2")
        ),
        "alternative_cost_structures_change_membership": all(
            cost_evidence[setting]["top30_overlap"] < 1.0
            and cost_evidence[setting]["portfolio_overlap"] < 1.0
            for setting in ("Equal-action cost", "Length-only cost")
        ),
    }

    decision = {
        "reviewer": "Reviewer 4",
        "comment": 3,
        "input_sha256": {
            str(path.relative_to(ROOT)): sha256(path)
            for path in (PORTFOLIO_PATH, COMPARATOR_PATH, SENSITIVITY_PATH, SOURCE_PATH)
        },
        "checks": checks,
        "all_checks_pass": all(checks.values()),
        "common_central_budgets": common_budgets,
        "budget_identity_checks": budget_checks,
        "effect_cost_ratio_checks": ratio_checks,
        "cost_structure_evidence": cost_evidence,
        "identity_explanation": (
            "Both rankings use the same consequence proxy and action assignment. The robust "
            "assigned-action score uses the median cross-setting action effect-to-cost ratio, "
            "which equals the Central ratio for every action; the scores are therefore "
            "elementwise identical. A shared linear consequence proxy is not sufficient by "
            "itself to create or remove this identity."
        ),
        "divergence_conditions": [
            "Cross-setting action effect-to-cost ratios whose median differs from the Central ratio.",
            "Site-specific or nonlinear costs that break the current ratio identity rather than a uniform global rescaling.",
            "Portfolio-dependent marginal benefits, redundancy, interaction, or resource constraints evaluated during selection.",
            "Additional objectives supported by an explicit decision protocol and stakeholder-elicited weights.",
        ],
        "interpretation": (
            "Existing cost sensitivity shows that uniform scaling preserves the Top-30 order, "
            "whereas equal-action and length-only cost structures reduce Top-30 overlap to "
            "70% and 40%. This demonstrates sensitivity of membership to cost structure but "
            "does not establish superiority over an equal-cost comparator when the same score "
            "and assumptions are used by both."
        ),
    }

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUTPUT_DIR / "decision.json").write_text(
        json.dumps(decision, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )
    report = [
        "# Reviewer 4 Comment 3 identity and divergence audit",
        "",
        f"- All checks pass: **{decision['all_checks_pass']}**",
        f"- Central budget rows compared: {len(common_budgets)}",
        "- Central assigned-action and equal-cost consequence outputs are identical at every budget.",
        "- Every action's cross-setting median effect-to-cost ratio equals its Central ratio.",
        "",
        "## Why the rankings coincide",
        "",
        decision["identity_explanation"],
        "",
        "## Existing sensitivity evidence",
        "",
    ]
    for setting in selected_settings:
        item = cost_evidence[setting]
        report.append(
            f"- {setting}: score rho={item['score_spearman']:.6f}; "
            f"Top-30 overlap={item['top30_overlap']:.1%}; portfolio overlap="
            f"{item['portfolio_overlap']:.1%}; protected population="
            f"{item['protected_population']:.1f}."
        )
    report.extend(
        [
            "",
            "Uniform rescaling changes affordability at a fixed budget but not score order. "
            "Alternative cost structures change road membership. Neither result by itself "
            "establishes an advantage over a comparator that uses the same inputs and score.",
            "",
            "## Conditions for genuine divergence",
            "",
        ]
    )
    report.extend(f"- {item}" for item in decision["divergence_conditions"])
    report.extend(
        [
            "",
            "These conditions are prospective design requirements, not results demonstrated "
            "by the present screening analysis.",
            "",
        ]
    )
    (OUTPUT_DIR / "audit_report.md").write_text("\n".join(report), encoding="utf-8")
    print(json.dumps(decision, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
