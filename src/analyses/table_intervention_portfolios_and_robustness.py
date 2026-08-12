#!/usr/bin/env python3
"""Intervention Portfolios and Robustness.

Plan: Compare seven budget-feasible road-access intervention portfolios under
Conservative, Central, and Optimistic cost-effect assumptions.
Framework: AnaSOP Sections 5-7 use the accepted Heavy-scenario closure mapping,
consequence-aware assigned-action road ranking, action-specific proportional effects,
relative planning costs, five independently seeded sets of 1,000 paired network draws, avoided
isolation, protected population, and selection overlap. Results are planning
screening outputs rather than engineering optima or guaranteed benefits.
"""
from __future__ import annotations

import os
from pathlib import Path

os.environ.pop("PROJ_LIB", None)
os.environ.pop("PROJ_DATA", None)

import numpy as np
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd
from PIL import Image, ImageDraw

import figure_intervention_priorities_and_budgeted_benefits as intervention
import figure_community_isolation_frequency_and_exposed_population as isolation
import table_priority_road_sections as priority_roads


ROOT = Path(__file__).resolve().parents[2]
PORTFOLIO_OUT = ROOT / "data/results/tables/Table_intervention_portfolios.xlsx"
COMPARATOR_OUT = ROOT / "data/results/tables/Table_comparator_robustness.xlsx"
PREVIEW_OUT = ROOT / "data/exp/table_previews/Table_intervention_portfolios.png"
COMPARATOR_PREVIEW_OUT = ROOT / "data/exp/table_previews/Table_comparator_robustness.png"
SHEET_NAME = "Portfolio Robustness"
COMPARATOR_SHEET = "Comparator Baselines"
TABLE_TITLE = "Intervention Portfolios"
COMPARATOR_TITLE = "Comparator Robustness"
SETTINGS = ("Conservative", "Central", "Optimistic")
BUDGET_COUNT = 7
PORTFOLIO_CANDIDATES = 150


def portfolio_selections(
    priority_order: np.ndarray,
    base_cost: np.ndarray,
    budgets: np.ndarray,
) -> dict[tuple[str, int], tuple[np.ndarray, float]]:
    """Return greedy assigned-action portfolios and realized setting-specific spend."""
    positions = priority_order[:PORTFOLIO_CANDIDATES]
    selections: dict[tuple[str, int], tuple[np.ndarray, float]] = {}
    for setting in SETTINGS:
        setting_cost = base_cost * intervention.COST_MULTIPLIER[setting]
        for budget_index, budget in enumerate(budgets):
            selected: list[int] = []
            spent = 0.0
            for position in positions:
                cost = float(setting_cost[position])
                if spent + cost <= budget + 1e-9:
                    selected.append(int(position))
                    spent += cost
            selections[(setting, budget_index)] = (
                np.asarray(selected, dtype="int32"),
                spent,
            )
    return selections


def build_table() -> tuple[pd.DataFrame, dict[str, object]]:
    """Build 21 budget-setting portfolio rows from the accepted road ranking."""
    _, context = priority_roads.build_table()
    priority_order = context["Priority Order"]
    base_cost = context["Base Cost"]
    actions = context["Actions"]
    section_propensity = context["Section Propensity"]
    community_population = context["Community Population"]
    community_older = context["Community Population Age 65+"]
    baseline_frequencies = {
        int(seed): np.asarray(frequency, dtype=float)
        for seed, frequency in context["Baseline Frequencies"].items()
    }
    baseline_frequency = np.mean(np.vstack(list(baseline_frequencies.values())), axis=0)
    baseline_expected = float(np.sum(community_population * baseline_frequency))
    portfolio_positions = priority_order[:PORTFOLIO_CANDIDATES]
    max_budget = float(base_cost[portfolio_positions[:100]].sum())
    budgets = np.linspace(0.0, max_budget, BUDGET_COUNT)
    selections = portfolio_selections(priority_order, base_cost, budgets)

    rows: list[dict[str, object]] = []
    for budget_index, budget in enumerate(budgets):
        central_selected = set(selections[("Central", budget_index)][0].tolist())
        for setting in SETTINGS:
            selected, spent = selections[(setting, budget_index)]
            adjusted_propensity = section_propensity.copy()
            if selected.size:
                effects = np.asarray(
                    [intervention.ACTION_EFFECT[str(actions[position])][setting] for position in selected],
                    dtype=float,
                )
                adjusted_propensity[selected] *= 1.0 - effects
            protected_by_seed: list[float] = []
            protected_older_by_seed: list[float] = []
            reduction_by_seed: list[np.ndarray] = []
            for seed in isolation.REPLICATE_SEEDS:
                adjusted_frequency = intervention.cached_intervention_frequency(
                    f"assigned_{setting.lower()}_b{budget_index}_seed_{seed}",
                    context["Candidate U"],
                    context["Candidate V"],
                    context["Candidate Edge Section"],
                    adjusted_propensity,
                    int(context["Root Count"]),
                    context["Target Roots"],
                    context["Attachment Community"],
                    context["Attachment Root"],
                    len(community_population),
                    seed,
                ).astype(float)
                reduction = np.maximum(baseline_frequencies[seed] - adjusted_frequency, 0.0)
                reduction_by_seed.append(reduction)
                protected_by_seed.append(float(np.sum(community_population * reduction)))
                protected_older_by_seed.append(float(np.sum(community_older * reduction)))
            protected_population = float(np.mean(protected_by_seed))
            protected_older = float(np.mean(protected_older_by_seed))
            frequency_reduction = np.mean(np.vstack(reduction_by_seed), axis=0)
            protected_communities = int(np.count_nonzero(frequency_reduction >= 0.001))
            selected_set = set(selected.tolist())
            union = selected_set | central_selected
            overlap = (
                len(selected_set & central_selected) / len(union)
                if union
                else 1.0
            )
            action_counts = pd.Series(actions[selected], dtype="string").value_counts()
            mix = " / ".join(
                str(
                    int(
                        action_counts.get(
                            action,
                            0,
                        )
                    )
                )
                for action in (
                    "Temporary reinforcement",
                    "Clearance pre-positioning",
                    "Alternative-route protection",
                )
            )
            rows.append(
                {
                    "Sensitivity Setting": setting,
                    "Budget (Relative Planning Units)": float(budget),
                    "Selected Road Count": int(selected.size),
                    "Intervention Mix (Reinforcement / Clearance / Alternative)": mix,
                    "Realized Cost (Relative Planning Units)": float(spent),
                    "Protected Community Count": protected_communities,
                    "Protected Population Mean [Seed Range] (Total / Age 65+)": (
                        f"{protected_population:,.1f} [{min(protected_by_seed):,.1f}–{max(protected_by_seed):,.1f}] / "
                        f"{protected_older:,.1f} [{min(protected_older_by_seed):,.1f}–{max(protected_older_by_seed):,.1f}]"
                    ),
                    "Avoided Isolation Share": (
                        protected_population / baseline_expected
                        if baseline_expected > 0
                        else np.nan
                    ),
                    "Protected Population per Relative Cost": (
                        protected_population / spent if spent > 0 else np.nan
                    ),
                    "Selection Overlap vs Central": float(overlap),
                }
            )

    table = pd.DataFrame(rows)
    if table.shape != (BUDGET_COUNT * len(SETTINGS), 10):
        raise RuntimeError(f"Unexpected portfolio-table shape: {table.shape}.")
    comparator = intervention.evaluate_comparator_portfolios(
        budgets,
        base_cost,
        actions,
        context["Candidate Score"],
        context["Emergency Candidate"],
        context["Candidate Road Category"],
        context["Consequence Proxy"],
        section_propensity,
        baseline_frequencies,
        community_population,
        community_older,
        context["Candidate U"],
        context["Candidate V"],
        context["Candidate Edge Section"],
        int(context["Root Count"]),
        context["Target Roots"],
        context["Attachment Community"],
        context["Attachment Root"],
    )
    comparator["Protected Population Mean [Seed Range] (Total / Age 65+)"] = comparator.apply(
        lambda row: (
            f"{row['Protected Population']:,.1f} "
            f"[{row['Protected Population Low']:,.1f}–{row['Protected Population High']:,.1f}] / "
            f"{row['Protected Population Age 65+']:,.1f} "
            f"[{row['Protected Population Age 65+ Low']:,.1f}–{row['Protected Population Age 65+ High']:,.1f}]"
        ),
        axis=1,
    )
    comparator["Avoided Isolation Share"] = comparator["Protected Population"] / baseline_expected
    comparator["Protected Population per Relative Cost"] = np.divide(
        comparator["Protected Population"],
        comparator["Realized Portfolio Cost"],
        out=np.full(len(comparator), np.nan, dtype=float),
        where=comparator["Realized Portfolio Cost"].to_numpy(dtype=float) > 0,
    )
    comparator = comparator.rename(
        columns={
            "Budget (Planning Units)": "Budget (Relative Planning Units)",
            "Realized Portfolio Cost": "Realized Cost (Relative Planning Units)",
        }
    )
    comparator = comparator.loc[
        :,
        [
            "Comparator",
            "Setting",
            "Budget (Relative Planning Units)",
            "Selected Road Count",
            "Realized Cost (Relative Planning Units)",
            "Protected Population Mean [Seed Range] (Total / Age 65+)",
            "Avoided Isolation Share",
            "Protected Population per Relative Cost",
        ],
    ]
    diagnostics = {
        "Baseline Expected Isolation": baseline_expected,
        "Maximum Budget": max_budget,
        "Maximum Protected Population": float(
            table["Protected Population Mean [Seed Range] (Total / Age 65+)"]
            .str.split(" ")
            .str[0]
            .str.replace(",", "", regex=False)
            .astype(float)
            .max()
        ),
        "Model Mode": context["Model Mode"],
        "Comparator Table": comparator,
    }
    return table, diagnostics


def style_workbooks(portfolio_path: Path, comparator_path: Path) -> None:
    """Apply compact Word-safe formatting to the two single-sheet workbooks."""
    workbook = load_workbook(portfolio_path)
    worksheet = workbook[SHEET_NAME]
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "B3"
    worksheet.auto_filter.ref = f"A2:J{worksheet.max_row}"
    worksheet.sheet_view.zoomScale = 90
    worksheet.print_area = f"A1:J{worksheet.max_row}"
    worksheet.print_title_rows = None
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A3
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 1
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_margins = PageMargins(
        left=0.20, right=0.20, top=0.30, bottom=0.30, header=0.10, footer=0.10
    )

    header_fill = PatternFill("solid", fgColor="17365D")
    header_font = Font(name="Aptos", size=9, bold=True, color="FFFFFF")
    body_font = Font(name="Aptos", size=9, color="172033")
    subtle_border = Border(bottom=Side(style="thin", color="D0D5DD"))
    setting_fill = {
        "Conservative": PatternFill("solid", fgColor="DCE9FA"),
        "Central": PatternFill("solid", fgColor="FCE5D9"),
        "Optimistic": PatternFill("solid", fgColor="D9EDE9"),
    }
    worksheet.merge_cells("A1:J1")
    title_cell = worksheet["A1"]
    title_cell.value = TABLE_TITLE
    title_cell.fill = PatternFill("solid", fgColor="D9EAF7")
    title_cell.font = Font(name="Aptos Display", size=15, bold=True, color="17365D")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    worksheet.row_dimensions[1].height = 28

    for cell in worksheet[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[2].height = 54

    for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = subtle_border
        row[0].fill = setting_fill[str(row[0].value)]
        row[1].number_format = "0.00"
        row[2].number_format = "#,##0"
        row[4].number_format = "0.00"
        row[5].number_format = "#,##0"
        row[7].number_format = "0.0%"
        row[8].number_format = "0.00"
        row[9].number_format = "0.0%"
        for cell in (*row[1:3], *row[4:6], *row[7:]):
            cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
        worksheet.row_dimensions[row[0].row].height = 29

    widths = {
        "A": 19,
        "B": 19,
        "C": 18,
        "D": 35,
        "E": 20,
        "F": 21,
        "G": 30,
        "H": 20,
        "I": 23,
        "J": 22,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width
    for column in ("H", "I", "J"):
        worksheet.conditional_formatting.add(
            f"{column}3:{column}{worksheet.max_row}",
            ColorScaleRule(
                start_type="min",
                start_color="FFFFFF",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="63BE7B",
            ),
        )
    excel_table = Table(
        displayName="InterventionPortfolioRobustness",
        ref=f"A2:J{worksheet.max_row}",
    )
    excel_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(excel_table)

    workbook.save(portfolio_path)

    workbook = load_workbook(comparator_path)
    comparator = workbook[COMPARATOR_SHEET]
    comparator.sheet_view.showGridLines = False
    comparator.freeze_panes = "C3"
    comparator.auto_filter.ref = f"A2:H{comparator.max_row}"
    comparator.page_setup.orientation = "landscape"
    comparator.page_setup.paperSize = comparator.PAPERSIZE_A3
    comparator.page_setup.fitToWidth = 1
    comparator.page_setup.fitToHeight = 0
    comparator.sheet_properties.pageSetUpPr.fitToPage = True
    comparator.print_area = f"A1:H{comparator.max_row}"
    comparator.print_title_rows = None
    comparator.merge_cells("A1:H1")
    comparator_title = comparator["A1"]
    comparator_title.value = COMPARATOR_TITLE
    comparator_title.fill = PatternFill("solid", fgColor="D9EAF7")
    comparator_title.font = Font(name="Aptos Display", size=15, bold=True, color="17365D")
    comparator_title.alignment = Alignment(horizontal="left", vertical="center")
    comparator.row_dimensions[1].height = 28
    for cell in comparator[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    comparator.row_dimensions[2].height = 48
    for row in comparator.iter_rows(min_row=3, max_row=comparator.max_row):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = subtle_border
        row[2].number_format = "0.00"
        row[3].number_format = "#,##0"
        row[4].number_format = "0.00"
        row[6].number_format = "0.0%"
        row[7].number_format = "0.00"
    for column, width in {
        "A": 27,
        "B": 17,
        "C": 20,
        "D": 18,
        "E": 21,
        "F": 31,
        "G": 20,
        "H": 24,
    }.items():
        comparator.column_dimensions[column].width = width
    comparator_table = Table(
        displayName="InterventionComparatorBaselines",
        ref=f"A2:H{comparator.max_row}",
    )
    comparator_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    comparator.add_table(comparator_table)
    workbook.save(comparator_path)


def render_preview(path: Path, output: Path = PREVIEW_OUT) -> None:
    """Render the full 21-row title-first table as one PNG."""
    workbook = load_workbook(path, data_only=True)
    values = list(workbook[SHEET_NAME].values)
    title = str(values[0][0])
    headers = [str(value) for value in values[1]]
    rows = [list(row) for row in values[2:]]
    widths = [175, 145, 140, 260, 155, 170, 255, 175, 195, 190]
    margin = 22
    title_height = 84
    header_height = 104
    row_height = 58
    table_width = sum(widths)
    image = Image.new(
        "RGB",
        (table_width + 2 * margin, title_height + header_height + row_height * len(rows) + margin),
        "white",
    )
    draw = ImageDraw.Draw(image)
    title_font = priority_roads._preview_font(28, bold=True)
    note_font = priority_roads._preview_font(16)
    header_font = priority_roads._preview_font(16, bold=True)
    body_font = priority_roads._preview_font(16)
    draw.text((margin, 16), title, font=title_font, fill="#17365D")
    draw.text(
        (margin, 52),
        "Relative planning costs and assumed action effects; Heavy-scenario screening",
        font=note_font,
        fill="#52606D",
    )
    x_positions = [margin]
    for width in widths:
        x_positions.append(x_positions[-1] + width)
    for column, header in enumerate(headers):
        x0, x1 = x_positions[column], x_positions[column + 1]
        draw.rectangle((x0, title_height, x1, title_height + header_height), fill="#17365D")
        wrapped = priority_roads._wrap_preview_text(
            draw, header, header_font, widths[column] - 14
        )
        bbox = draw.multiline_textbbox((0, 0), wrapped, font=header_font, spacing=3)
        text_height = bbox[3] - bbox[1]
        draw.multiline_text(
            ((x0 + x1) / 2, title_height + (header_height - text_height) / 2),
            wrapped,
            font=header_font,
            fill="white",
            spacing=3,
            anchor="ma",
            align="center",
        )

    setting_colours = {
        "Conservative": "#DCE9FA",
        "Central": "#FCE5D9",
        "Optimistic": "#D9EDE9",
    }
    numeric_right = {1, 2, 4, 5, 7, 8, 9}
    formats = {
        1: lambda value: f"{float(value):,.2f}",
        2: lambda value: f"{int(value):,}",
        4: lambda value: f"{float(value):,.2f}",
        5: lambda value: f"{int(value):,}",
        7: lambda value: f"{float(value):.1%}",
        8: lambda value: "—" if value is None else f"{float(value):,.2f}",
        9: lambda value: f"{float(value):.1%}",
    }
    conditional_columns = (7, 8, 9)
    ranges = {
        column: (
            min(float(row[column]) for row in rows if row[column] is not None),
            max(float(row[column]) for row in rows if row[column] is not None),
        )
        for column in conditional_columns
    }

    def scale_colour(value: float, low: float, high: float) -> str:
        ratio = 0.5 if high == low else (value - low) / (high - low)
        if ratio <= 0.5:
            local, start, end = ratio / 0.5, (255, 255, 255), (255, 235, 132)
        else:
            local, start, end = (ratio - 0.5) / 0.5, (255, 235, 132), (99, 190, 123)
        rgb = tuple(round(a + (b - a) * local) for a, b in zip(start, end))
        return "#" + "".join(f"{channel:02X}" for channel in rgb)

    for row_number, row in enumerate(rows):
        y0 = title_height + header_height + row_number * row_height
        y1 = y0 + row_height
        base_fill = "#FFFFFF" if row_number % 2 == 0 else "#F8FAFC"
        for column, value in enumerate(row):
            x0, x1 = x_positions[column], x_positions[column + 1]
            fill = setting_colours[str(row[0])] if column == 0 else base_fill
            if column in conditional_columns and value is not None:
                low, high = ranges[column]
                fill = scale_colour(float(value), low, high)
            draw.rectangle((x0, y0, x1, y1), fill=fill, outline="#D0D5DD", width=1)
            display_value = formats.get(column, lambda item: "" if item is None else str(item))(value)
            wrapped = priority_roads._wrap_preview_text(
                draw, display_value, body_font, widths[column] - 14
            )
            if column in numeric_right:
                draw.multiline_text(
                    (x1 - 7, y0 + 8),
                    wrapped,
                    font=body_font,
                    fill="#172033",
                    spacing=2,
                    anchor="ra",
                    align="right",
                )
            else:
                draw.multiline_text(
                    (x0 + 7, y0 + 8),
                    wrapped,
                    font=body_font,
                    fill="#172033",
                    spacing=2,
                )
    output.parent.mkdir(parents=True, exist_ok=True)
    image.save(output, optimize=True)


def render_comparator_preview(
    path: Path,
    output: Path = COMPARATOR_PREVIEW_OUT,
) -> None:
    """Render all 84 matched comparator rows as one readable continuous PNG."""
    workbook = load_workbook(path, data_only=True)
    values = list(workbook[COMPARATOR_SHEET].values)
    title = str(values[0][0])
    headers = [str(value) for value in values[1]]
    rows = [list(row) for row in values[2:]]
    widths = [235, 145, 175, 145, 190, 260, 175, 210]
    margin = 22
    title_height = 84
    header_height = 100
    row_height = 48
    table_width = sum(widths)
    image = Image.new(
        "RGB",
        (table_width + 2 * margin, title_height + header_height + row_height * len(rows) + margin),
        "white",
    )
    draw = ImageDraw.Draw(image)
    title_font = priority_roads._preview_font(28, bold=True)
    note_font = priority_roads._preview_font(16)
    header_font = priority_roads._preview_font(15, bold=True)
    body_font = priority_roads._preview_font(15)
    draw.text((margin, 16), title, font=title_font, fill="#17365D")
    draw.text(
        (margin, 52),
        "Four comparator rankings under matched costs and effects; Heavy-scenario screening",
        font=note_font,
        fill="#52606D",
    )
    x_positions = [margin]
    for width in widths:
        x_positions.append(x_positions[-1] + width)
    for column, header in enumerate(headers):
        x0, x1 = x_positions[column], x_positions[column + 1]
        draw.rectangle((x0, title_height, x1, title_height + header_height), fill="#17365D")
        wrapped = priority_roads._wrap_preview_text(
            draw, header, header_font, widths[column] - 14
        )
        bbox = draw.multiline_textbbox((0, 0), wrapped, font=header_font, spacing=3)
        text_height = bbox[3] - bbox[1]
        draw.multiline_text(
            ((x0 + x1) / 2, title_height + (header_height - text_height) / 2),
            wrapped,
            font=header_font,
            fill="white",
            spacing=3,
            anchor="ma",
            align="center",
        )

    setting_colours = {
        "Conservative": "#DCE9FA",
        "Central": "#FCE5D9",
        "Optimistic": "#D9EDE9",
    }
    comparator_colours = {
        "Hazard only": "#F2F4F7",
        "Emergency route only": "#F3E8FF",
        "Road class only": "#FFF4D6",
        "Equal-cost consequence": "#E5E7EB",
    }
    numeric_right = {2, 3, 4, 6, 7}
    formats = {
        2: lambda value: f"{float(value):,.2f}",
        3: lambda value: f"{int(value):,}",
        4: lambda value: f"{float(value):,.2f}",
        6: lambda value: f"{float(value):.1%}",
        7: lambda value: "—" if value is None else f"{float(value):,.3f}",
    }
    for row_number, row in enumerate(rows):
        y0 = title_height + header_height + row_number * row_height
        y1 = y0 + row_height
        for column, value in enumerate(row):
            x0, x1 = x_positions[column], x_positions[column + 1]
            if column == 0:
                fill = comparator_colours[str(row[0])]
            elif column == 1:
                fill = setting_colours[str(row[1])]
            else:
                fill = "#FFFFFF" if row_number % 2 == 0 else "#F8FAFC"
            draw.rectangle((x0, y0, x1, y1), fill=fill, outline="#D0D5DD", width=1)
            display_value = formats.get(
                column,
                lambda item: "" if item is None else str(item),
            )(value)
            wrapped = priority_roads._wrap_preview_text(
                draw, display_value, body_font, widths[column] - 14
            )
            if column in numeric_right:
                draw.multiline_text(
                    (x1 - 7, y0 + 7),
                    wrapped,
                    font=body_font,
                    fill="#172033",
                    spacing=2,
                    anchor="ra",
                    align="right",
                )
            else:
                draw.multiline_text(
                    (x0 + 7, y0 + 7),
                    wrapped,
                    font=body_font,
                    fill="#172033",
                    spacing=2,
                )
    output.parent.mkdir(parents=True, exist_ok=True)
    image.save(output, optimize=True)


def verify_workbook(path: Path, sheet_name: str, rows: int, columns: int) -> None:
    """Verify one Word-safe workbook, dimensions, and metric ranges."""
    workbook = load_workbook(path, data_only=False)
    if workbook.sheetnames != [sheet_name]:
        raise RuntimeError(f"Unexpected workbook sheets: {workbook.sheetnames}")
    worksheet = workbook[sheet_name]
    if worksheet.max_row != rows or worksheet.max_column != columns:
        raise RuntimeError(
            f"Unexpected workbook dimensions: {worksheet.max_row} × {worksheet.max_column}."
        )
    error_tokens = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"}
    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value in error_tokens:
                raise RuntimeError(f"Spreadsheet error token in {cell.coordinate}: {cell.value}")
    if worksheet["A1"].value not in {TABLE_TITLE, COMPARATOR_TITLE}:
        raise RuntimeError("Workbook title row is missing or incorrect.")
    share_columns = (8, 10) if sheet_name == SHEET_NAME else (7,)
    for row in range(3, worksheet.max_row + 1):
        for column in share_columns:
            value = worksheet.cell(row, column).value
            if value is not None and not 0 <= float(value) <= 1:
                raise RuntimeError(f"Share outside [0, 1] at row {row}, column {column}.")


def main() -> None:
    table, diagnostics = build_table()
    comparator = diagnostics["Comparator Table"]
    PORTFOLIO_OUT.parent.mkdir(parents=True, exist_ok=True)
    table.to_excel(
        PORTFOLIO_OUT,
        index=False,
        sheet_name=SHEET_NAME,
        engine="openpyxl",
        startrow=1,
    )
    comparator.to_excel(
        COMPARATOR_OUT,
        index=False,
        sheet_name=COMPARATOR_SHEET,
        engine="openpyxl",
        startrow=1,
    )
    style_workbooks(PORTFOLIO_OUT, COMPARATOR_OUT)
    verify_workbook(PORTFOLIO_OUT, SHEET_NAME, BUDGET_COUNT * len(SETTINGS) + 2, 10)
    verify_workbook(COMPARATOR_OUT, COMPARATOR_SHEET, len(comparator) + 2, 8)
    render_preview(PORTFOLIO_OUT)
    render_comparator_preview(COMPARATOR_OUT)
    print(f"Saved: {PORTFOLIO_OUT.relative_to(ROOT)}")
    print(f"Saved: {COMPARATOR_OUT.relative_to(ROOT)}")
    print(f"Preview: {PREVIEW_OUT.relative_to(ROOT)}")
    print(f"Preview: {COMPARATOR_PREVIEW_OUT.relative_to(ROOT)}")
    print(f"Rows: {len(table):,}; columns: {len(table.columns):,}")
    print(
        f"Heavy baseline expected isolated population: "
        f"{diagnostics['Baseline Expected Isolation']:,.1f}"
    )
    print(f"Maximum budget: {diagnostics['Maximum Budget']:,.2f} planning units")
    print(
        f"Maximum protected population: "
        f"{diagnostics['Maximum Protected Population']:,.1f}"
    )
    print(f"Terrain-score construction: {diagnostics['Model Mode']}")
    print("Workbook verification: passed")


if __name__ == "__main__":
    main()
