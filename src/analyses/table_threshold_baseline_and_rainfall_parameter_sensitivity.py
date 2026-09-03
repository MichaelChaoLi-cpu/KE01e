#!/usr/bin/env python3
"""Build Appendix Table B5 from validated Reviewer 2 Comment 3/4 outputs.

The publication-facing sheet is intentionally compact. A hidden Evidence sheet
stores the linked numerical inputs, source rows, and source-file hashes so that
the formulas in the displayed table remain auditable.
"""
from __future__ import annotations

import csv
import hashlib
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[2]
COMMENT3 = ROOT / "data/exp/revision/reviewer-2-comment-3"
COMMENT4 = ROOT / "data/exp/revision/reviewer-2-comment-4"
OUT = (
    ROOT
    / "data/results/tables/Table_threshold_baseline_and_rainfall_parameter_sensitivity.xlsx"
)

SUMMARY_SHEET = "Table B5"
EVIDENCE_SHEET = "Evidence"
TITLE = "Table B5. Baseline-Threshold Comparison and Rainfall-Parameter Sensitivity"
HEADERS = (
    "Analysis",
    "Outcome",
    "Reference estimate",
    "Alternative estimate or range",
    "Relative effect",
    "Stability or support",
    "Interpretation",
)


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def exactly_one(rows: Iterable[dict[str, str]], **criteria: object) -> dict[str, str]:
    matches = [
        row
        for row in rows
        if all(str(row.get(key)) == str(value) for key, value in criteria.items())
    ]
    if len(matches) != 1:
        raise RuntimeError(f"Expected one row for {criteria}, found {len(matches)}.")
    return matches[0]


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def add_evidence(
    worksheet,
    *,
    analysis: str,
    outcome: str,
    setting: str,
    value: float,
    unit: str,
    source: Path,
    source_key: str,
) -> int:
    row = worksheet.max_row + 1
    worksheet.append(
        [
            analysis,
            outcome,
            setting,
            float(value),
            unit,
            source.relative_to(ROOT).as_posix(),
            source_key,
            sha256(source),
        ]
    )
    return row


def build_workbook() -> None:
    isolation_path = COMMENT3 / "community_isolation_threshold_comparison.csv"
    service_path = COMMENT3 / "service_reachability_threshold_comparison.csv"
    score_path = COMMENT3 / "slope_and_road_threshold_comparison.csv"
    road_sensitivity_path = COMMENT4 / "road_score_sensitivity.csv"
    matched_path = COMMENT4 / "matched_validation_and_ordering.csv"
    isolation_sensitivity_path = COMMENT4 / "community_isolation_parameter_sensitivity.csv"
    compatibility_path = COMMENT4 / "weight_scheme_jma_compatibility.csv"

    isolation = read_csv(isolation_path)
    service = read_csv(service_path)
    scores = read_csv(score_path)
    road_sensitivity = read_csv(road_sensitivity_path)
    matched = read_csv(matched_path)
    isolation_sensitivity = read_csv(isolation_sensitivity_path)
    compatibility = read_csv(compatibility_path)

    workbook = Workbook()
    summary = workbook.active
    summary.title = SUMMARY_SHEET
    evidence = workbook.create_sheet(EVIDENCE_SHEET)

    evidence.append(["Validated numerical inputs for Table B5"])
    evidence.append(
        [
            "Analysis",
            "Outcome",
            "Setting",
            "Value",
            "Unit",
            "Source file",
            "Source row key",
            "Source SHA-256",
        ]
    )

    evidence_rows: dict[str, int] = {}

    heavy_road = exactly_one(scores, analysis_level="Road section", scenario="Heavy")
    evidence_rows["road_rho"] = add_evidence(
        evidence,
        analysis="Threshold geography",
        outcome="Heavy road-section score ranking",
        setting="Baseline f=1.00 versus official geography",
        value=float(heavy_road["spearman_rho"]),
        unit="Spearman rho",
        source=score_path,
        source_key="analysis_level=Road section; scenario=Heavy",
    )
    evidence_rows["road_top1"] = add_evidence(
        evidence,
        analysis="Threshold geography",
        outcome="Heavy road-section score ranking",
        setting="Baseline f=1.00 versus official geography",
        value=float(heavy_road["top1_overlap"]),
        unit="fraction",
        source=score_path,
        source_key="analysis_level=Road section; scenario=Heavy",
    )

    for scenario in ("Moderate", "Heavy", "Extreme"):
        for setting, key in (
            ("Baseline f=1.00", "baseline"),
            ("Official geography", "official"),
        ):
            record = exactly_one(isolation, threshold_geography=setting, scenario=scenario)
            evidence_rows[f"iso_{scenario.lower()}_{key}"] = add_evidence(
                evidence,
                analysis="Threshold geography",
                outcome=f"{scenario} expected isolated population",
                setting=setting,
                value=float(record["expected_isolated_population_mean"]),
                unit="persons",
                source=isolation_path,
                source_key=f"threshold_geography={setting}; scenario={scenario}",
            )

    service_names = ("Shelter", "Fire service", "Municipal facility", "Emergency water")
    service_keys = {
        "Shelter": "shelter",
        "Fire service": "fire",
        "Municipal facility": "municipal",
        "Emergency water": "water",
    }
    for service_name in service_names:
        for setting, key in (
            ("Baseline f=1.00", "baseline"),
            ("Official geography", "official"),
        ):
            record = exactly_one(
                service, threshold_geography=setting, service_class=service_name
            )
            evidence_rows[f"service_{service_keys[service_name]}_{key}"] = add_evidence(
                evidence,
                analysis="Threshold geography",
                outcome=f"Heavy {service_name.lower()} reachability loss",
                setting=setting,
                value=float(record["expected_population_losing_reachability_mean"]),
                unit="persons",
                source=service_path,
                source_key=f"threshold_geography={setting}; service_class={service_name}",
            )

    evidence_rows["sensitivity_min_rho"] = add_evidence(
        evidence,
        analysis="Rainfall parameters",
        outcome="Road-priority ranking",
        setting="Minimum across five weights x three gamma values x three scenarios",
        value=min(float(row["central_rank_spearman_rho"]) for row in road_sensitivity),
        unit="Spearman rho",
        source=road_sensitivity_path,
        source_key="minimum central_rank_spearman_rho across 45 rows",
    )
    evidence_rows["sensitivity_min_top1"] = add_evidence(
        evidence,
        analysis="Rainfall parameters",
        outcome="Road-priority ranking",
        setting="Minimum across five weights x three gamma values x three scenarios",
        value=min(float(row["central_top1_overlap"]) for row in road_sensitivity),
        unit="fraction",
        source=road_sensitivity_path,
        source_key="minimum central_top1_overlap across 45 rows",
    )

    matched_central = exactly_one(matched, weight_scheme="equal", gamma="1.0")
    for key, value, setting in (
        ("matched_central", matched_central["matched_concordance"], "Equal weights; gamma=1.0"),
        ("matched_min", min(float(row["matched_concordance"]) for row in matched), "Minimum across 15 combinations"),
        ("matched_max", max(float(row["matched_concordance"]) for row in matched), "Maximum across 15 combinations"),
    ):
        evidence_rows[key] = add_evidence(
            evidence,
            analysis="Rainfall parameters",
            outcome="Matched road-evidence concordance",
            setting=setting,
            value=float(value),
            unit="concordance",
            source=matched_path,
            source_key=setting,
        )

    isolation_central = exactly_one(isolation_sensitivity, weight_scheme="equal", gamma="1.0")
    for key, value, setting, unit in (
        ("isolation_central", isolation_central["expected_isolated_population_mean"], "Equal weights; gamma=1.0", "persons"),
        ("isolation_min", min(float(row["expected_isolated_population_mean"]) for row in isolation_sensitivity), "Minimum across five prespecified network settings", "persons"),
        ("isolation_max", max(float(row["expected_isolated_population_mean"]) for row in isolation_sensitivity), "Maximum across five prespecified network settings", "persons"),
        ("isolation_ratio_min", min(float(row["ratio_to_central"]) for row in isolation_sensitivity), "Minimum ratio to central", "ratio"),
        ("isolation_ratio_max", max(float(row["ratio_to_central"]) for row in isolation_sensitivity), "Maximum ratio to central", "ratio"),
    ):
        evidence_rows[key] = add_evidence(
            evidence,
            analysis="Rainfall parameters",
            outcome="Heavy expected isolated population",
            setting=setting,
            value=float(value),
            unit=unit,
            source=isolation_sensitivity_path,
            source_key=setting,
        )

    compatibility_central = exactly_one(compatibility, weight_scheme="equal")
    for key, value, setting in (
        ("compatibility_central", compatibility_central["pooled_spearman_rho"], "Equal weights"),
        ("compatibility_min", min(float(row["pooled_spearman_rho"]) for row in compatibility), "Minimum across five weight schemes"),
        ("compatibility_max", max(float(row["pooled_spearman_rho"]) for row in compatibility), "Maximum across five weight schemes"),
        ("compatibility_events", compatibility_central["eligible_events"], "Eligible independent rainfall events"),
    ):
        evidence_rows[key] = add_evidence(
            evidence,
            analysis="Rainfall parameters",
            outcome="JMA-type indicator compatibility",
            setting=setting,
            value=float(value),
            unit="events" if key.endswith("events") else "Spearman rho",
            source=compatibility_path,
            source_key=setting,
        )

    summary.append([TITLE])
    summary.append(list(HEADERS))

    def ref(key: str) -> str:
        return f"='{EVIDENCE_SHEET}'!D{evidence_rows[key]}"

    summary_rows = [
        [
            "Threshold geography",
            "Heavy road-section score ranking",
            "All-area f = 1.00",
            "Corrected official geography",
            f'=\"Spearman rho = \"&TEXT({ref("road_rho")[1:]},\"0.000\")',
            f'=\"Top 1% overlap = \"&TEXT({ref("road_top1")[1:]},\"0.0%\")',
            "Rank order is largely preserved",
        ],
    ]
    for scenario in ("Moderate", "Heavy", "Extreme"):
        lower = scenario.lower()
        summary_rows.append(
            [
                "Threshold geography",
                f"{scenario} expected isolated population (persons)",
                ref(f"iso_{lower}_baseline"),
                ref(f"iso_{lower}_official"),
                f"=(D{len(summary_rows)+3}-C{len(summary_rows)+3})/C{len(summary_rows)+3}",
                "5 seeds × 1,000 draws",
                "Magnitude increases under official retention factors",
            ]
        )
    for service_name, key in (
        ("Shelter", "shelter"),
        ("Fire service", "fire"),
        ("Municipal facility", "municipal"),
        ("Emergency water*", "water"),
    ):
        summary_rows.append(
            [
                "Threshold geography",
                f"Heavy {service_name.lower()} reachability loss (persons)",
                ref(f"service_{key}_baseline"),
                ref(f"service_{key}_official"),
                f"=(D{len(summary_rows)+3}-C{len(summary_rows)+3})/C{len(summary_rows)+3}",
                "5 seeds × 1,000 draws",
                "Conditional sensitivity" if key == "water" else "Magnitude increases",
            ]
        )
    summary_rows.extend(
        [
            [
                "Rainfall parameters",
                "Road-priority ranking (45 score comparisons)",
                "Equal weights; gamma = 1.0",
                "Five weights × three gamma values × three scenarios",
                f'=\"Minimum rho = \"&TEXT({ref("sensitivity_min_rho")[1:]},\"0.000\")',
                f'=\"Minimum Top 1% overlap = \"&TEXT({ref("sensitivity_min_top1")[1:]},\"0.0%\")',
                "High ranking stability",
            ],
            [
                "Rainfall parameters",
                "Matched road-evidence concordance",
                ref("matched_central"),
                f'=TEXT({ref("matched_min")[1:]},\"0.000\")&\"–\"&TEXT({ref("matched_max")[1:]},\"0.000\")',
                "Range across 15 combinations",
                "93 matched road sections",
                "Validation signal is stable",
            ],
            [
                "Rainfall parameters",
                "Heavy expected isolated population (persons)",
                ref("isolation_central"),
                f'=TEXT({ref("isolation_min")[1:]},\"#,##0.0\")&\"–\"&TEXT({ref("isolation_max")[1:]},\"#,##0.0\")',
                f'=TEXT({ref("isolation_ratio_min")[1:]},\"0.000\")&\"–\"&TEXT({ref("isolation_ratio_max")[1:]},\"0.000\")&\" × central\"',
                "5 prespecified settings; 5 seeds each",
                "Consequence magnitude is sensitive",
            ],
            [
                "Rainfall parameters",
                "JMA-type indicator compatibility (pooled Spearman rho)",
                ref("compatibility_central"),
                f'=TEXT({ref("compatibility_min")[1:]},\"0.000\")&\"–\"&TEXT({ref("compatibility_max")[1:]},\"0.000\")',
                "Range across five weight schemes",
                f'=TEXT({ref("compatibility_events")[1:]},\"#,##0\")&\" eligible events\"',
                "Directional compatibility remains positive",
            ],
        ]
    )
    for row in summary_rows:
        summary.append(row)

    notes = [
        "Notes: Baseline f = 1.00 switches off temporary threshold retention; it is not a physical no-earthquake counterfactual.",
        "Corrected official geography uses official resolved factors plus the analyst-defined Yatsushiro midpoint of 0.75, bounded by separate municipality-wide 0.70 and 0.80 cases.",
        "Expected-population results are means over five deterministic seeds with 1,000 Monte Carlo draws per seed.",
        "* Emergency-water results are conditional because 10 of 36 source features were geocoded and attached to the network.",
        "The 15 rainfall-parameter combinations comprise five window-weight schemes and gamma values of 0.5, 1.0, and 2.0; the consequence range uses five prespecified boundary settings.",
    ]
    summary.append([None] * len(HEADERS))
    for note in notes:
        summary.append([note])

    style_workbook(workbook)
    add_comments(summary, evidence_rows)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    OUT.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUT)


def style_workbook(workbook: Workbook) -> None:
    summary = workbook[SUMMARY_SHEET]
    evidence = workbook[EVIDENCE_SHEET]

    navy = "17365D"
    title_blue = "D9EAF7"
    threshold_blue = "EAF2F8"
    rainfall_orange = "FCE5D4"
    stripe = "F8FAFC"
    green = "E2F0D9"
    amber = "FFF2CC"
    lavender = "EDE3F7"
    text = "172033"
    muted = "52606D"
    line = Side(style="thin", color="D0D5DD")
    vertical_line = Side(style="thin", color="E4E7EC")

    summary.merge_cells("A1:G1")
    summary["A1"].fill = PatternFill("solid", fgColor=title_blue)
    summary["A1"].font = Font(name="Aptos Display", size=15, bold=True, color=navy)
    summary["A1"].alignment = Alignment(horizontal="left", vertical="center")
    summary.row_dimensions[1].height = 40

    for cell in summary[2]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(name="Aptos", size=9.5, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    summary.row_dimensions[2].height = 42

    data_end = 2 + 12
    for row_number in range(3, data_end + 1):
        analysis = summary.cell(row_number, 1).value
        base_fill = threshold_blue if analysis == "Threshold geography" else rainfall_orange
        for column in range(1, 8):
            cell = summary.cell(row_number, column)
            cell.font = Font(name="Aptos", size=9.2, color=text)
            cell.fill = PatternFill(
                "solid", fgColor=base_fill if column == 1 else (stripe if row_number % 2 == 0 else "FFFFFF")
            )
            cell.border = Border(
                right=vertical_line if column < 7 else Side(style=None), bottom=line
            )
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for column in (3, 4, 5):
            summary.cell(row_number, column).alignment = Alignment(
                horizontal="right", vertical="center", wrap_text=True, indent=1
            )
        summary.cell(row_number, 6).alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True, indent=1
        )
        if row_number >= 11:
            for column in (3, 4, 5, 6):
                summary.cell(row_number, column).alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
        if row_number in range(4, 11):
            summary.cell(row_number, 3).number_format = "#,##0.0"
            summary.cell(row_number, 4).number_format = "#,##0.0"
            summary.cell(row_number, 5).number_format = "+0.0%;-0.0%"
        if row_number in (12, 14):
            summary.cell(row_number, 3).number_format = "0.000"
        if row_number == 13:
            summary.cell(row_number, 3).number_format = "#,##0.0"
        interpretation = str(summary.cell(row_number, 7).value)
        status_fill = (
            amber
            if "sensitive" in interpretation.lower()
            else lavender
            if "conditional" in interpretation.lower()
            else green
        )
        summary.cell(row_number, 7).fill = PatternFill("solid", fgColor=status_fill)
        summary.row_dimensions[row_number].height = 48 if row_number in (11, 12, 14) else 39

    note_start = data_end + 2
    for row_number in range(note_start, note_start + 5):
        summary.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=7)
        cell = summary.cell(row_number, 1)
        cell.font = Font(name="Aptos", size=8.5, italic=True, color=muted)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        summary.row_dimensions[row_number].height = 25 if row_number != note_start + 1 else 33

    widths = {"A": 21, "B": 39, "C": 23, "D": 30, "E": 23, "F": 29, "G": 31}
    for column, width in widths.items():
        summary.column_dimensions[column].width = width
    summary.sheet_view.showGridLines = False
    summary.freeze_panes = "C3"
    summary.auto_filter.ref = f"A2:G{data_end}"
    summary.sheet_view.zoomScale = 90
    summary.print_area = f"A1:G{note_start + 4}"
    summary.page_setup.orientation = "landscape"
    summary.page_setup.paperSize = summary.PAPERSIZE_A3
    summary.page_setup.fitToWidth = 1
    summary.page_setup.fitToHeight = 1
    summary.sheet_properties.pageSetUpPr.fitToPage = True
    summary.page_margins = PageMargins(
        left=0.25, right=0.25, top=0.30, bottom=0.30, header=0.10, footer=0.10
    )
    table = Table(displayName="ThresholdAndRainfallSensitivity", ref=f"A2:G{data_end}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    summary.add_table(table)

    evidence.merge_cells("A1:H1")
    evidence["A1"].fill = PatternFill("solid", fgColor=title_blue)
    evidence["A1"].font = Font(name="Aptos Display", size=14, bold=True, color=navy)
    for cell in evidence[2]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(name="Aptos", size=9, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in evidence.iter_rows(min_row=3, max_row=evidence.max_row):
        for cell in row:
            cell.font = Font(name="Aptos", size=8.5, color=text)
            cell.border = Border(bottom=line)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row[3].number_format = "0.000000"
    evidence_widths = {"A": 22, "B": 36, "C": 40, "D": 16, "E": 17, "F": 58, "G": 50, "H": 68}
    for column, width in evidence_widths.items():
        evidence.column_dimensions[column].width = width
    evidence.sheet_view.showGridLines = False
    evidence.freeze_panes = "A3"
    evidence.auto_filter.ref = f"A2:H{evidence.max_row}"
    evidence.sheet_state = "hidden"


def add_comments(summary, evidence_rows: dict[str, int]) -> None:
    author = "Mike Li"
    summary["C4"].comment = Comment(
        f"Linked to {EVIDENCE_SHEET}!D{evidence_rows['iso_moderate_baseline']}; source path and hash are retained on the Evidence sheet.",
        author,
    )
    summary["C12"].comment = Comment(
        f"Linked to {EVIDENCE_SHEET}!D{evidence_rows['matched_central']}; the displayed range spans all 15 prespecified parameter combinations.",
        author,
    )
    summary["C13"].comment = Comment(
        f"Linked to {EVIDENCE_SHEET}!D{evidence_rows['isolation_central']}; the boundary range uses the five prespecified network settings.",
        author,
    )


def verify_workbook() -> None:
    workbook = load_workbook(OUT, data_only=False)
    if workbook.sheetnames != [SUMMARY_SHEET, EVIDENCE_SHEET]:
        raise RuntimeError(f"Unexpected sheet order: {workbook.sheetnames}")
    summary = workbook[SUMMARY_SHEET]
    evidence = workbook[EVIDENCE_SHEET]
    if summary["A1"].value != TITLE or tuple(cell.value for cell in summary[2]) != HEADERS:
        raise RuntimeError("The title/header contract is not satisfied.")
    if summary.max_column != 7 or summary.max_row != 20:
        raise RuntimeError(f"Unexpected summary dimensions: {summary.max_row} x {summary.max_column}")
    if evidence.sheet_state != "hidden":
        raise RuntimeError("Evidence sheet must remain hidden in the publication workbook.")
    if evidence.max_row != 32:
        raise RuntimeError(f"Unexpected evidence row count: {evidence.max_row}")
    error_tokens = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and any(token in cell.value for token in error_tokens):
                    raise RuntimeError(f"Formula/error token in {worksheet.title}!{cell.coordinate}: {cell.value}")
    if not all(summary.cell(row, 5).data_type == "f" for row in range(4, 11)):
        raise RuntimeError("Expected formula-driven relative changes in E4:E10.")
    if len(summary.tables) != 1 or summary.auto_filter.ref != "A2:G14":
        raise RuntimeError("Summary table/filter range is incorrect.")
    print(f"Saved: {OUT.relative_to(ROOT)}")
    print("Summary: 12 result rows, 7 columns, 5 notes, one-page A3 landscape")
    print(f"Evidence: {evidence.max_row - 2} linked values with source paths and SHA-256 hashes")
    print("Workbook verification: passed")


def main() -> None:
    build_workbook()
    verify_workbook()


if __name__ == "__main__":
    main()
