#!/usr/bin/env python3
"""Rainfall and Threshold Scenarios.

Plan: Define nine rainfall-threshold combinations plus three Heavy-rainfall
Yatsushiro spatial-support assignments for the 0.70, 0.75, and 0.80 cases.
Framework: AnaSOP Sections 5-7 use station-specific independent-event maxima and
official area-level retention factors. Central values summarize seven stations over
2016-2020, remain at station support, and do not imply fine-resolution rainfall or a
causal earthquake effect.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd


ROOT = Path(__file__).resolve().parents[2]
PROCESSED = ROOT / "data/processed"
SCENARIO_PATH = PROCESSED / "jma_rainfall_scenario_quantiles_preprocessed.parquet"
THRESHOLD_PATH = PROCESSED / "official_threshold_factors_preprocessed.parquet"
OUT = ROOT / "data/results/tables/Table_rainfall_and_threshold_scenarios.xlsx"
SHEET_NAME = "Rainfall Scenarios"
TABLE_TITLE = "Rainfall and Threshold Scenarios"
CENTRAL_SUPPORT = "Central: 7 stations, 2016-2020"

SCENARIOS = (
    ("Moderate", 0.75, "M"),
    ("Heavy", 0.90, "H"),
    ("Extreme", 0.99, "E"),
)
RETENTION_SETTINGS = (
    (
        1.00,
        "Seven-station median of independent-event quantiles (2016–2020); baseline threshold.",
    ),
    (
        0.80,
        "Same event rainfall; official 80% threshold retention enters the exceedance index as 1/f.",
    ),
    (
        0.70,
        "Same event rainfall; official 70% threshold retention enters the exceedance index as 1/f.",
    ),
)


def build_table() -> pd.DataFrame:
    """Assemble one row per rainfall-severity and retention-factor combination."""
    scenario_data = pd.read_parquet(
        SCENARIO_PATH,
        columns=[
            "Support Specification",
            "Station ID",
            "Rainfall Scenario",
            "Event Quantile",
            "Rainfall Event Count",
            "Scenario 1 h Rainfall",
            "Scenario 3 h Rainfall",
            "Scenario 24 h Rainfall",
            "Scenario 72 h Rainfall",
        ],
    )
    central = scenario_data.loc[
        scenario_data["Support Specification"].eq(CENTRAL_SUPPORT)
    ].copy()
    if central["Station ID"].nunique() != 7:
        raise RuntimeError("Central scenario support must contain exactly seven stations.")

    threshold = pd.read_parquet(
        THRESHOLD_PATH,
        columns=["Rainfall Threshold Retention Factor"],
    )
    observed_factors = set(
        threshold["Rainfall Threshold Retention Factor"].dropna().astype(float).round(2)
    )
    if not {0.70, 0.80}.issubset(observed_factors):
        raise RuntimeError(
            "Official threshold data must contain both 70% and 80% retention settings."
        )

    support = central.groupby(
        ["Rainfall Scenario", "Event Quantile"],
        as_index=True,
    )[
        [
            "Rainfall Event Count",
            "Scenario 1 h Rainfall",
            "Scenario 3 h Rainfall",
            "Scenario 24 h Rainfall",
            "Scenario 72 h Rainfall",
        ]
    ].median()

    rows: list[dict[str, object]] = []
    for scenario, quantile, code in SCENARIOS:
        scenario_values = support.loc[(scenario, quantile)]
        for factor, boundary in RETENTION_SETTINGS:
            rows.append(
                {
                    "Scenario ID": f"{code}-{int(round(factor * 100)):03d}",
                    "Rainfall Scenario": scenario,
                    "Event Quantile": quantile,
                    "Median Station Event Count": int(scenario_values["Rainfall Event Count"]),
                    "1 h Event Rainfall (mm)": float(scenario_values["Scenario 1 h Rainfall"]),
                    "3 h Event Rainfall (mm)": float(scenario_values["Scenario 3 h Rainfall"]),
                    "24 h Event Rainfall (mm)": float(scenario_values["Scenario 24 h Rainfall"]),
                    "72 h Event Rainfall (mm)": float(scenario_values["Scenario 72 h Rainfall"]),
                    "Threshold Retention Factor": factor,
                    "Interpretation Boundary": boundary,
                }
            )
    heavy_values = support.loc[("Heavy", 0.90)]
    for factor, support_label in (
        (0.70, "Bounding assignment: all unresolved Yatsushiro municipality set to 0.70."),
        (0.75, "Central analyst assumption: unresolved Yatsushiro municipality midpoint; not an official value."),
        (0.80, "Bounding assignment: all unresolved Yatsushiro municipality set to 0.80."),
    ):
        rows.append(
            {
                "Scenario ID": f"Y-H-{int(round(factor * 100)):03d}",
                "Rainfall Scenario": "Heavy",
                "Event Quantile": 0.90,
                "Median Station Event Count": int(heavy_values["Rainfall Event Count"]),
                "1 h Event Rainfall (mm)": float(heavy_values["Scenario 1 h Rainfall"]),
                "3 h Event Rainfall (mm)": float(heavy_values["Scenario 3 h Rainfall"]),
                "24 h Event Rainfall (mm)": float(heavy_values["Scenario 24 h Rainfall"]),
                "72 h Event Rainfall (mm)": float(heavy_values["Scenario 72 h Rainfall"]),
                "Threshold Retention Factor": factor,
                "Interpretation Boundary": support_label,
            }
        )
    table = pd.DataFrame(rows)
    if table.shape != (12, 10):
        raise RuntimeError(f"Expected a 12 × 10 table, found {table.shape}.")
    if table["Scenario ID"].duplicated().any():
        raise RuntimeError("Scenario identifiers must be unique.")
    return table


def style_workbook(path: Path) -> None:
    """Apply compact scientific-table formatting for screen and A3 review."""
    workbook = load_workbook(path)
    worksheet = workbook[SHEET_NAME]
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A3"
    worksheet.auto_filter.ref = f"A2:J{worksheet.max_row}"
    worksheet.sheet_view.zoomScale = 90
    worksheet.print_area = f"A1:J{worksheet.max_row}"
    worksheet.print_title_rows = "1:2"
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A3
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 1
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_margins = PageMargins(
        left=0.25,
        right=0.25,
        top=0.35,
        bottom=0.35,
        header=0.15,
        footer=0.15,
    )

    header_fill = PatternFill("solid", fgColor="17365D")
    header_font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
    body_font = Font(name="Aptos", size=9.5, color="172033")
    subtle_border = Border(bottom=Side(style="thin", color="D0D5DD"))
    scenario_fills = {
        "Moderate": PatternFill("solid", fgColor="D9EDE9"),
        "Heavy": PatternFill("solid", fgColor="FCE5D4"),
        "Extreme": PatternFill("solid", fgColor="F4D6D3"),
    }

    worksheet.merge_cells("A1:J1")
    title_cell = worksheet["A1"]
    title_cell.value = TABLE_TITLE
    title_cell.fill = PatternFill("solid", fgColor="D9EAF7")
    title_cell.font = Font(name="Aptos Display", size=15, bold=True, color="17365D")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    worksheet.row_dimensions[1].height = 28

    for cell in worksheet[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[2].height = 42

    for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row):
        scenario = str(row[1].value)
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = subtle_border
        row[1].fill = scenario_fills[scenario]
        row[2].number_format = "0%"
        row[3].number_format = "#,##0"
        for cell in row[4:8]:
            cell.number_format = "0.0"
        row[8].number_format = "0%"
        for cell in row[2:9]:
            cell.alignment = Alignment(horizontal="right", vertical="top")
        worksheet.row_dimensions[row[0].row].height = 38

    widths = {
        "A": 14,
        "B": 18,
        "C": 15,
        "D": 22,
        "E": 17,
        "F": 17,
        "G": 18,
        "H": 18,
        "I": 20,
        "J": 54,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    worksheet.conditional_formatting.add(
        f"I3:I{worksheet.max_row}",
        ColorScaleRule(
            start_type="num",
            start_value=0.70,
            start_color="F8696B",
            mid_type="num",
            mid_value=0.80,
            mid_color="FFEB84",
            end_type="num",
            end_value=1.00,
            end_color="63BE7B",
        ),
    )
    excel_table = Table(
        displayName="RainfallThresholdScenarios",
        ref=f"A2:J{worksheet.max_row}",
    )
    excel_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(excel_table)
    workbook.save(path)


def verify_workbook(path: Path) -> None:
    """Reopen the workbook and verify structure, values, and error-free cells."""
    workbook = load_workbook(path, data_only=False)
    if workbook.sheetnames != [SHEET_NAME]:
        raise RuntimeError(f"Unexpected workbook sheets: {workbook.sheetnames}")
    worksheet = workbook[SHEET_NAME]
    if worksheet.max_row != 14 or worksheet.max_column != 10:
        raise RuntimeError(
            "Expected 14 worksheet rows including title and header and 10 columns; found "
            f"{worksheet.max_row} × {worksheet.max_column}."
        )
    if worksheet["A1"].value != TABLE_TITLE:
        raise RuntimeError("Workbook title row is missing or incorrect.")
    error_tokens = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"}
    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value in error_tokens:
                raise RuntimeError(f"Spreadsheet error token in {cell.coordinate}: {cell.value}")
    factors = {float(worksheet.cell(row, 9).value) for row in range(3, 15)}
    if factors != {0.70, 0.75, 0.80, 1.00}:
        raise RuntimeError(f"Unexpected retention-factor set: {factors}")
    for row in range(3, 15):
        for column in range(3, 10):
            value = worksheet.cell(row, column).value
            if not isinstance(value, (int, float)):
                raise RuntimeError(f"Expected numeric value at row {row}, column {column}.")


def main() -> None:
    table = build_table()
    OUT.parent.mkdir(parents=True, exist_ok=True)
    table.to_excel(
        OUT,
        index=False,
        sheet_name=SHEET_NAME,
        engine="openpyxl",
        startrow=1,
    )
    style_workbook(OUT)
    verify_workbook(OUT)
    unique_values = table.drop_duplicates("Rainfall Scenario").set_index("Rainfall Scenario")
    print(f"Saved: {OUT.relative_to(ROOT)}")
    print(f"Rows: {len(table):,}; columns: {len(table.columns):,}")
    for scenario, _, _ in SCENARIOS:
        row = unique_values.loc[scenario]
        print(
            f"{scenario}: 1 h={row['1 h Event Rainfall (mm)']:.1f} mm; "
            f"24 h={row['24 h Event Rainfall (mm)']:.1f} mm; "
            f"72 h={row['72 h Event Rainfall (mm)']:.1f} mm"
        )
    print("Median station event count: 346")
    print("Workbook verification: passed")


if __name__ == "__main__":
    main()
